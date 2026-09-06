#!/usr/bin/env python3
"""
gcd_build_local_db.py — extract a Marshall-Comics-scoped subset of the GCD
(Grand Comics Database) public dump into a small local SQLite file.

Why this shape: GCD's public REST API has no working search/filter
(`?name=`/`?search=` on /api/series/ are silently ignored — confirmed by
direct request, not assumed) — the only complete access path is the bi-weekly
MySQL dump. This is a ONE-TIME streaming extraction, not a live query: scan
the ~3.8GB dump for gcd_series rows whose name matches a title actually in
our inventory, then cascade to gcd_issue and gcd_story rows that hang off
those matched series only — everything else (231k series, most of it
irrelevant) is discarded during the scan, never held in memory or written
to the output DB.

Column orders below were extracted directly from the dump's own CREATE TABLE
statements at verified byte offsets (not assumed from a possibly-truncated
terminal view — one table's schema silently dropped its `id` column in a
naive first pass; ground-truth byte-offset extraction caught it).

Usage:
    python3 gcd_build_local_db.py /path/to/2026-07-15.sql
    python3 gcd_build_local_db.py /path/to/2026-07-15.sql --stage series   # test one stage
"""
import argparse, glob, json, os, re, sqlite3, subprocess, sys, time

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
OUT_DB = os.path.join(ROOT, "gcd_local.sqlite")

# ── Verified column orders (byte-offset extracted from the dump's own DDL) ──
COLUMNS = {
    "gcd_series": ["id","name","sort_name","format","year_began","year_began_uncertain",
        "year_ended","year_ended_uncertain","publication_dates","first_issue_id",
        "last_issue_id","is_current","publisher_id","country_id","language_id",
        "tracking_notes","notes","has_gallery","issue_count","created","modified",
        "deleted","has_indicia_frequency","has_isbn","has_barcode","has_issue_title",
        "has_volume","is_comics_publication","color","dimensions","paper_stock",
        "binding","publishing_format","has_rating","publication_type_id",
        "is_singleton","has_about_comics","has_indicia_printer","has_publisher_code_number"],
    "gcd_publisher": ["id","name","country_id","year_began","year_ended","notes","url",
        "brand_count","indicia_publisher_count","series_count","created","modified",
        "issue_count","deleted","year_began_uncertain","year_ended_uncertain",
        "year_overall_began","year_overall_began_uncertain","year_overall_ended",
        "year_overall_ended_uncertain"],
    "gcd_issue": ["id","number","volume","no_volume","display_volume_with_number",
        "series_id","indicia_publisher_id","indicia_pub_not_printed","brand_id",
        "no_brand","publication_date","key_date","sort_code","price","page_count",
        "page_count_uncertain","indicia_frequency","no_indicia_frequency","editing",
        "no_editing","notes","created","modified","deleted","is_indexed","isbn",
        "valid_isbn","no_isbn","variant_of_id","variant_name","barcode","no_barcode",
        "title","no_title","on_sale_date","on_sale_date_uncertain","rating",
        "no_rating","volume_not_printed","indicia_printer_not_printed",
        "variant_cover_status","indicia_printer_sourced_by"],
    "gcd_story": ["id","title","title_inferred","feature","sequence_number",
        "page_count","issue_id","script","pencils","inks","colors","letters",
        "editing","genre","characters","synopsis","reprint_notes","created",
        "modified","notes","no_script","no_pencils","no_inks","no_colors",
        "no_letters","no_editing","page_count_uncertain","type_id","job_number",
        "deleted","first_line"],
    "gcd_story_credit": ["id","created","modified","deleted","is_credited",
        "is_signed","uncertain","signed_as","credited_as","credit_name",
        "creator_id","credit_type_id","story_id","signature_id","is_sourced",
        "sourced_by"],
    "gcd_creator_name_detail": ["id","name","created","modified","deleted",
        "creator_id","type_id","sort_name","is_official_name","in_script_id",
        "family_name","given_name"],
    "gcd_credit_type": ["id","name","sort_code"],
}

# ── Streaming mysqldump tuple parser (verified against escape-sequence edge
#    cases: \' \\ \n NULL, and doubled '' quoting) ──────────────────────────
def iter_tuples(fh, chunk_size=4 * 1024 * 1024):
    """Read text chunks from an already-positioned file handle (positioned
    right after 'VALUES ') and yield each (v1,v2,...) row as a list of raw
    (still-quoted) field strings. Stops at the statement-terminating ';'."""
    depth = 0
    in_str = False
    esc = False
    field = []
    row = []
    started = False
    while True:
        chunk = fh.read(chunk_size)
        if not chunk:
            return
        i, n = 0, len(chunk)
        while i < n:
            c = chunk[i]
            if esc:
                field.append(c); esc = False; i += 1; continue
            if in_str:
                if c == '\\':
                    field.append(c); esc = True
                elif c == "'":
                    if i + 1 < n and chunk[i+1] == "'":
                        field.append("'"); i += 1
                    else:
                        in_str = False; field.append(c)
                else:
                    field.append(c)
                i += 1; continue
            if c == "'":
                in_str = True; field.append(c); i += 1; continue
            if c == '(':
                depth += 1
                if depth == 1:
                    started = True; field = []; row = []
                else:
                    field.append(c)
                i += 1; continue
            if c == ')':
                depth -= 1
                if depth == 0 and started:
                    row.append(''.join(field))
                    yield row
                    field = []; row = []; started = False
                else:
                    field.append(c)
                i += 1; continue
            if depth == 1 and c == ',':
                row.append(''.join(field)); field = []; i += 1; continue
            if depth == 0 and c == ';':
                return
            if depth >= 1:
                field.append(c)
            i += 1


_ESC = {'n': '\n', 'r': '\r', 't': '\t', '0': '\0', 'Z': '\x1a', 'b': '\b'}
def unquote(v):
    if v == 'NULL':
        return None
    if not (v.startswith("'") and v.endswith("'")):
        return v
    v = v[1:-1]
    out = []
    i = 0
    while i < len(v):
        c = v[i]
        if c == '\\' and i + 1 < len(v):
            out.append(_ESC.get(v[i+1], v[i+1])); i += 2
        else:
            out.append(c); i += 1
    return ''.join(out)


def insert_offsets(dump_path, table):
    """All byte offsets of `INSERT INTO \`table\` VALUES` in the dump, via grep
    (fast C scan of a 3.8GB file; far faster than doing this in Python)."""
    pattern = f"^INSERT INTO `{table}` VALUES"
    out = subprocess.run(["grep", "-aboE", pattern, dump_path], capture_output=True, text=True, check=True).stdout
    offsets = []
    for line in out.splitlines():
        off_str, _ = line.split(":", 1)
        offsets.append(int(off_str))
    return offsets


def scan_table(dump_path, table, row_filter, progress_label=""):
    """Yield dict(colname->raw_value) for every row in `table` where
    row_filter(dict) returns True. row_filter receives UNQUOTED values."""
    cols = COLUMNS[table]
    offsets = insert_offsets(dump_path, table)
    prefix = f"INSERT INTO `{table}` VALUES "
    matched = 0
    t0 = time.time()
    with open(dump_path, "r", encoding="utf-8", errors="replace") as f:
        for bi, off in enumerate(offsets):
            f.seek(off + len(prefix))
            for row in iter_tuples(f):
                if len(row) != len(cols):
                    continue  # defensive: skip malformed rows rather than misalign
                d = {cols[i]: unquote(row[i]) for i in range(len(cols))}
                if row_filter(d):
                    matched += 1
                    yield d
            if progress_label:
                elapsed = time.time() - t0
                print(f"  [{progress_label}] block {bi+1}/{len(offsets)}  matched so far: {matched}  ({elapsed:.0f}s)", file=sys.stderr)


def normalize_title(t):
    return re.sub(r"[^a-z0-9]+", " ", str(t or "").lower()).strip()


def load_inventory_titles():
    files = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    xlsx = max(files, key=os.path.getmtime)
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    inv = next(n for n in wb.sheetnames if n.startswith("✅ Clean Inventory"))
    ws = wb[inv]
    rows = list(ws.iter_rows(values_only=True))
    H = list(rows[0])
    ti = H.index("Title")
    titles = {str(r[ti]).strip() for r in rows[1:] if r[ti]}
    print(f"Loaded {len(titles)} distinct titles from {os.path.basename(xlsx)}", file=sys.stderr)
    return titles, xlsx


def init_db(path):
    if os.path.exists(path):
        os.remove(path)
    conn = sqlite3.connect(path)
    conn.executescript("""
        CREATE TABLE gcd_series (
            id INTEGER PRIMARY KEY, name TEXT, year_began INTEGER, year_ended INTEGER,
            publisher_id INTEGER, issue_count INTEGER, matched_title TEXT
        );
        CREATE TABLE gcd_publisher (id INTEGER PRIMARY KEY, name TEXT);
        CREATE TABLE gcd_issue (
            id INTEGER PRIMARY KEY, series_id INTEGER, number TEXT, key_date TEXT,
            publication_date TEXT
        );
        CREATE TABLE gcd_story (
            id INTEGER PRIMARY KEY, issue_id INTEGER, sequence_number INTEGER,
            script TEXT, pencils TEXT, type_id INTEGER
        );
        CREATE INDEX idx_series_name ON gcd_series(matched_title);
        CREATE INDEX idx_issue_series ON gcd_issue(series_id);
        CREATE INDEX idx_story_issue ON gcd_story(issue_id);
    """)
    return conn


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("dump")
    ap.add_argument("--stage", choices=["series", "issue", "story", "all"], default="all")
    ap.add_argument("--limit-titles", type=int, default=0, help="debug: only match first N titles")
    args = ap.parse_args()

    titles, xlsx_name = load_inventory_titles()
    if args.limit_titles:
        titles = set(sorted(titles)[:args.limit_titles])
    norm_titles = {normalize_title(t): t for t in titles}
    print(f"Normalized to {len(norm_titles)} unique title-keys", file=sys.stderr)

    conn = init_db(OUT_DB)

    # ── Stage 1: gcd_series ──────────────────────────────────────────────────
    def series_filter(d):
        return normalize_title(d["name"]) in norm_titles

    n = 0
    matched_series_ids = set()
    for d in scan_table(args.dump, "gcd_series", series_filter, progress_label="series"):
        matched_series_ids.add(d["id"])
        conn.execute(
            "INSERT OR IGNORE INTO gcd_series (id,name,year_began,year_ended,publisher_id,issue_count,matched_title) VALUES (?,?,?,?,?,?,?)",
            (d["id"], d["name"], d["year_began"], d["year_ended"], d["publisher_id"], d["issue_count"],
             norm_titles.get(normalize_title(d["name"]))),
        )
        n += 1
    conn.commit()
    print(f"Stage 1 (gcd_series): {n} matched series rows, {len(matched_series_ids)} distinct ids", file=sys.stderr)
    if args.stage == "series":
        return

    # ── Publishers (small table, keep all — cheap) ───────────────────────────
    def pub_filter(d):
        return True
    npub = 0
    for d in scan_table(args.dump, "gcd_publisher", pub_filter):
        conn.execute("INSERT OR IGNORE INTO gcd_publisher (id,name) VALUES (?,?)", (d["id"], d["name"]))
        npub += 1
    conn.commit()
    print(f"gcd_publisher: {npub} rows loaded", file=sys.stderr)

    # ── Stage 2: gcd_issue, filtered to matched series ───────────────────────
    def issue_filter(d):
        return d["series_id"] in matched_series_ids

    matched_issue_ids = set()
    n = 0
    for d in scan_table(args.dump, "gcd_issue", issue_filter, progress_label="issue"):
        matched_issue_ids.add(d["id"])
        conn.execute(
            "INSERT OR IGNORE INTO gcd_issue (id,series_id,number,key_date,publication_date) VALUES (?,?,?,?,?)",
            (d["id"], d["series_id"], d["number"], d["key_date"], d["publication_date"]),
        )
        n += 1
        if n % 500 == 0:
            conn.commit()
    conn.commit()
    print(f"Stage 2 (gcd_issue): {n} matched issue rows", file=sys.stderr)
    if args.stage == "issue":
        return

    # ── Stage 3: gcd_story, filtered to matched issues ───────────────────────
    def story_filter(d):
        return d["issue_id"] in matched_issue_ids

    n = 0
    for d in scan_table(args.dump, "gcd_story", story_filter, progress_label="story"):
        conn.execute(
            "INSERT OR IGNORE INTO gcd_story (id,issue_id,sequence_number,script,pencils,type_id) VALUES (?,?,?,?,?,?)",
            (d["id"], d["issue_id"], d["sequence_number"], d["script"], d["pencils"], d["type_id"]),
        )
        n += 1
        if n % 1000 == 0:
            conn.commit()
    conn.commit()
    print(f"Stage 3 (gcd_story): {n} matched story rows", file=sys.stderr)

    conn.close()
    print(f"\nDone. Local DB: {OUT_DB}", file=sys.stderr)


if __name__ == "__main__":
    main()
