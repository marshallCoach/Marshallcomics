#!/usr/bin/env python3
"""Apply the legitimate title/formatting corrections (item 1 + item 2). Skips
the CV mis-matches that are not formatting fixes (e.g. 'Chadwick Boseman: Rest
in Power' -> 'Rest'). Matched by expected current title for safety; clears the
'check title' note on each applied row. Non-destructive: flattens formulas,
writes a new timestamped canonical, only Title / Issue Note change."""
import openpyxl, glob, os, datetime

SRC = max((f for f in glob.glob("attached_assets/comics_inventory_*.xlsx")
           if " copy" not in f and not os.path.basename(f).startswith("~$")),
          key=os.path.getmtime)

# expected current title -> corrected (applied to every matching flagged row)
RENAMES = {
    # item 2: GCD + CV both give the base title
    "G.I. Joe: A Real American Hero — MIA":        "G.I. Joe: A Real American Hero",
    # item 1: genuine formatting/title corrections
    "Terrifics":                                    "The Terrifics",
    "Good Asian":                                   "The Good Asian",
    "Justice League: The Darkseid War - Shazam":    "Justice League: Darkseid War: Shazam",
    "Justice League: The Darkseid War: The Flash":  "Justice League: Darkseid War: Flash",
    "Icon and Rocket":                              "Icon & Rocket: Season One",
}

srcvals = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory"))
cached = [list(r) for r in srcvals.iter_rows(values_only=True)]
wb = openpyxl.load_workbook(SRC)
ws = next(w for w in wb.worksheets if w.title.startswith("✅ Clean Inventory"))
for ri, row in enumerate(ws.iter_rows()):
    for cj, c in enumerate(row):
        if isinstance(c.value, str) and c.value.startswith("=") and ri < len(cached) and cj < len(cached[ri]):
            c.value = cached[ri][cj]
H = [c.value for c in ws[1]]; C = {h: i for i, h in enumerate(H) if h}
TCOL, NCOL = C["Title"]+1, C["Issue Note"]+1
def sval(r, col):
    v = ws.cell(r, col).value
    return "" if v is None else str(v).strip()

applied = []
for r in range(2, ws.max_row + 1):
    t = sval(r, TCOL)
    if t in RENAMES:
        ws.cell(r, TCOL, RENAMES[t])
        ws.cell(r, NCOL).value = None            # correct clear idiom
        applied.append((r, t, RENAMES[t]))

out = f"attached_assets/comics_inventory_{datetime.datetime.now():%d%m_%H%M}.xlsx"
wb.save(out)
so = list(srcvals.iter_rows(values_only=True))
no = list(next(w for w in openpyxl.load_workbook(out, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory")).iter_rows(values_only=True))
allowed = {C["Title"], C["Issue Note"]}
off = sum(1 for a,b in zip(so,no) for j in range(len(a)) if j not in allowed and (a[j] or "")!=(b[j] or ""))
print(f"OUTPUT: {os.path.basename(out)}   renames applied: {len(applied)}")
for r,a,b in applied: print(f"   row {r}: {a!r} -> {b!r}")
print(f"CONTAMINATION (non Title/Note diffs): {off}  (must be 0)")
