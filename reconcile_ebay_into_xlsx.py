#!/usr/bin/env python3
"""
reconcile_ebay_into_xlsx.py — bake the newest eBay comps into the inventory
xlsx so the SPREADSHEET is the single source of truth (retiring the JSON
overlay the web app currently leans on).

Self-contained: needs only `openpyxl` + stdlib, no repo imports — so it runs
unchanged inside Claude chat's Python sandbox on two attached files.

Decision baked in (per the fetch-date evidence): the JSON is the NEWER fetch
(2026-07-08..07-15) and the xlsx eBay columns are OLDER (2026-07-04/06), so
JSON WINS on every conflict — this is a full refresh, not a fill-only pass.

Matching mirrors the web app's overlay exactly, so the reconciled xlsx will
agree with what the site already shows:
  key = Title + "|||" + normalized Issue   (case-sensitive Title)
  normalize issue: strip leading '#', collapse a trailing '.0'
                   ("656.0","656","#656" -> "656"; "1 of 5","3.1" untouched)
  write the same comps to EVERY row sharing that Title+Issue (multiple
  physical copies share one market price — correct).

Never writes the source in place; emits a new comics_inventory_DDMM_HHMM.xlsx.

Usage:
    python3 reconcile_ebay_into_xlsx.py <inventory.xlsx> <ebay_pricing_results.json>
    python3 reconcile_ebay_into_xlsx.py            # auto-detect in cwd / attached_assets
"""
import json, os, glob, sys, datetime

COLS = {  # xlsx column  ->  JSON field
    "eBay Median Sold $": "median",
    "eBay Avg Sold $":    "avg",
    "eBay Low $":         "low",
    "eBay High $":        "high",
    "eBay Comp Count":    "count",
}
FETCHED_COL = "eBay Fetched"


def norm_issue(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def find(patterns):
    for p in patterns:
        hits = [f for f in glob.glob(p) if not os.path.basename(f).startswith("~$")]
        if hits:
            return max(hits, key=os.path.getmtime)
    return None


def main():
    args = [a for a in sys.argv[1:]]
    xlsx = args[0] if len(args) > 0 else find(
        ["comics_inventory_*.xlsx", "attached_assets/comics_inventory_*.xlsx"])
    jpath = args[1] if len(args) > 1 else find(
        ["ebay_pricing_results.json", "attached_assets/ebay_pricing_results.json"])
    if not xlsx or not jpath:
        sys.exit("ERROR: need <inventory.xlsx> <ebay_pricing_results.json> (none auto-detected).")

    import openpyxl
    print(f"Inventory : {os.path.basename(xlsx)}")
    print(f"Pricing   : {os.path.basename(jpath)}")

    res = json.load(open(jpath))
    # newest-priced entry per normalized key (a title|issue can recur; keep the
    # freshest by fetched_at so a re-fetch supersedes an earlier one)
    jmap = {}
    for k, v in res.items():
        if "|||" not in k or not isinstance(v, dict) or v.get("avg") is None:
            continue
        t, iss = k.split("|||", 1)
        key = f"{t}|||{norm_issue(iss)}"
        prev = jmap.get(key)
        if prev is None or str(v.get("fetched_at", "")) >= str(prev.get("fetched_at", "")):
            jmap[key] = v
    print(f"Priced JSON entries (unique key): {len(jmap)}")

    wb = openpyxl.load_workbook(xlsx)
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("✅ Clean Inventory"))
    H = [c.value for c in next(ws.iter_rows(max_row=1))]
    ti, ii = H.index("Title"), H.index("Issue #")
    col_idx = {c: H.index(c) + 1 for c in COLS if c in H}
    missing = [c for c in COLS if c not in H]
    if missing:
        sys.exit(f"ERROR: xlsx missing expected columns: {missing}")
    fetched_idx = H.index(FETCHED_COL) + 1 if FETCHED_COL in H else None
    avg_idx = col_idx["eBay Avg Sold $"]

    refreshed = newly_filled = unchanged_blank = 0
    matched_keys = set()
    samples = []
    for row in ws.iter_rows(min_row=2):
        title = str(row[ti].value or "").strip()
        if not title:
            continue
        key = f"{title}|||{norm_issue(row[ii].value)}"
        eb = jmap.get(key)
        if not eb:
            continue
        matched_keys.add(key)
        old_avg = row[avg_idx - 1].value
        had = not (old_avg is None or str(old_avg).strip() in ("", "nan"))
        for col, field in COLS.items():
            row[col_idx[col] - 1].value = eb.get(field)
        if fetched_idx and eb.get("fetched_at"):
            row[fetched_idx - 1].value = str(eb["fetched_at"])[:10]
        if had:
            refreshed += 1
            if len(samples) < 5:
                samples.append((title, norm_issue(row[ii].value), old_avg, eb.get("avg")))
        else:
            newly_filled += 1

    # rows still blank after the pass (report only)
    for row in ws.iter_rows(min_row=2):
        v = row[avg_idx - 1].value
        if (v is None or str(v).strip() in ("", "nan")):
            unchanged_blank += 1

    total_rows = ws.max_row - 1
    print(f"\n{'='*58}\n  RECONCILE — JSON (newer) written over xlsx (older)\n{'='*58}")
    print(f"  Rows refreshed (had a price, updated to newer): {refreshed}")
    print(f"  Rows newly filled (were blank):                 {newly_filled}")
    print(f"  Rows still blank (no comps in JSON):            {unchanged_blank}")
    print(f"  Distinct JSON keys applied:                     {len(matched_keys)} / {len(jmap)}")
    print(f"\n  before/after sample (refreshed rows):")
    for t, i, o, n in samples:
        print(f"    {t} #{i}: {o} -> {n}")

    stamp = datetime.datetime.now().strftime("%d%m_%H%M")
    out = os.path.join(os.path.dirname(xlsx) or ".", f"comics_inventory_{stamp}.xlsx")
    wb.save(out)
    print(f"\n  Written: {out}   (rows: {total_rows})")


if __name__ == "__main__":
    main()
