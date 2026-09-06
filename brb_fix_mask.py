#!/usr/bin/env python3
"""Fix the Energon Universe M.A.S.K. #3 row (Box 103): it was mistitled "Mask"
and mispublished "Dark Horse". Sets Title -> "M.A.S.K." and Publisher ->
"Image/Skybound" to match the #2 row. Content-keyed on Title+Issue+Box so it
survives row-number shifts from the weekly reingest. Only Title and Publisher
change (contamination check). Year is left as-is and flagged for the human:
Energon M.A.S.K. #3 is ~2024, the row says 2025 — verify, don't guess."""
import openpyxl, glob, os, datetime

SRC = max((f for f in glob.glob("attached_assets/comics_inventory_*.xlsx")
           if " copy" not in f and not os.path.basename(f).startswith("~$")),
          key=os.path.getmtime)

# match target row by CONTENT (not row number)
def norm_iss(v): return str(v or "").strip().replace(".0", "")
MATCH = {"title": "Mask", "issue": "3", "box": "103"}
NEW_TITLE = "M.A.S.K."
NEW_PUB   = "Image/Skybound"

srcvals = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory"))
cached = [list(r) for r in srcvals.iter_rows(values_only=True)]

wb = openpyxl.load_workbook(SRC)
ws = next(w for w in wb.worksheets if w.title.startswith("✅ Clean Inventory"))
# flatten formulas so openpyxl doesn't null cached formula columns on save
for ri, row in enumerate(ws.iter_rows()):
    for cj, c in enumerate(row):
        if isinstance(c.value, str) and c.value.startswith("=") and ri < len(cached) and cj < len(cached[ri]):
            c.value = cached[ri][cj]

H = [c.value for c in ws[1]]; C = {h: i for i, h in enumerate(H) if h}
TCOL = C["Title"] + 1; PCOL = C["Publisher"] + 1
ICOL = C["Issue #"] + 1; BCOL = C["Box #"] + 1

applied = []
for r in range(2, ws.max_row + 1):
    t = str(ws.cell(r, TCOL).value or "").strip()
    i = norm_iss(ws.cell(r, ICOL).value)
    b = str(ws.cell(r, BCOL).value or "").strip()
    if t == MATCH["title"] and i == MATCH["issue"] and b == MATCH["box"]:
        ws.cell(r, TCOL, NEW_TITLE)
        ws.cell(r, PCOL, NEW_PUB)
        applied.append(r)

if not applied:
    print("NO MATCH — 'Mask' #3 in Box 103 not found (already fixed, or box changed). Nothing written.")
    raise SystemExit(0)

out = f"attached_assets/comics_inventory_{datetime.datetime.now():%d%m_%H%M}.xlsx"
wb.save(out)

# contamination check: only Title + Publisher may differ
so = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
          if w.title.startswith("✅ Clean Inventory"))
no = next(w for w in openpyxl.load_workbook(out, read_only=True, data_only=True).worksheets
          if w.title.startswith("✅ Clean Inventory"))
allowed = {C["Title"], C["Publisher"]}
off = sum(1 for a, b in zip(so.iter_rows(values_only=True), no.iter_rows(values_only=True))
          for j in range(len(a)) if j not in allowed and (a[j] or "") != (b[j] or ""))
print(f"SOURCE: {os.path.basename(SRC)}")
print(f"OUTPUT: {os.path.basename(out)}")
print(f"rows fixed: {len(applied)} (row {applied})  Title->'{NEW_TITLE}', Publisher->'{NEW_PUB}'")
print(f"contamination (non Title/Publisher cells changed): {off}")
print("REMINDER: verify Year (row says 2025; Energon M.A.S.K. #3 is ~2024) before it ships.")
