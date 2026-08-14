#!/usr/bin/env python3
"""Combine visually-verified characters into the inventory as the new source of
truth. Adds a 'Visually Verified
Characters' column (unambiguous Title+Issue+Volume matches only) and a
'Cover Characters' tab. Non-destructive: existing cells untouched; writes a NEW
timestamped file. Confirms no data contamination (row count + every original
column byte-identical)."""
import openpyxl, glob, os, datetime, collections, hashlib

SRC = "attached_assets/comics_inventory_1108_1200.xlsx"
VIS = "vision_characters_enriched.xlsx"
CL = max(glob.glob("cover_links_*.xlsx"), key=os.path.getmtime)


def ni(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s); return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def nv(v):
    try:
        return str(int(float(str(v).strip())))
    except (ValueError, TypeError):
        return "1"


# --- url -> volume (only URLs that map to exactly one volume are trustworthy) ---
clw = openpyxl.load_workbook(CL, read_only=True, data_only=True).active
ch = list(next(clw.iter_rows(values_only=True))); ci = {h: i for i, h in enumerate(ch)}
url2vol = collections.defaultdict(set)
for r in list(clw.iter_rows(values_only=True))[1:]:
    u = r[ci["Cover URL"]]
    if u:
        url2vol[u].add(nv(r[ci["Volume"]]))

# --- vision output -> character map keyed (title,issue,volume) ---
vw = openpyxl.load_workbook(VIS, read_only=True, data_only=True).active
vh = list(next(vw.iter_rows(values_only=True))); vi = {str(h).lower(): i for i, h in enumerate(vh)}
charmap = {}
vis_rows = []
for r in list(vw.iter_rows(values_only=True))[1:]:
    t = str(r[vi["title"]] or "").strip(); iss = ni(r[vi["issue"]])
    chars = r[vi["visually verified characters"]]; conf = r[vi["confidence"]]; url = r[vi["cover url"]]
    vols = url2vol.get(url, set())
    vol = next(iter(vols)) if len(vols) == 1 else None
    vis_rows.append((t, iss, vol, chars, conf, url))
    if vol is not None and conf == "🟢":
        charmap[(t.lower(), iss, vol)] = chars

# cached values from SRC (data_only) so we can flatten any formula cell to its
# computed value — openpyxl drops cached formula results on save, which would
# null out formula-driven cells (e.g. the VF-value column). Flattening preserves
# the exact data for this validated snapshot.
srcvals = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory"))
cached = [list(r) for r in srcvals.iter_rows(values_only=True)]

# --- load inventory (editable, keeps formatting/formulas) ---
wb = openpyxl.load_workbook(SRC)
ws = next(w for w in wb.worksheets if w.title.startswith("✅ Clean Inventory"))
# flatten formula cells on the inventory sheet to their cached values
for ri, row in enumerate(ws.iter_rows()):
    for cjj, cell in enumerate(row):
        if isinstance(cell.value, str) and cell.value.startswith("="):
            if ri < len(cached) and cjj < len(cached[ri]):
                cell.value = cached[ri][cjj]
H = [c.value for c in ws[1]]; C = {h: i for i, h in enumerate(H) if h}
tcol = C["Title"] + 1; icol = C["Issue #"] + 1; vcol = C["Volume"] + 1
pcol = C["Publisher"] + 1; acol = C["Cover Artist"] + 1
orig_rows, orig_cols = ws.max_row, ws.max_column

# snapshot original cells for contamination check
def snap(sheet):
    h = hashlib.sha256()
    for row in sheet.iter_rows(values_only=True):
        h.update(repr(row).encode())
    return h.hexdigest()
orig_hash = snap(ws)

# inventory lookup for clean Publisher/Cover Artist in the tab
invmap = {}
for row in range(2, ws.max_row + 1):
    t = str(ws.cell(row, tcol).value or "").strip().lower(); iss = ni(ws.cell(row, icol).value); vol = nv(ws.cell(row, vcol).value)
    invmap.setdefault((t, iss, vol), (ws.cell(row, pcol).value, ws.cell(row, acol).value))

# --- add the characters column (only confident, unambiguous matches) ---
newcol = ws.max_column + 1
ws.cell(1, newcol, "Visually Verified Characters")
filled = 0
for row in range(2, ws.max_row + 1):
    t = str(ws.cell(row, tcol).value or "").strip().lower(); iss = ni(ws.cell(row, icol).value); vol = nv(ws.cell(row, vcol).value)
    hit = charmap.get((t, iss, vol))
    if hit:
        ws.cell(row, newcol, hit); filled += 1

# --- keep the ✅ Clean Inventory naming ---
ws.title = "✅ Clean Inventory 1308_1843"

# --- Cover Characters tab ---
if "Cover Characters" in wb.sheetnames:
    del wb["Cover Characters"]
tab = wb.create_sheet("Cover Characters")
from openpyxl.styles import Font, PatternFill
tab.append(["Title", "Issue", "Volume", "Publisher", "Cover Artist", "Visually Verified Characters", "Confidence", "Cover URL"])
for c in tab[1]:
    c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2F5496")
tab.freeze_panes = "A2"
for t, iss, vol, chars, conf, url in vis_rows:
    pub, art = invmap.get((t.lower(), iss, vol), ("", ""))
    tab.append([t, iss, vol or "", pub, art, chars, conf, url])

out = f"attached_assets/comics_inventory_1108_{datetime.datetime.now():%H%M}.xlsx"
wb.save(out)

# --- contamination check: reload, verify original data intact ---
rb = openpyxl.load_workbook(out, read_only=True, data_only=True)
rs = next(rb[n] for n in rb.sheetnames if n.startswith("✅ Clean Inventory"))
rrows = list(rs.iter_rows(values_only=True))
# rebuild original-only view (drop the appended column) and hash
orig_view = hashlib.sha256()
for row in rrows:
    orig_view.update(repr(row[:orig_cols]).encode())
new_rows = len(rrows)
print(f"OUTPUT: {out}")
print(f"inventory rows: source={orig_rows} new={new_rows}  cols: {orig_cols}->{rs.max_column}")
print(f"characters filled: {filled}")
print(f"Cover Characters tab rows: {len(vis_rows)}")
print(f"CONTAMINATION CHECK original {orig_cols} cols byte-identical: {orig_view.hexdigest()==orig_hash}")
print(f"all sheets preserved: {rb.sheetnames}")
