#!/usr/bin/env python3
"""
brb_legacy_credits.py — fill blank Avengers credits using the Fandom legacy
numbering key, with a YEAR GATE.

The problem: many Avengers rows carry LEGACY issue numbers, so "#672" is not
Vol 7's 672nd issue in the naive sense and "#27" may be Vol 5 #27 rather than
legacy #27. The legacy key is at:
    https://marvel.fandom.com/wiki/Avengers_Vol_1/Legacy_Numbers
mined into avengers_legacy_map.json (713 entries, legacy# -> Vol/issue page).

WHY THIS IS NOT A BLIND LOOKUP — the map alone is wrong for modern volumes:
    "The Avengers #27, declared vol 5, year 2012"
       legacy map  -> Avengers Vol 1 27  (1966, Stan Lee)      WRONG era
       per-volume  -> Avengers Vol 5 27  (2014, Hickman)       right
So for every row we build BOTH candidates, fetch each, and keep only the one
whose Fandom release year corroborates the row's own declared Year (+/- 2).
No corroboration -> write nothing. This is the same in-range gate that made the
earlier Fandom year fixes safe (see brb_fandom_years.py CALIBRATION note).

Fandom's API is public and free — it does NOT touch the Comic Vine quota, so
this is safe to run alongside a CV job.

Only writes cells that are currently BLANK. Never writes the source file.

Usage:
    python3 brb_legacy_credits.py            # dry-run, no writes
    python3 brb_legacy_credits.py --apply
"""
import argparse, glob, json, os, re, time, urllib.parse, urllib.request
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
MAP = os.path.join(ROOT, "legacy_maps.json")
API = "https://marvel.fandom.com/api.php"
UA = "MarshallComicsInventory/1.0"
DELAY = 0.4
YEAR_TOL = 2
VERBOSE = True


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx"))
         if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def blank(v):
    return v is None or str(v).strip() in ("", "nan", "None")


def norm_issue(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def year_of(v):
    n = [int(x) for x in re.findall(r"\d{4}", str(v or "")) if 1900 < int(x) < 2100]
    return n[0] if n else None


_canon = {}


def canon(title):
    return _canon.get(re.sub(r"[^a-z0-9]", "", re.sub(r"^the ", "", str(title).strip().lower())), "")


_cache = {}


def fandom_issue(page):
    """Return {'year', 'writer', 'artist', 'cover'} for a Fandom issue page."""
    if page in _cache:
        return _cache[page]
    out = None
    try:
        u = (f"{API}?action=parse&page={urllib.parse.quote(page)}"
             "&prop=wikitext&format=json&formatversion=2&redirects=1")
        req = urllib.request.Request(u, headers={"User-Agent": UA})
        with urllib.request.urlopen(req, timeout=30) as r:
            d = json.load(r)
        wt = d.get("parse", {}).get("wikitext", "")
        if wt:
            def field(*names):
                for n in names:
                    m = re.search(r"\|\s*" + n + r"\s*=\s*([^\n|]+)", wt)
                    if m:
                        v = re.sub(r"\[\[|\]\]", "", m.group(1)).split("|")[-1].strip()
                        if v:
                            return v
                return None
            rd = field("ReleaseDate", "CoverDate", "Month", "Year")
            out = {"year": year_of(rd) or year_of(wt[:4000]),
                   "writer": field(r"Writer1_1", r"Writer_1", r"Writer1"),
                   "artist": field(r"Penciler1_1", r"Penciler_1", r"Artist1_1", r"Penciler1"),
                   "cover": field(r"CoverArtist1_1", r"CoverArtist_1", r"CoverArtist1")}
    except Exception as e:
        print(f"    [ERR] {page}: {type(e).__name__}: {e}", flush=True)
    _cache[page] = out
    time.sleep(DELAY)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=None)
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--verbose", action="store_true")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()
    global VERBOSE
    VERBOSE = args.verbose

    legacy = json.load(open(MAP))
    global _canon
    _canon = {re.sub(r"[^a-z0-9]", "", re.sub(r"^the ", "", k.lower())): k for k in legacy}
    xlsx = os.path.abspath(args.xlsx) if args.xlsx else latest_xlsx()
    print(f"Source: {os.path.basename(xlsx)}   legacy map: {len(legacy)} entries   "
          f"(mode: {'APPLY' if args.apply else 'DRY-RUN'})")

    wb = openpyxl.load_workbook(xlsx)
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("✅ Clean Inventory"))
    H = [c.value for c in next(ws.iter_rows(max_row=1))]
    C = {n: H.index(n) for n in H if n}
    ti, ii, yi, vi = C["Title"], C["Issue #"], C["Year"], C["Volume"]
    wi, ai, ci = C["Writer(s)"], C["Artist(s)"], C["Cover Artist"]

    targets = []
    for row in ws.iter_rows(min_row=2):
        title = str(row[ti].value or "").strip()
        if canon(title) not in legacy:
            continue
        if blank(row[wi].value) or blank(row[ai].value) or blank(row[ci].value):
            targets.append(row)
    if args.limit:
        targets = targets[: args.limit]
    print(f"  rows needing credits (legacy-keyed titles): {len(targets)}")

    fw = fa = fc = 0
    used_legacy = used_declared = rejected = 0
    for row in targets:
        iss = norm_issue(row[ii].value)
        yr = year_of(row[yi].value)
        vol = norm_issue(row[vi].value)
        cands = []
        base = canon(title)
        tmap = legacy.get(base, {})
        if iss in tmap:
            cands.append(("legacy", tmap[iss]["page"]))
        if vol and vol.isdigit():
            cands.append(("declared", f"{base} Vol {vol} {iss}"))
        pick = pickkind = None
        for kind, page in cands:
            info = fandom_issue(page)
            if not info or not info.get("year") or not yr:
                continue
            if abs(info["year"] - yr) <= YEAR_TOL:      # the gate
                pick, pickkind = info, kind
                break
        if not pick:
            rejected += 1
            if VERBOSE:
                print(f"    {title[:22]:<23}#{iss:<5} yr={yr} vol={vol} -> no year support", flush=True)
            continue
        used_legacy += pickkind == "legacy"
        used_declared += pickkind == "declared"
        if blank(row[wi].value) and pick["writer"]:
            row[wi].value = pick["writer"]; fw += 1
        if blank(row[ai].value) and pick["artist"]:
            row[ai].value = pick["artist"]; fa += 1
        if blank(row[ci].value) and pick["cover"]:
            row[ci].value = pick["cover"]; fc += 1
        d = pick["year"] - yr
        print(f"    {title[:22]:<23}#{iss:<5} yr={yr} -> {pickkind:<8} "
              f"fandom={pick['year']}({d:+d})  W={pick['writer']} A={pick['artist']}", flush=True)

    print(f"\n  matched via LEGACY number : {used_legacy}")
    print(f"  matched via DECLARED volume: {used_declared}")
    print(f"  rejected (no year support) : {rejected}")
    print(f"  Writers +{fw}   Artists +{fa}   Cover Artists +{fc}")
    if args.apply:
        out = args.out or xlsx.replace(".xlsx", "_LEGACY.xlsx")
        wb.save(out)
        print(f"  Written: {out}")
    else:
        print("  [DRY-RUN] nothing written — re-run with --apply")


if __name__ == "__main__":
    main()

