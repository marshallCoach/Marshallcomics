#!/usr/bin/env python3
"""
brb.py — one command that runs the whole Marshall Comics local pipeline.

Instead of calling gen_data / gen_quest / validators individually, this
conductor runs them in the right order, from the single source of truth
(newest comics_inventory*.xlsx in attached_assets/), with the tripwire safety
rules baked in.

    python3 brb.py                 # the safe local pipeline (default)
    python3 brb.py --check         # read-only: detect + validate + fill-rates, NO writes
    python3 brb.py --skip-reingest # refresh quest data + reports, but don't rebuild data3.ts
    python3 brb.py --covers        # + re-fetch null covers        (needs COMIC_VINE_API_KEY)
    python3 brb.py --ebay          # + eBay pricing pass           (needs EBAY_APP_ID/CERT_ID)
    python3 brb.py --writers       # + overnight writer/artist fill, launched in background
    python3 brb.py --commit "msg"  # + git add generated files, commit, push (gated)
    python3 brb.py --yes           # don't pause at the pre-flight confirmation

Default pipeline (each step gates the next):
    1. detect     newest xlsx in attached_assets/ + row count + sheet
    2. validate   brb_validate.py — HARD-STOP before any write only if a
                  STRUCTURAL check fails (columns / blank title / blank box /
                  box-integer / issue-blank). Dupes, clones, capacity and year
                  are known-accepted todos and do NOT stop the run.
    3. reingest   node gen_data.mjs      -> data3.ts (+ copies covers.json)
    4. gen-quest  node gen_quest.mjs     -> box-quest.html (consolidation planner)
    5. quest-data node gen_quest_data.mjs-> quest-data.js (snapshot/guide/board)
    6. fill-rates brb_fill_rates.py      (report)

Credentials are never handled here — credentialed phases only CHECK that the
env var is set and tell you what to export if it isn't; you run those yourself.
"""
import argparse, json, os, re, subprocess, sys, glob

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
STATE_FILE = os.path.join(ROOT, ".brb_pipeline_state.json")

# Validator checks that must PASS before we write anything downstream. Dupes,
# clones, capacity and year are accepted/known states, not blockers.
HARD_STOP_CHECKS = [
    "Required columns",
    "Blank Title",
    "Blank Box #",
    "Box # values are positive integers",
    "Issue # blank rate",
]
# Row-count tripwires (vs. the previous recorded run of the same pipeline).
ROW_DROP_LIMIT = 50      # rows shouldn't fall by more than this without a purge
ROW_GROW_LIMIT = 1000    # a jump this large is almost certainly the wrong file

C = {"g": "\033[32m", "r": "\033[31m", "y": "\033[33m", "b": "\033[1m", "d": "\033[2m", "x": "\033[0m"}
def color(s, k): return f"{C[k]}{s}{C['x']}" if sys.stdout.isatty() else s
def banner(s): print("\n" + color("═" * 64, "d") + f"\n  {color(s, 'b')}\n" + color("═" * 64, "d"))
def ok(s):   print(color("  ✓ " + s, "g"))
def warn(s): print(color("  ⚠ " + s, "y"))
def err(s):  print(color("  ✗ " + s, "r"))


def run(cmd, label, cwd=ROOT, env=None):
    """Run a subprocess, stream nothing-fancy, return (exit_code, stdout)."""
    print(color(f"\n$ {' '.join(cmd)}", "d"))
    proc = subprocess.run(cmd, cwd=cwd, env=env, capture_output=True, text=True)
    if proc.stdout: print(proc.stdout.rstrip())
    if proc.stderr.strip(): print(proc.stderr.rstrip())
    return proc.returncode, proc.stdout + proc.stderr


def detect_xlsx():
    files = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx"))
             if not os.path.basename(f).startswith("~$")]
    if not files:
        err(f"No comics_inventory_*.xlsx in {ASSETS}"); sys.exit(1)
    newest = max(files, key=os.path.getmtime)
    return newest


def xlsx_rowcount(path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        for name in wb.sheetnames:
            if name.startswith("✅ Clean Inventory"):
                ws = wb[name]
                return ws.max_row - 1, name  # minus header
        ws = wb[wb.sheetnames[0]]
        return ws.max_row - 1, wb.sheetnames[0]
    except Exception as e:
        warn(f"Could not read row count ({e})")
        return None, None


def parse_validator(output):
    """Return (hard_failures, soft_failures) — lists of check names that hit ✗."""
    blocks = re.split(r"CHECK\s+\w+\s+—\s+", output)
    hard, soft = [], []
    for blk in blocks[1:]:
        name = blk.splitlines()[0].strip()
        failed = "✗" in blk.split("CHECK")[0]
        if not failed:
            continue
        if any(h in name for h in HARD_STOP_CHECKS):
            hard.append(name)
        else:
            soft.append(name)
    return hard, soft


def load_state():
    try:
        with open(STATE_FILE) as f: return json.load(f)
    except Exception:
        return {}


def save_state(file, rows):
    try:
        with open(STATE_FILE, "w") as f:
            json.dump({"file": os.path.basename(file), "rows": rows}, f, indent=2)
    except Exception:
        pass


def credentialed_phase(name, env_var, cmd, background=False, hint=""):
    if not os.environ.get(env_var):
        warn(f"Skipping {name}: {env_var} is not set in this shell.")
        print(color(f"    Run it yourself after: export {env_var}=...  {hint}", "d"))
        return
    banner(f"{name}")
    if background:
        log = os.path.join(ROOT, f"{name.replace(' ', '_').lower()}.log")
        with open(log, "a") as lf:
            p = subprocess.Popen(cmd, cwd=ROOT, stdout=lf, stderr=lf)
        ok(f"Launched in background (PID {p.pid}). Log: {os.path.relpath(log, ROOT)}")
        print(color(f"    Watch it: tail -f {os.path.relpath(log, ROOT)}", "d"))
    else:
        run(cmd, name)


def main():
    ap = argparse.ArgumentParser(add_help=True, description="Marshall Comics pipeline conductor")
    ap.add_argument("--check", "--dry-run", dest="check", action="store_true",
                    help="Read-only: detect + validate + fill-rates. No writes.")
    ap.add_argument("--skip-reingest", action="store_true", help="Skip the data3.ts rebuild.")
    ap.add_argument("--covers", action="store_true", help="Also re-fetch null covers (needs COMIC_VINE_API_KEY).")
    ap.add_argument("--ebay", action="store_true", help="Also run eBay pricing (needs EBAY_APP_ID/CERT_ID).")
    ap.add_argument("--writers", action="store_true", help="Also launch the overnight writer/artist fill (background).")
    ap.add_argument("--commit", metavar="MSG", help="git add generated files, commit with MSG, and push.")
    ap.add_argument("--yes", action="store_true", help="Don't pause at the pre-flight confirmation.")
    args = ap.parse_args()

    # ── 1. DETECT ────────────────────────────────────────────────────────────
    banner("1 · DETECT — source of truth")
    xlsx = detect_xlsx()
    rows, sheet = xlsx_rowcount(xlsx)
    print(f"  File  : {color(os.path.basename(xlsx), 'b')}")
    print(f"  Sheet : {sheet}")
    print(f"  Rows  : {rows:,}" if rows is not None else "  Rows  : (unreadable)")

    state = load_state()
    if state.get("rows") and rows is not None:
        delta = rows - state["rows"]
        if delta:
            same = state.get("file") == os.path.basename(xlsx)
            msg = f"Row count {'unchanged file' if same else 'new file'}: {state['rows']:,} → {rows:,} ({'+' if delta>=0 else ''}{delta})"
            if delta < -ROW_DROP_LIMIT or delta > ROW_GROW_LIMIT:
                err("TRIPWIRE — " + msg)
                if not args.yes and not args.check:
                    err("Unexpected row-count change. Re-run with --yes if this is intended, or --check to inspect first.")
                    sys.exit(2)
            else:
                warn(msg)

    # Pre-flight summary
    phases = ["validate", "fill-rates"] if args.check else \
        (["validate"] + ([] if args.skip_reingest else ["reingest"]) +
         ["gen-quest", "quest-data", "fill-rates"])
    print(f"\n  Mode  : {color('READ-ONLY (--check)', 'y') if args.check else color('WRITE', 'b')}")
    print(f"  Steps : {' → '.join(phases)}")
    if not args.yes and not args.check:
        try:
            resp = input(color("\n  Proceed? [y/N] ", "b"))
        except EOFError:
            resp = "n"
        if resp.strip().lower() not in ("y", "yes"):
            print("  Aborted."); sys.exit(0)

    # ── 2. VALIDATE ──────────────────────────────────────────────────────────
    banner("2 · VALIDATE")
    code, out = run(["python3", "brb_validate.py", xlsx], "validate")
    hard, soft = parse_validator(out)
    if hard:
        err("Structural check(s) failed — refusing to reingest:")
        for h in hard: err("   " + h)
        err("Fix the data first. Nothing was written.")
        sys.exit(3)
    if soft:
        warn("Known/accepted checks failing (not blockers): " + "; ".join(soft))
    else:
        ok("All checks clean.")

    if args.check:
        banner("FILL RATES (read-only)")
        run(["python3", "brb_fill_rates.py", xlsx], "fill-rates")
        save_state(xlsx, rows)
        banner("DONE — read-only, nothing written")
        return

    # ── 3. REINGEST ──────────────────────────────────────────────────────────
    if not args.skip_reingest:
        banner("3 · REINGEST — node gen_data.mjs")
        code, out = run(["node", "gen_data.mjs"], "reingest")
        if code != 0:
            err("Reingest failed (see above). Stopping before derived files diverge.")
            sys.exit(4)

    # ── 4. BOX-QUEST consolidation planner ───────────────────────────────────
    banner("4 · GEN-QUEST — node gen_quest.mjs (consolidation planner)")
    run(["node", "gen_quest.mjs"], "gen-quest")

    # ── 5. QUEST DATA for the new pages ──────────────────────────────────────
    banner("5 · QUEST-DATA — node gen_quest_data.mjs (snapshot/guide/board)")
    run(["node", "gen_quest_data.mjs"], "quest-data")

    # ── 6. FILL RATES report ─────────────────────────────────────────────────
    banner("6 · FILL RATES")
    run(["python3", "brb_fill_rates.py", xlsx], "fill-rates")

    save_state(xlsx, rows)

    # ── Optional credentialed phases ─────────────────────────────────────────
    if args.covers:
        credentialed_phase("Covers re-fetch", "COMIC_VINE_API_KEY",
                            ["node", "fetchCovers.mjs", "--retry-nulls"],
                            hint="then: node fetchCovers.mjs --retry-nulls")
    if args.ebay:
        credentialed_phase("eBay pricing", "EBAY_APP_ID",
                            ["python3", "brb_ebay_pricing.py", "--min-value", "10"],
                            hint="(also needs EBAY_CERT_ID)")
    if args.writers:
        credentialed_phase("Writer/artist fill", "COMIC_VINE_API_KEY",
                            ["python3", "BRB_overnight_script_v2/run_overnight_v2.py"],
                            background=True, hint="(runs ~8h; safe-shutdown handler saves progress on kill)")

    # ── Optional git commit ──────────────────────────────────────────────────
    if args.commit:
        banner("GIT — commit generated files")
        gen_files = [
            "artifacts/comics-inventory/src/data/data3.ts",
            "artifacts/comics-inventory/public/covers.json",
            "artifacts/comics-inventory/public/box-quest.html",
            "artifacts/comics-inventory/public/quest-data.js",
        ]
        run(["git", "add"] + gen_files, "git add")
        code, _ = run(["git", "commit", "-m", args.commit], "git commit")
        if code == 0:
            run(["git", "push"], "git push")
        else:
            warn("Nothing to commit (no changes) — skipped push.")

    banner("DONE — pipeline complete")


if __name__ == "__main__":
    main()
