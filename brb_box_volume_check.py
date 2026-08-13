#!/usr/bin/env python3
"""
brb_box_volume_check.py — resolve the wrong Volume/Year on rows that COLLIDE as
false same-box duplicates.

Root cause (Roberto, 2707): the same-box dup validator keys on Title+Issue+Year
+Box with no Volume, so distinct volumes of one issue number look like dupes.
Underneath, some rows carry a WRONG Volume/Year — e.g. a Juann Cabal (2022
Ridley) Black Panther #2 mislabeled Vol 7 / year 2005 — which collides with the
genuine 2005 Vol 4 issue.

Method — the cover artist is the anchor:
  1. find same-box collision groups (Title+Issue+Year, >1 row) in the scoped boxes
  2. for each row, look up every real Fandom volume that has that issue number
     (volume_index.json), fetch each volume's release year + cover artist
  3. pick the volume whose cover-artist SURNAME matches the row's Cover Artist
  4. if that volume / year differs from the row's declared values, PROPOSE the
     correction (never applied here — dry-run only)

Fandom API is free (no Comic Vine quota). READ-ONLY: writes a review xlsx, never
the source file.

Usage:
    python3 brb_box_volume_check.py --boxes 7,10
    python3 brb_box_volume_check.py --boxes 7,10 --out box_vol_review.xlsx
"""
import argparse, glob, json, os, re, time, urllib.parse, urllib.request
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
API = "https://marvel.fandom.com/api.php"
UA = "MarshallComicsInventory/1.0"
DELAY = 0.3


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def ni(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s); return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def yr(v):
    n = [int(t) for t in re.findall(r"\d{4}", str(v or "")) if 1900 < int(t) < 2100]
    return n[0] if n else None


def surname(s):
    s = re.sub(r"[^a-z ]", " ", str(s or "").lower())
    # take the first credited artist (before &/,), then its last word
    first = re.split(r"[&,]", s)[0]
    parts = [p for p in first.split() if len(p) > 2]
    return parts[-1] if parts else ""


_cache = {}


def fandom(page):
    if page in _cache:
        return _cache[page]
    out = None
    try:
        u = (f"{API}?action=parse&page={urllib.parse.quote(page)}&prop=wikitext"
             "&format=json&formatversion=2&redirects=1")
        with urllib.request.urlopen(urllib.request.Request(u, headers={"User-Agent": UA}), timeout=30) as r:
            d = json.load(r)
        wt = d.get("parse", {}).get("wikitext", "")
        if wt:
            def f(*names):
                for n in names:
                    m = re.search(r"\|\s*" + n + r"\s*=\s*([^\n|]+)", wt)
                    if m:
                        v = re.sub(r"\[\[|\]\]", "", m.group(1)).split("|")[-1].strip()
                        if v:
                            return v
                return None
            out = {"year": yr(f("ReleaseDate", "CoverDate", "Year")),
                   "cover": f("CoverArtist1_1", "CoverArtist_1", "CoverArtist1"),
                   "penciler": f("Penciler1_1", "Penciler_1", "Artist1_1")}
    except Exception:
        pass
    _cache[page] = out
    time.sleep(DELAY)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--boxes", default="7,10")
    ap.add_argument("--out", default="box_vol_review.xlsx")
    args = ap.parse_args()
    boxes = {b.strip() for b in args.boxes.split(",")}

    vindex = json.load(open(os.path.join(ROOT, "volume_index.json"))) if os.path.exists(os.path.join(ROOT, "volume_index.json")) else {}
    xlsx = latest_xlsx()
    print(f"Source: {os.path.basename(xlsx)}   boxes: {sorted(boxes)}")
    wb = openpyxl.load_workbook(xlsx)
    ws = next(wb[n] for n in wb.sheetnames if (n == "Sheet X" or n.startswith("✅ Clean Inventory")))
    H = [c.value for c in next(ws.iter_rows(max_row=1))]
    C = {n: H.index(n) for n in H if n}
    ti, ii, yi, vi, cai = C["Title"], C["Issue #"], C["Year"], C["Volume"], C["Cover Artist"]

    # collision groups: same Title+Issue+Year+Box, >1 row
    groups = defaultdict(list)
    for row in ws.iter_rows(min_row=2):
        b = str(row[C["Box #"]].value or "").strip()
        if b not in boxes:
            continue
        k = (str(row[ti].value or "").strip(), ni(row[ii].value), str(row[yi].value or "").strip(), b)
        groups[k].append(row)
    coll = {k: v for k, v in groups.items() if len(v) > 1}
    print(f"  same-box collision groups in scope: {len(coll)}")

    props = []
    for (title, iss, y, box), rowlist in coll.items():
        cands = vindex.get(title, {}).get(iss, [])
        if not cands:
            continue
        # fetch each candidate volume once
        volinfo = {}
        for v in cands:
            volinfo[v] = fandom(f"{title} Vol {v} {iss}")
        for row in rowlist:
            rcov = surname(row[cai].value)
            dv = ni(row[vi].value); dy = yr(row[yi].value)
            match = None
            for v in cands:
                info = volinfo.get(v)
                if info and rcov and surname(info.get("cover")) == rcov:
                    match = (v, info); break
                if info and rcov and surname(info.get("penciler")) == rcov:
                    match = (v, info); break
            if not match:
                continue
            mv, minfo = match
            newy = minfo["year"]
            volwrong = str(mv) != str(dv)
            yearwrong = newy and dy and abs(newy - dy) > 1
            if volwrong or yearwrong:
                props.append(dict(title=title, issue=iss, box=box, cover=row[cai].value,
                                  old_vol=dv, new_vol=mv, old_year=dy, new_year=newy,
                                  why=f"cover artist '{row[cai].value}' matches {title} Vol {mv} ({newy})"))

    print(f"\n  PROPOSED corrections (cover-artist anchored): {len(props)}")
    print(f"  {'Title':<24}{'#':<5}{'box':<4}{'vol':>8}{'year':>12}  cover")
    for p in sorted(props, key=lambda x: (x["title"], x["issue"]))[:40]:
        vc = f"{p['old_vol']}->{p['new_vol']}" if str(p['old_vol']) != str(p['new_vol']) else str(p['old_vol'])
        yc = f"{p['old_year']}->{p['new_year']}" if p['old_year'] != p['new_year'] else str(p['old_year'])
        print(f"  {p['title'][:23]:<24}#{p['issue']:<4}{p['box']:<4}{vc:>8}{yc:>12}  {str(p['cover'])[:22]}")

    out = openpyxl.Workbook(); sh = out.active; sh.title = "Box vol-year corrections"
    sh.append(["Title", "Issue", "Box", "Cover Artist", "Vol now", "Vol proposed", "Year now", "Year proposed", "Reason", "Approve? (y/n)"])
    for c in sh[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="C00000")
    sh.freeze_panes = "A2"
    for i, w in enumerate((24, 7, 5, 22, 8, 12, 9, 13, 50, 14), 1):
        sh.column_dimensions[chr(64 + i)].width = w
    for p in sorted(props, key=lambda x: (x["title"], x["issue"])):
        sh.append([p["title"], p["issue"], p["box"], p["cover"], p["old_vol"], p["new_vol"],
                   p["old_year"], p["new_year"], p["why"], ""])
    out.save(args.out)
    print(f"\n  Written: {args.out}  (DRY-RUN — nothing applied)")


if __name__ == "__main__":
    main()
