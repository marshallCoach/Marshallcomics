#!/usr/bin/env python3
"""Append this/last week's new comics (28 rows) to the current canonical.
Runs on the Mac against the newest attached_assets/comics_inventory_*.xlsx so it
picks up the post-weekly enrichment (no cloud regression). Adds rows only —
never edits existing ones. Confident creators are filled; uncertain ones are
left BLANK for the weekly's GCD/CV pass. Defaults applied (override in DATA if
wrong): Box = "UNKNOWN — needs physical reassignment", Year = 2026 (except the
Iron Man #282 back-issue key = 1992), Condition = NM (blank for the IM #282
back-issue). After running: python3 brb.py --commit "add new comics" --yes
does the validation, dup-matching, regen and push."""
import openpyxl, glob, os, datetime

SRC = max((f for f in glob.glob("attached_assets/comics_inventory_*.xlsx")
           if " copy" not in f and not os.path.basename(f).startswith("~$")),
          key=os.path.getmtime)
UNK = "UNKNOWN — needs physical reassignment"

# fields: title, issue, pub, year, w, a, c, variant, key(bool), keywhy, first, cond, imprint
def B(title, issue, pub="", year="2026", w="", a="", c="", variant="",
      key=False, keywhy="", first="", cond="NM", imprint=""):
    return dict(title=title, issue=issue, pub=pub, year=year, w=w, a=a, c=c,
                variant=variant, key=key, keywhy=keywhy, first=first, cond=cond, imprint=imprint)

DATA = [
 B("Black Cat", "13", "Marvel", variant="Variant edition — Cover C (bought for the cover)"),
 B("Iron Man", "282", "Marvel", year="1992", w="Len Kaminski", a="Kevin Hopgood", c="Kevin Hopgood",
   key=True, keywhy="First full appearance of War Machine (James Rhodes in the War Machine armor) [verify]",
   first="War Machine (James Rhodes) [verify]", cond=""),
 B("Ben 10", "5", "", variant="Cover A"),
 B("Capes", "10", "Image", variant="Cover B (bought for the cover)"),
 B("Doom Patrol", "1", "DC", variant="Cover A — new series"),
 B("Absolute Wonder Woman", "23", "DC", variant="Cover A"),
 B("Absolute Green Lantern", "18", "DC", variant="Variant cover — 1 copy"),
 B("The World to Come", "6", "", variant="Variant edition — Cover B (completes the run)"),
 B("Fantastic Four", "16", "Marvel", variant="Alex Ross Timeless variant (Crystal)"),
 B("Captain America", "15", "Marvel", variant="Alex Ross Timeless variant (Wonder Man)"),
 B("Spider-Woman", "1", "Marvel", c="Olivier Coipel", variant="Anniversary Special — Cover A"),
 B("X-Men: Omega Red Dawn", "1", "Marvel", variant="new series"),
 B("Bishop", "3", "Marvel", w="Saladin Ahmed", variant="Bishop cover"),
 B("Doctor Strange & Doctor Doom: Triumph and Torment", "1", "Marvel",
   w="Roger Stern", a="Mike Mignola", variant="Special — reprints the 1989 graphic novel"),
 B("Black Panther & Namor: Doomed", "1", "Marvel", c="Taurin Clarke", variant="Cover B (Taurin Clarke)"),
 B("Absolute Catwoman", "3", "DC", w="Ann Nocenti", variant="Cover A"),
 B("Justice League Unlimited", "22", "DC", w="Mark Waid", a="Dan Mora", variant="Cover A"),
 B("Batman", "13", "DC", w="Matt Fraction", a="Matteo Scalera", c="José Ladrönn",
   variant="Bane variant — Hispanic Heritage Month cover by José Ladrönn"),
 B("JSA", "23", "DC", w="Jeff Lemire", c="Teddy Kristiansen", variant="Variant cover (Teddy Kristiansen)"),
 B("Teen Titans", "1", "DC", variant="Variant cover — new series"),
 B("Dark Knights of Steel: Heir to the Sea", "2", "DC", w="Tom Taylor", a="Riccardo Federici",
   variant="Elseworlds one-shot"),
 B("Supergirl: Survive", "4", "DC", c="Taurin Clarke", variant="Variant cover (Taurin Clarke)"),
 B("Batgirl", "23", "DC", w="Tate Brombal", variant="Cover A"),
 B("Void Rivals", "33", "Image", variant="Cover A", imprint="Skybound"),
 B("Rom", "1", "Image", w="Robert Kirkman", a="Lorenzo De Felici", variant="Cover A — new series", imprint="Skybound"),
 B("M.A.S.K.", "4", "Image", imprint="Skybound"),
 B("M.A.S.K. Origins", "1", "Image", imprint="Skybound"),
 B("Terminal", "2", "Image", variant="Kirkman/Casey [verify creators]", imprint="Skybound"),
]

wb = openpyxl.load_workbook(SRC)
ws = next(w for w in wb.worksheets if w.title.startswith("✅ Clean Inventory"))
H = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
def ci(name):
    return H.index(name) + 1 if name in H else None
COLS = {k: ci(v) for k, v in {
    "title":"Title", "issue":"Issue #", "pub":"Publisher", "year":"Year",
    "w":"Writer(s)", "a":"Artist(s)", "c":"Cover Artist", "key":"Key Issue?",
    "keywhy":"Key Issue — Why", "first":"1st Appearances", "cond":"Condition",
    "box":"Box #", "variant":"Seller Notes / Variants / Caveats", "date":"Date Added",
    "imprint":"Imprint", "entry":"#",
}.items()}

# next entry index
maxentry = 0
if COLS["entry"]:
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, COLS["entry"]).value
        try: maxentry = max(maxentry, int(float(v)))
        except (TypeError, ValueError): pass

today = datetime.date.today().strftime("%b %-d, %Y") if os.name != "nt" else datetime.date.today().strftime("%b %d, %Y")
start = ws.max_row + 1
for n, bk in enumerate(DATA):
    r = start + n
    def put(key, val):
        if COLS[key] and val != "": ws.cell(r, COLS[key], val)
    put("title", bk["title"]); put("issue", bk["issue"]); put("pub", bk["pub"])
    put("year", bk["year"]); put("w", bk["w"]); put("a", bk["a"]); put("c", bk["c"])
    put("cond", bk["cond"]); put("variant", bk["variant"]); put("imprint", bk["imprint"])
    if COLS["box"]: ws.cell(r, COLS["box"], UNK)
    if bk["key"]:
        put("key", "YES"); put("keywhy", bk["keywhy"]); put("first", bk["first"])
    if COLS["entry"]: ws.cell(r, COLS["entry"], maxentry + 1 + n)
    put("date", f"{today} (new-comics intake)")

out = f"attached_assets/comics_inventory_{datetime.datetime.now():%d%m_%H%M}.xlsx"
wb.save(out)
print(f"SOURCE: {os.path.basename(SRC)}")
print(f"OUTPUT: {os.path.basename(out)}")
print(f"rows appended: {len(DATA)}  (was {start-1} data rows -> {ws.max_row})")
print(f"Box set to: {UNK!r} for all — reassign physically.")
print("Next: python3 brb.py --commit \"add new comics (2 weeks)\" --yes   (validates + dup-matches + pushes)")
