#!/usr/bin/env python3
"""
brb_gcd_volume_check.py — compare the inventory's declared Volume column
against GCD's series catalog, offline (the volume-numbering tiebreaker use
case from the original GCD brief).

Method:
  1. For each Title+Publisher, take ALL matching GCD series from that
     publisher (publisher filtering excludes GCD's many foreign-reprint series
     that share the name), ordered chronologically by year_began — that
     ordering is the derived volume numbering (Vol 1 = earliest), and unlike a
     collection-local ranking it includes volumes you don't own.
  2. Resolve each inventory row to one of those series by Year overlap.
  3. Compare the derived number to the declared Volume.

Read-only: writes nothing except an optional CSV worklist. Disagreements are a
REVIEW list, not corrections — publishers occasionally skip/renumber volumes,
and GCD itself can catalog specials as separate series.

Usage:
    python3 brb_gcd_volume_check.py [--csv volume_check.csv]
"""
import argparse, glob, os, re, sqlite3, sys
from collections import defaultdict

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
DB = os.path.join(ROOT, "gcd_local.sqlite")


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def parse_year_range(y):
    nums = [int(n) for n in re.findall(r"\d{4}", str(y or ""))]
    nums = [n for n in nums if 1900 < n < 2100]
    return (min(nums), max(nums)) if nums else None


def pub_match(inv_pub, gcd_pub):
    if not inv_pub or not gcd_pub:
        return False
    a, b = inv_pub.lower().strip(), gcd_pub.lower().strip()
    return a in b or b in a


def tight(t):
    """Same loose key as gcd_rescan_missing.py — '&'='and', drop all
    punctuation AND spaces, so split inventory spellings of one real title
    ('Ultimate Comics: X-Men' vs 'Ultimate Comics X-Men') merge into one
    series pool instead of undercounting both."""
    t = re.sub(r"&", " and ", str(t or "").lower())
    return re.sub(r"[^a-z0-9]", "", t)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", default=None)
    args = ap.parse_args()

    xlsx = latest_xlsx()
    print(f"Inventory: {os.path.basename(xlsx)}")
    conn = sqlite3.connect(DB)

    # All GCD series pooled by the loose tight-key (merges matched_title
    # spelling variants), plus explicit aliases from gcd_rescan_missing.py.
    gcd_by_key = defaultdict(dict)  # tight-key -> {series_id: series}
    for sid, mt, name, yb, ye, pub in conn.execute(
        "SELECT s.id, s.matched_title, s.name, s.year_began, s.year_ended, p.name "
        "FROM gcd_series s LEFT JOIN gcd_publisher p ON p.id = s.publisher_id"):
        gcd_by_key[tight(mt)][sid] = {"id": sid, "name": name, "yb": yb, "ye": ye, "pub": pub}
    try:
        aliases = dict(conn.execute("SELECT alias, matched_title FROM gcd_title_alias"))
    except sqlite3.OperationalError:
        aliases = {}

    def series_for(title):
        pool = dict(gcd_by_key.get(tight(title), {}))
        al = aliases.get(title)
        if al:
            pool.update(gcd_by_key.get(tight(al), {}))
        return list(pool.values())

    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = next(wb[n] for n in wb.sheetnames if (n == "Sheet X" or n.startswith("✅ Clean Inventory")))
    rows = list(ws.iter_rows(values_only=True))
    H = list(rows[0])
    ti, ii, yi, vi, pi, bi = (H.index(c) for c in ("Title", "Issue #", "Year", "Volume", "Publisher", "Box #"))

    agree = disagree = no_declared = no_gcd = ambiguous = 0
    worklist = []
    # cache the derived numbering per (title, publisher)
    numbering_cache = {}

    for r in rows[1:]:
        title = str(r[ti] or "").strip()
        if not title:
            continue
        declared = r[vi]
        if declared is None or str(declared).strip() in ("", "nan"):
            no_declared += 1
            continue
        try:
            declared_n = int(float(declared))
        except (ValueError, TypeError):
            no_declared += 1
            continue

        pub = str(r[pi] or "").strip()
        ck = (title, pub)
        if ck not in numbering_cache:
            cands = [s for s in series_for(title) if pub_match(pub, s["pub"]) and s["yb"]]
            # de-duplicate same-name-same-year (GCD sometimes has printings as
            # separate series entries starting the same year) by keeping the
            # earliest id per year_began
            by_year = {}
            for s in sorted(cands, key=lambda s: (s["yb"], s["id"])):
                by_year.setdefault(s["yb"], s)
            ordered = sorted(by_year.values(), key=lambda s: s["yb"])
            numbering_cache[ck] = ordered
        ordered = numbering_cache[ck]
        if not ordered:
            no_gcd += 1
            continue

        yr = parse_year_range(r[yi])
        if not yr:
            ambiguous += 1
            continue
        # resolve the row to the series whose year window overlaps the row's Year
        hits = [i for i, s in enumerate(ordered)
                if s["yb"] <= yr[1] and (s["ye"] or 2100) >= yr[0]]
        if len(hits) != 1:
            ambiguous += 1
            continue
        derived_n = hits[0] + 1

        if derived_n == declared_n:
            agree += 1
        else:
            disagree += 1
            s = ordered[hits[0]]
            worklist.append({
                "Title": title, "Issue": str(r[ii]), "Box": str(r[bi]), "Year": str(r[yi]),
                "Declared_Volume": declared_n, "GCD_Derived_Volume": derived_n,
                "GCD_Series": f"{s['name']} ({s['yb']}-{s['ye'] or '?'})",
                "GCD_volume_count_for_title": len(ordered),
            })

    total = agree + disagree
    print(f"\n{'='*64}")
    print(f"  VOLUME CHECK — inventory declared vs GCD-derived (publisher-filtered)")
    print(f"{'='*64}")
    print(f"  Rows compared (resolved to exactly one GCD series): {total:,}")
    print(f"  ✓ Agree:    {agree:,}  ({100*agree/total:.1f}%)" if total else "")
    print(f"  ✗ Disagree: {disagree:,}  ({100*disagree/total:.1f}%)" if total else "")
    print(f"  · No declared Volume on row: {no_declared:,}")
    print(f"  · Title has no GCD series for this publisher: {no_gcd:,}")
    print(f"  · Year missing/ambiguous (0 or 2+ series overlap): {ambiguous:,}")

    from collections import Counter
    by_title = Counter(w["Title"] for w in worklist)
    print(f"\n  Top disagreeing titles:")
    for t, n in by_title.most_common(15):
        print(f"    {n:>4}  {t}")

    print(f"\n  Sample disagreements:")
    seen_titles = set()
    shown = 0
    for w in worklist:
        if w["Title"] in seen_titles:
            continue
        seen_titles.add(w["Title"])
        print(f"    {w['Title']} #{w['Issue']} ({w['Year']}, Box {w['Box']}): declared Vol {w['Declared_Volume']} vs GCD Vol {w['GCD_Derived_Volume']} [{w['GCD_Series']}]")
        shown += 1
        if shown >= 12:
            break

    if args.csv and worklist:
        import csv
        with open(args.csv, "w", newline="") as f:
            wr = csv.DictWriter(f, fieldnames=list(worklist[0].keys()))
            wr.writeheader(); wr.writerows(worklist)
        print(f"\n  Worklist written: {args.csv} ({len(worklist)} rows)")


if __name__ == "__main__":
    main()
