#!/usr/bin/env python3
"""
brb_null_covers.py — Null out incorrect cover entries so --retry-nulls can re-fetch them.
Run from: ~/marshallcomics/

Usage (manual, Title+Issue only - ambiguous if a title+issue exists across
multiple volumes/boxes, e.g. the same book bought twice in different printings):
    python3 brb_null_covers.py "Batman and Robin" 4.9
    python3 brb_null_covers.py "U.S. Agent" 1 "Batwoman" 4 "Uncanny X-Men" 406

Usage (from a Cover Review export - PREFERRED when you have one, since it
carries Box + Year and can resolve the exact volume instead of guessing):
    python3 brb_null_covers.py --from-json ~/Downloads/flagged-covers-2026-07-12.json [--xlsx path]

    Cross-references each flagged Title+Issue+Box(+Year) against the current
    inventory xlsx (auto-detects the newest in attached_assets/ unless --xlsx
    is given) to find the exact Volume, then nulls only that specific
    covers.json entry. Only falls back to nulling every matching volume when
    Box+Year genuinely can't disambiguate (reported explicitly, not silent).

    Pass title+issue pairs in sequence. Quotes required for titles with spaces.
    After running, execute: node fetchCovers.mjs --retry-nulls
"""

from __future__ import annotations
import sys, os, json, glob as _glob

REPO_ROOT   = os.path.dirname(os.path.abspath(__file__))
COVERS_PATH = os.path.join(REPO_ROOT, "covers.json")
ASSETS_DIR  = os.path.join(REPO_ROOT, "attached_assets")


def _latest_xlsx() -> str | None:
    """Newest comics_inventory_*.xlsx by mtime in attached_assets/ (same
    auto-detect convention as gen_data.mjs — reads the one current file
    regardless of whether its name contains 'VALIDATED')."""
    matches = [f for f in _glob.glob(os.path.join(ASSETS_DIR, "comics_inventory_*.xlsx"))
               if not os.path.basename(f).startswith("~$")]
    if not matches:
        return None
    return max(matches, key=os.path.getmtime)


def _norm_issue(v) -> str:
    s = str(v).strip()
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else s
    except ValueError:
        return s


def resolve_volumes_from_xlsx(flagged: list[dict], xlsx_path: str) -> tuple[dict, list]:
    """For each flagged entry ({Title, Issue, Box, Year}), find the exact
    Volume by cross-referencing Title+Issue+Box in the xlsx, falling back to
    Title+Issue+Box+Year when Box alone matches more than one row. Returns
    (resolved: {id -> volume}, unresolved: [(id, xlsx_matches, flagged_year)])
    for entries that stay ambiguous even after that."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = next((wb[n] for n in wb.sheetnames if n.startswith("✅ Clean Inventory")), None)
    if ws is None:
        return {}, [(f["id"], [], f.get("Year")) for f in flagged]

    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    ti, ii, bi = headers.index("Title"), headers.index("Issue #"), headers.index("Box #")
    vi = headers.index("Volume") if "Volume" in headers else None
    yi = headers.index("Year") if "Year" in headers else None

    by_tib: dict[tuple, list] = {}
    for r in rows[1:]:
        title, issue, box = str(r[ti]), _norm_issue(r[ii]), str(r[bi])
        vol  = r[vi] if vi is not None else None
        year = r[yi] if yi is not None else None
        by_tib.setdefault((title, issue, box), []).append((vol, year))

    resolved, unresolved = {}, []
    for f in flagged:
        key = (f["Title"], _norm_issue(f["Issue"]), str(f.get("Box", "")))
        matches = by_tib.get(key, [])
        if len(matches) == 1:
            resolved[f["id"]] = matches[0][0]
        elif len(matches) > 1:
            year_matches = {v for v, y in matches if str(y) == str(f.get("Year"))}
            if len(year_matches) == 1:
                resolved[f["id"]] = next(iter(year_matches))
            else:
                unresolved.append((f["id"], matches, f.get("Year")))
        else:
            unresolved.append((f["id"], [], f.get("Year")))
    return resolved, unresolved


def find_keys(covers, title, issue, volume=None):
    """Find covers.json key(s) for a title+issue, tolerating float formatting
    and the Title|||Issue|||Volume key format. If `volume` is given (resolved
    via Box+Year cross-reference against the xlsx), only that specific
    volume's key is targeted - no guessing. Without a volume, multiple
    physical copies of the same title+issue in different volumes/runs share
    the same Title+Issue text but get distinct keys, so this falls back to
    returning every match (caller nulls all of them, since it can't
    disambiguate from title+issue alone)."""
    title_lower = title.lower()
    issue_str   = str(issue)
    issue_is_num = issue_str.replace(".", "", 1).isdigit()

    def issue_matches(stored_issue):
        try:
            return abs(float(stored_issue) - float(issue_str)) < 0.01
        except ValueError:
            return stored_issue == issue_str

    if volume is not None:
        candidates = [
            f"{title}|||{issue}|||{volume}",
            f"{title}|||{issue}|||{float(volume)}",
            f"{title}|||{issue}|||{int(float(volume))}.0" if str(volume).replace(".", "", 1).isdigit() else None,
        ]
        exact = [c for c in candidates if c and c in covers]
        if exact:
            return exact
        # Volume given but no 3-part key exists - this title/issue only has a
        # plain 2-part entry (doesn't distinguish by volume). Fall through.

    # Exact 2-part candidates (fast path)
    candidates = [
        f"{title}|||{issue}",
        f"{title}|||{float(issue)}" if issue_is_num else None,
        f"{title}|||{int(float(issue))}.0" if issue_is_num else None,
    ]
    exact = [c for c in candidates if c and c in covers]
    if exact:
        return exact

    # Fuzzy scan: case-insensitive title match, both 2-part and 3-part keys.
    # Collect every match - do not stop at the first one.
    matches = []
    for k in covers:
        parts = k.split("|||")
        if len(parts) not in (2, 3):
            continue
        if parts[0].lower() != title_lower:
            continue
        if issue_matches(parts[1]):
            matches.append(k)
    return matches


def _null_keys(covers, keys, nulled_log):
    for key in keys:
        old = covers[key]
        covers[key] = None
        old_url = (old or {}).get("url", "no url") if old else "already null"
        nulled_log.append(key)
        print(f"  ✓  {key}  →  null  (was: {str(old_url)[:60]})")


def run_from_json(json_path: str, xlsx_arg: str | None):
    with open(json_path) as f:
        flagged = json.load(f)
    if not isinstance(flagged, list) or not flagged:
        print(f"ERROR: {json_path} is not a non-empty JSON array of flagged-cover entries.")
        sys.exit(1)

    xlsx_path = xlsx_arg or _latest_xlsx()
    if not xlsx_path or not os.path.exists(xlsx_path):
        print(f"ERROR: No inventory xlsx found (looked in {ASSETS_DIR}). Pass --xlsx explicitly.")
        sys.exit(1)
    print(f"Cross-referencing against: {os.path.relpath(xlsx_path, REPO_ROOT)}")

    resolved, unresolved = resolve_volumes_from_xlsx(flagged, xlsx_path)
    print(f"Resolved to exactly one volume via Title+Issue+Box(+Year): {len(resolved)} / {len(flagged)}")
    if unresolved:
        print(f"Genuinely ambiguous or not found in xlsx (falling back to title+issue match): {len(unresolved)}")
        for fid, matches, year in unresolved:
            print(f"  ⚠  {fid}  (flagged year={year!r})  xlsx candidates: {matches or 'none'}")

    with open(COVERS_PATH) as f:
        covers = json.load(f)

    nulled, missing = [], []
    for f in flagged:
        title, issue = f["Title"], f["Issue"]
        volume = resolved.get(f["id"])
        keys = find_keys(covers, title, issue, volume)
        if not keys:
            missing.append(f"  NOT FOUND: '{title}' #{issue}")
            continue
        if len(keys) > 1:
            print(f"  ⚠  '{title}' #{issue}  matches {len(keys)} volumes even after "
                  f"Box+Year - nulling all of them")
        _null_keys(covers, keys, nulled)

    _finish(covers, nulled, missing)


def run_from_pairs(pairs):
    with open(COVERS_PATH) as f:
        covers = json.load(f)

    nulled, missing = [], []
    for title, issue in pairs:
        keys = find_keys(covers, title, issue)
        if not keys:
            missing.append(f"  NOT FOUND: '{title}' #{issue}")
            continue
        if len(keys) > 1:
            print(f"  ⚠  '{title}' #{issue}  matches {len(keys)} volumes - "
                  f"nulling all of them (can't disambiguate from title+issue alone - "
                  f"use --from-json with a Cover Review export to resolve this precisely)")
        _null_keys(covers, keys, nulled)

    _finish(covers, nulled, missing)


def _finish(covers, nulled, missing):
    if missing:
        print()
        for m in missing:
            print(f"  ✗  {m}")
    if nulled:
        with open(COVERS_PATH, "w") as f:
            json.dump(covers, f, indent=2)
        print(f"\n{len(nulled)} cover(s) nulled in covers.json")
        print("Run next:  node fetchCovers.mjs --retry-nulls")
    else:
        print("\nNothing changed.")


def main():
    args = sys.argv[1:]

    if "--from-json" in args:
        i = args.index("--from-json")
        if i + 1 >= len(args):
            print("Usage: python3 brb_null_covers.py --from-json <path.json> [--xlsx <path>]")
            sys.exit(1)
        json_path = args[i + 1]
        xlsx_arg = None
        if "--xlsx" in args:
            xi = args.index("--xlsx")
            if xi + 1 < len(args):
                xlsx_arg = args[xi + 1]
        if not os.path.exists(COVERS_PATH):
            print(f"ERROR: covers.json not found at {COVERS_PATH}")
            sys.exit(1)
        run_from_json(json_path, xlsx_arg)
        return

    if not args or len(args) % 2 != 0:
        print("Usage: python3 brb_null_covers.py \"Title\" issue# [\"Title2\" issue2# ...]")
        print("  e.g: python3 brb_null_covers.py \"U.S. Agent\" 1 \"Batwoman\" 4")
        print("  or:  python3 brb_null_covers.py --from-json ~/Downloads/flagged-covers-....json")
        sys.exit(1)

    if not os.path.exists(COVERS_PATH):
        print(f"ERROR: covers.json not found at {COVERS_PATH}")
        sys.exit(1)

    pairs = [(args[i], args[i + 1]) for i in range(0, len(args), 2)]
    run_from_pairs(pairs)


if __name__ == "__main__":
    main()
