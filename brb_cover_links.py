#!/usr/bin/env python3
"""brb_cover_links.py — build a human-browsable index of every inventory book and
its resolved cover URL. One row per book; reads covers.json with the same lookup
the app uses (volume key -> legacy -> #-form). Output: cover_links_DDMM_HHMM.xlsx
with a clickable HYPERLINK on the cover cell, plus a summary of fill/blank."""
import glob, os, json, re, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
covers = json.load(open(os.path.join(ROOT, "covers.json")))


def ni(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s); return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def nv(v):
    try:
        return str(int(float(str(v).strip())))
    except (ValueError, TypeError):
        return "1"


def yr(v):
    n = [int(t) for t in re.findall(r"\d{4}", str(v or "")) if 1900 < int(t) < 2100]
    return n[0] if n else None


def cover(t, iss, vol):
    for k in (f"{t}|||{iss}|||{vol}", f"{t}|||{iss}", f"{t}|||#{iss}"):
        e = covers.get(k)
        if e and e.get("url"):
            return e
    return None


x = max(glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")), key=os.path.getmtime)
wb = openpyxl.load_workbook(x, read_only=True, data_only=True)
ws = next(wb[n] for n in wb.sheetnames if (n == "Sheet X" or n.startswith("✅ Clean Inventory")))
rows = list(ws.iter_rows(values_only=True)); H = list(rows[0]); C = {n: H.index(n) for n in H if n}


def g(r, n):
    i = C.get(n); return r[i] if i is not None else None


out = openpyxl.Workbook(); sh = out.active; sh.title = "Cover links"
sh.append(["Title", "Issue", "Volume", "Year", "Box", "Publisher", "Status", "Source", "Cover URL", "Large URL"])
for c in sh[1]:
    c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2F5496")
sh.freeze_panes = "A2"
for i, w in enumerate((30, 8, 7, 6, 8, 14, 8, 20, 60, 60), 1):
    sh.column_dimensions[chr(64 + i)].width = w

filled = blank = 0
for r in rows[1:]:
    t = str(g(r, "Title") or "").strip()
    if not t:
        continue
    iss = ni(g(r, "Issue #")); vol = nv(g(r, "Volume"))
    e = cover(t, iss, vol)
    if e:
        filled += 1; status = "filled"; src = e.get("source", ""); url = e.get("url"); large = e.get("large") or ""
    else:
        blank += 1; status = "BLANK"; src = ""; url = ""; large = ""
    sh.append([t, g(r, "Issue #"), vol, g(r, "Year"), g(r, "Box #"),
               g(r, "Publisher"), status, src, url, large])
    cell = sh.cell(row=sh.max_row, column=9)
    if url:
        cell.hyperlink = url; cell.font = Font(color="0563C1", underline="single")
    else:
        cell.fill = PatternFill("solid", fgColor="FFF2CC")

# Summary sheet
ss = out.create_sheet("Summary")
ss.append(["Metric", "Count"])
for c in ss[1]:
    c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2F5496")
ss.append(["Inventory books", filled + blank])
ss.append(["Covers filled", filled])
ss.append(["Covers blank", blank])
ss.append(["Coverage %", round(100 * filled / (filled + blank), 1) if (filled + blank) else 0])
ss.column_dimensions["A"].width = 22; ss.column_dimensions["B"].width = 12

stamp = datetime.datetime.now().strftime("%d%m_%H%M")
name = f"cover_links_{stamp}.xlsx"
out.save(os.path.join(ROOT, name))
print(f"books: {filled + blank}  filled: {filled}  blank: {blank}  coverage: {round(100*filled/(filled+blank),1)}%")
print(f"Written: {name}")
