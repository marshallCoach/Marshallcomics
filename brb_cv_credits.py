#!/usr/bin/env python3
"""
brb_cv_credits.py — fill blank Writer(s) / Artist(s) / Cover Artist from Comic
Vine, for the rows local GCD cannot resolve.

Why: after the GCD passes, ~1,200 cells remain blank (Cover Artist 758, Artist
437, Writer 164) across ~797 rows GCD has no series/issue for. Comic Vine has
them. This is the last automated lever before manual research.

Method — two steps, mirroring the proven covers.ts pattern:
  1. resolve the row's series to a CV volume  (filter=name, scored by year)
  2. fetch that volume's issue and read person_credits, mapping:
       writer                       -> Writer(s)
       penciler/artist/inker        -> Artist(s)
       cover                        -> Cover Artist

SAFETY — the hard-won rules from this project:
  - ONLY writes cells that are currently BLANK. Never overwrites an existing
    credit (this is the is_blank guard whose absence once clobbered good data).
  - Volume resolution requires a year match, so a same-named wrong-era volume
    cannot be used (the JLA #100 / X-Men 1963 failure mode).
  - Never writes the source file; emits a new *_CVCREDITS.xlsx.
  - Resumable: --state keeps a JSON of processed rows so a crash/restart does
    not redo work, and results are flushed incrementally.

Needs COMIC_VINE_API_KEY. Rate-limited (CV allows ~200/hr) -> overnight job.
Do NOT run alongside a cover fetch; they share the same quota.

EFFICIENCY: the 982 rows needing credits span only ~336 distinct title+year
groups, so this resolves ONE volume per group and pulls that volume's issues in
a single call, then maps credits onto every row of the group. That is ~672 calls
(~3.5 hrs) instead of ~1,964 (~10.4 hrs).

Dry-run makes NO API calls — it prints the plan (groups, calls, ETA) so you can
size the job without spending quota.

Usage:
    export COMIC_VINE_API_KEY=...            # or rely on ~/.zshrc
    python3 brb_cv_credits.py                # plan only, no API calls
    python3 brb_cv_credits.py --limit 5 --apply   # small live sample
    python3 brb_cv_credits.py --apply             # full run
"""
import argparse, glob, json, os, re, time, urllib.parse, urllib.request
from collections import defaultdict
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
API = "https://comicvine.gamespot.com/api"
KEY = os.environ.get("COMIC_VINE_API_KEY", "")
UA = "MarshallComicsInventory/1.0"
DELAY = 19.0          # CV free tier ~200/hr
STATE = os.path.join(ROOT, ".cv_credits_state.json")

WRITER_ROLES = {"writer", "script", "plot"}
ARTIST_ROLES = {"penciler", "penciller", "artist", "inker", "painter", "layouts"}
COVER_ROLES = {"cover", "coverartist", "cover artist"}


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def blank(v):
    return v is None or str(v).strip() in ("", "nan", "None")


def norm_issue(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def year_of(v):
    m = re.findall(r"(19|20)\d{2}", str(v or ""))
    n = [int(x) for x in re.findall(r"\d{4}", str(v or "")) if 1900 < int(x) < 2100]
    return n[0] if n else None


VERBOSE = False


def cv(path, **params):
    p = {"api_key": KEY, "format": "json", **params}
    url = f"{API}/{path}/?{urllib.parse.urlencode(p)}"
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=30) as r:
        d = json.load(r)
    # CV signals problems in-band with status_code, not HTTP errors
    if d.get("status_code") not in (1, None):
        print(f"    [CV ERROR] status_code={d.get('status_code')} {d.get('error')}", flush=True)
    return d


_volcache = {}


def resolve_volume(title, year):
    """Find the CV volume for this title, REQUIRING a year corroboration."""
    k = (title.lower(), year)
    if k in _volcache:
        return _volcache[k]
    out = None
    try:
        d = cv("volumes", filter=f"name:{title}", field_list="id,name,start_year", limit=50)
        results = d.get("results", []) or []
        best, bs = None, 0
        near = []
        for v in results:
            if re.sub(r"[^a-z0-9]", "", (v.get("name") or "").lower()) != re.sub(r"[^a-z0-9]", "", title.lower()):
                continue
            sy = v.get("start_year")
            try:
                sy = int(sy)
            except (TypeError, ValueError):
                continue
            near.append((v.get("name"), sy))
            if year and abs(sy - year) <= 2:          # year gate — blocks wrong-era volumes
                score = 10 - abs(sy - year)
                if score > bs:
                    best, bs = v, score
        out = best
        if VERBOSE and not out:
            print(f"    [no volume] {title!r} year={year} — {len(results)} CV results, "
                  f"name-matched years={[y for _, y in near][:6]}", flush=True)
    except Exception as e:
        print(f"    [EXC resolve_volume] {title!r}: {type(e).__name__}: {e}", flush=True)
        out = None
    _volcache[k] = out
    time.sleep(DELAY)
    return out


def volume_issue_credits(volume_id):
    """All issues of a volume WITH credits, in as few calls as possible.
    CV's issue-list endpoint sometimes omits person_credits; if so, fall back to
    per-issue detail calls for the issues we actually need."""
    out = {}
    try:
        d = cv("issues", filter=f"volume:{volume_id}",
               field_list="id,issue_number,person_credits", limit=100)
        for res in d.get("results", []) or []:
            num = norm_issue(res.get("issue_number"))
            pc = res.get("person_credits")
            if pc is None:
                return None          # list endpoint has no credits -> caller falls back
            out[num] = _map_credits(pc)
    except Exception as e:
        print(f"    [EXC issue-list] volume {volume_id}: {type(e).__name__}: {e}", flush=True)
        return None
    finally:
        time.sleep(DELAY)
    if VERBOSE:
        print(f"    [volume {volume_id}] {len(out)} issues with credits", flush=True)
    return out


def _map_credits(pc):
    w = a = c = None
    for p in pc or []:
        roles = {x.strip().lower() for x in str(p.get("role", "")).split(",")}
        name = p.get("name")
        if not name:
            continue
        if not w and roles & WRITER_ROLES:
            w = name
        if not a and roles & ARTIST_ROLES:
            a = name
        if not c and roles & COVER_ROLES:
            c = name
    return {"writer": w, "artist": a, "cover": c}


def credits_for(volume_id, issue):
    try:
        d = cv("issues", filter=f"volume:{volume_id},issue_number:{issue}",
               field_list="id,person_credits,issue_number", limit=1)
        res = (d.get("results") or [None])[0]
        if not res:
            return None
        w = a = c = None
        for p in res.get("person_credits", []) or []:
            roles = {x.strip().lower() for x in str(p.get("role", "")).split(",")}
            name = p.get("name")
            if not name:
                continue
            if not w and roles & WRITER_ROLES:
                w = name
            if not a and roles & ARTIST_ROLES:
                a = name
            if not c and roles & COVER_ROLES:
                c = name
        return {"writer": w, "artist": a, "cover": c}
    except Exception:
        return None
    finally:
        time.sleep(DELAY)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=None)
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--out", default=None)
    ap.add_argument("--verbose", action="store_true", help="explain every lookup (use with --limit)")
    args = ap.parse_args()
    global VERBOSE
    VERBOSE = args.verbose

    if not KEY:
        raise SystemExit("ERROR: COMIC_VINE_API_KEY not set (source ~/.zshrc)")
    xlsx = os.path.abspath(args.xlsx) if args.xlsx else latest_xlsx()
    print(f"Source: {os.path.basename(xlsx)}   (mode: {'APPLY' if args.apply else 'DRY-RUN'})")

    state = {}
    if os.path.exists(STATE):
        state = json.load(open(STATE))

    wb = openpyxl.load_workbook(xlsx)
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("✅ Clean Inventory"))
    H = [c.value for c in next(ws.iter_rows(max_row=1))]
    C = {n: H.index(n) for n in H if n}
    ti, ii, yi, wi, ai, ci = (C["Title"], C["Issue #"], C["Year"],
                              C["Writer(s)"], C["Artist(s)"], C["Cover Artist"])

    # ---- group rows by (title, year): one volume resolve per GROUP, not per row
    groups = defaultdict(list)
    for row in ws.iter_rows(min_row=2):
        if blank(row[wi].value) or blank(row[ai].value) or blank(row[ci].value):
            groups[(str(row[ti].value or "").strip(), year_of(row[yi].value))].append(row)
    nrows = sum(len(v) for v in groups.values())
    todo = [g for g in groups if state.get(f"G::{g[0]}::{g[1]}") != "done"]
    if args.limit:
        todo = todo[: args.limit]
    print(f"  rows needing credits : {nrows}")
    print(f"  distinct title+year groups: {len(groups)}  ({len(todo)} not yet done)")
    print(f"  estimated CV calls   : ~{len(todo)*2}  (~{len(todo)*2*DELAY/3600:.1f} hrs)")

    if not args.apply:
        print("\n[DRY RUN] No API calls made, nothing written.")
        print("  run a small live sample:  python3 brb_cv_credits.py --limit 5 --apply")
        return

    fw = fa = fc = 0
    nogroup = 0
    for n, gkey in enumerate(todo, 1):
        title, year = gkey
        rows_in = groups[gkey]
        if VERBOSE:
            print(f"  [{n}/{len(todo)}] {title!r} year={year} ({len(rows_in)} rows)", flush=True)
        vol = resolve_volume(title, year)
        if not vol:
            nogroup += 1
            state[f"G::{title}::{year}"] = "done"
            continue
        # one call for the whole volume; fall back per-issue if credits absent
        if VERBOSE:
            print(f"    -> volume {vol['id']} {vol.get('name')!r} ({vol.get('start_year')})", flush=True)
        table = volume_issue_credits(vol["id"])
        for row in rows_in:
            iss = norm_issue(row[ii].value)
            cr = table.get(iss) if table is not None else credits_for(vol["id"], iss)
            if not cr:
                continue
            if blank(row[wi].value) and cr["writer"]:
                row[wi].value = cr["writer"]; fw += 1
            if blank(row[ai].value) and cr["artist"]:
                row[ai].value = cr["artist"]; fa += 1
            if blank(row[ci].value) and cr["cover"]:
                row[ci].value = cr["cover"]; fc += 1
        state[f"G::{title}::{year}"] = "done"
        if n % 5 == 0:
            json.dump(state, open(STATE, "w"))
            out = args.out or xlsx.replace(".xlsx", "_CVCREDITS.xlsx")
            wb.save(out)                      # incremental flush — survives a crash
            print(f"  [{n}/{len(todo)} groups] W+{fw} A+{fa} CA+{fc} unresolved={nogroup}", flush=True)

    json.dump(state, open(STATE, "w"))
    out = args.out or xlsx.replace(".xlsx", "_CVCREDITS.xlsx")
    wb.save(out)
    print(f"\n  Writers +{fw}   Artists +{fa}   Cover Artists +{fc}   groups unresolved {nogroup}")
    print(f"  Written: {out}")


if __name__ == "__main__":
    main()
