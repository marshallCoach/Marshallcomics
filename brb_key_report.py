#!/usr/bin/env python3
"""
brb_key_report.py — standing ordered list of key issues above a value threshold.

Roberto wants this re-runnable as the data tightens, so it derives everything
live from the current canonical xlsx + eBay results. READ-ONLY: it never touches
the Key Issue? flag (see KEY_CLASSIFICATION.md rule 2 — key flags are never
batch-changed).

Value used, in priority order:
  1. eBay median sold  (real market data, most trustworthy)
  2. eBay average sold
  3. Est. Raw Value (NM) $  (the flat import bucket — least trustworthy)
The source is shown per row so a $400 bucket-guess is never mistaken for a
$400 comp.

Also flags, per KEY_CLASSIFICATION.md:
  - signed books (Tier 1 #4 — ALWAYS keys, at any value)
  - CGC-graded / CGC-bound (Tier 1 #5)
  - keys with no value data at all (can't be ranked — surfaced separately)

Usage:
    python3 brb_key_report.py                 # >= $8, console + xlsx
    python3 brb_key_report.py --min 25
    python3 brb_key_report.py --out keys.xlsx
"""
import argparse, glob, json, os, re
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def num(v):
    try:
        return float(str(v).strip().replace("$", "").replace(",", ""))
    except (ValueError, TypeError, AttributeError):
        return None


def is_yes(v):
    return str(v).strip().upper() in ("YES", "Y", "TRUE", "1")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=None)
    ap.add_argument("--min", type=float, default=8.0)
    ap.add_argument("--out", default="key_issues_ranked.xlsx")
    ap.add_argument("--top", type=int, default=40, help="how many to print to console")
    args = ap.parse_args()

    xlsx = os.path.abspath(args.xlsx) if args.xlsx else latest_xlsx()
    print(f"Source: {os.path.basename(xlsx)}   threshold: ${args.min:g}")

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("✅ Clean Inventory"))
    rows = list(ws.iter_rows(values_only=True))
    H = list(rows[0])
    C = {n: H.index(n) for n in H if n}

    def col(row, name):
        i = C.get(name)
        return row[i] if i is not None else None

    keys, unpriced, signed_low = [], [], []
    for r in rows[1:]:
        if not is_yes(col(r, "Key Issue?")):
            continue
        med, avg = num(col(r, "eBay Median Sold $")), num(col(r, "eBay Avg Sold $"))
        nm = num(col(r, "Est. Raw Value (NM) $"))
        if med:
            val, src = med, "eBay median"
        elif avg:
            val, src = avg, "eBay avg"
        elif nm:
            val, src = nm, "NM estimate"
        else:
            val, src = None, "none"
        signed = is_yes(col(r, "Signed?"))
        box = str(col(r, "Box #") or "")
        cgc = "CGC" in box.upper() or is_yes(col(r, "CGC Worth It?"))
        rec = dict(title=str(col(r, "Title") or "").strip(), issue=str(col(r, "Issue #") or ""),
                   year=str(col(r, "Year") or ""), vol=col(r, "Volume"), box=box,
                   val=val, src=src, signed=signed, cgc=cgc,
                   why=str(col(r, "Key Issue — Why") or "").strip()[:80],
                   comps=col(r, "eBay Comp Count"))
        if val is None:
            unpriced.append(rec)
        elif val >= args.min:
            keys.append(rec)
        elif signed:
            signed_low.append(rec)      # Tier 1 #4: signed is always a key, whatever the value

    keys.sort(key=lambda x: -x["val"])
    print(f"\n  Keys at/above ${args.min:g}: {len(keys)}")
    print(f"  Keys below threshold but SIGNED (Tier 1 #4 — still keys): {len(signed_low)}")
    print(f"  Keys with NO value data (can't rank): {len(unpriced)}")
    tot = sum(k["val"] for k in keys)
    print(f"  Total value of ranked keys: ${tot:,.0f}")

    print(f"\n  TOP {min(args.top, len(keys))}:")
    print(f"  {'#':>3}  {'Value':>9}  {'Source':<12} {'Box':<6} Title")
    for i, k in enumerate(keys[:args.top], 1):
        flag = ("✍" if k["signed"] else " ") + ("C" if k["cgc"] else " ")
        print(f"  {i:>3}  ${k['val']:>8,.2f}  {k['src']:<12} {k['box'][:6]:<6} {flag} {k['title'][:38]} #{k['issue']}")

    out = openpyxl.Workbook()
    sh = out.active; sh.title = f"Keys over ${args.min:g}"
    hdr = ["Rank", "Value", "Value source", "eBay comps", "Title", "Issue", "Year", "Vol", "Box", "Signed", "CGC", "Key Issue — Why"]
    sh.append(hdr)
    from openpyxl.styles import Font, PatternFill
    for c in sh[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2F5496")
    sh.freeze_panes = "A2"
    for i, w in enumerate((6, 11, 13, 11, 36, 8, 11, 6, 8, 8, 6, 52), 1):
        sh.column_dimensions[chr(64 + i)].width = w
    for i, k in enumerate(keys, 1):
        sh.append([i, round(k["val"], 2), k["src"], k["comps"], k["title"], k["issue"], k["year"],
                   k["vol"], k["box"], "YES" if k["signed"] else "", "YES" if k["cgc"] else "", k["why"]])
    for name, data in (("Signed under threshold", signed_low), ("Keys with no value", unpriced)):
        s2 = out.create_sheet(name[:31]); s2.append(hdr)
        for c in s2[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="7F7F7F")
        s2.freeze_panes = "A2"
        for i, w in enumerate((6, 11, 13, 11, 36, 8, 11, 6, 8, 8, 6, 52), 1):
            s2.column_dimensions[chr(64 + i)].width = w
        for i, k in enumerate(data, 1):
            s2.append([i, k["val"], k["src"], k["comps"], k["title"], k["issue"], k["year"],
                       k["vol"], k["box"], "YES" if k["signed"] else "", "YES" if k["cgc"] else "", k["why"]])
    out.save(args.out)
    print(f"\n  Written: {args.out}")


if __name__ == "__main__":
    main()
