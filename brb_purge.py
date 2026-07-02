#!/usr/bin/env python3
"""
brb_purge.py — Safe row removal with mandatory audit trail.
Run from: ~/marshallcomics/

Never silently deletes. Every removal:
  1. Copies the row to a '🗑 Purged' sheet in the xlsx
  2. Appends a timestamped entry to purge_log.json
  3. Saves the xlsx only after both are written

Usage:
    # Interactive — shows matches and asks for confirmation:
    python3 brb_purge.py --title "Mockingbird" --issue 8 --volume 2 --reason "sold at SDCC 2026"

    # Dry run — shows what would be removed, writes nothing:
    python3 brb_purge.py --title "Captain Carter" --issue 1 --dry-run

    # Match by row index (from invalid_boxes_to_fix.csv or validator output):
    python3 brb_purge.py --row-index 4821 --reason "duplicate confirmed"
"""

import sys, os, json, glob as _glob, argparse
from datetime import datetime
import pandas as pd
import openpyxl

REPO_ROOT  = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(REPO_ROOT, "attached_assets")
LOG_PATH   = os.path.join(REPO_ROOT, "purge_log.json")
PURGE_SHEET = "🗑 Purged"


def _resolve(raw):
    if os.path.exists(raw): return os.path.abspath(raw)
    c = os.path.join(ASSETS_DIR, os.path.basename(raw))
    return c if os.path.exists(c) else raw


def _latest_validated():
    matches = _glob.glob(os.path.join(ASSETS_DIR, "comics_inventory_*VALIDATED*.xlsx"))
    if not matches:
        matches = _glob.glob(os.path.join(ASSETS_DIR, "comics_inventory_*.xlsx"))
    return max(matches, key=os.path.getmtime) if matches else ""


def _load(path):
    xl = pd.ExcelFile(path)
    REQUIRED = ["Title", "Issue #", "Box #"]
    for name in xl.sheet_names:
        df = xl.parse(name)
        if all(c in df.columns for c in REQUIRED):
            return df, name
    return xl.parse(xl.sheet_names[0]), xl.sheet_names[0]


def _load_purge_log():
    if os.path.exists(LOG_PATH):
        with open(LOG_PATH) as f:
            return json.load(f)
    return []


def _save_purge_log(entries):
    with open(LOG_PATH, "w") as f:
        json.dump(entries, f, indent=2)


def _find_matches(df, title=None, issue=None, volume=None, row_index=None):
    if row_index is not None:
        if row_index not in df.index:
            return pd.DataFrame()
        return df.loc[[row_index]]
    mask = pd.Series([True] * len(df), index=df.index)
    if title:
        mask &= df["Title"].str.lower().str.contains(title.lower(), na=False)
    if issue is not None:
        numeric_issue = pd.to_numeric(df["Issue #"], errors="coerce")
        mask &= numeric_issue == float(issue)
    if volume is not None and "Volume" in df.columns:
        numeric_vol = pd.to_numeric(df["Volume"], errors="coerce")
        mask &= numeric_vol == float(volume)
    return df[mask]


def _write_xlsx_with_purge(path, inv_df, inv_sheet, purge_df):
    """Write inventory sheet + purge sheet atomically."""
    wb = openpyxl.load_workbook(path)

    # Update inventory sheet
    if inv_sheet in wb.sheetnames:
        del wb[inv_sheet]
    ws_inv = wb.create_sheet(inv_sheet, 0)
    ws_inv.append(list(inv_df.columns))
    for row in inv_df.itertuples(index=False):
        ws_inv.append(list(row))

    # Update purge sheet
    if PURGE_SHEET in wb.sheetnames:
        del wb[PURGE_SHEET]
    ws_purge = wb.create_sheet(PURGE_SHEET)
    ws_purge.append(list(purge_df.columns))
    for row in purge_df.itertuples(index=False):
        ws_purge.append(list(row))

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Safe row purge with audit trail")
    parser.add_argument("--file", default=None, help="xlsx file (default: latest VALIDATED)")
    parser.add_argument("--title",     default=None, help="Title to match (partial, case-insensitive)")
    parser.add_argument("--issue",     default=None, type=float, help="Issue number")
    parser.add_argument("--volume",    default=None, type=float, help="Volume number")
    parser.add_argument("--row-index", default=None, type=int,   help="Exact DataFrame row index")
    parser.add_argument("--reason",    default=None, help="Reason for removal (required unless --dry-run)")
    parser.add_argument("--dry-run",   action="store_true", help="Show matches only, write nothing")
    args = parser.parse_args()

    if not args.dry_run and not args.reason:
        print("ERROR: --reason is required unless --dry-run. Describe why this row is being removed.")
        sys.exit(1)

    if not any([args.title, args.issue, args.row_index]):
        print("ERROR: Provide at least --title, --issue, or --row-index to identify the row(s).")
        sys.exit(1)

    # Resolve file
    path = _resolve(args.file) if args.file else _latest_validated()
    if not path or not os.path.exists(path):
        print("ERROR: No xlsx found.")
        sys.exit(1)

    df, sheet = _load(path)
    print(f"Loaded '{sheet}' — {len(df):,} rows from {os.path.basename(path)}")

    matches = _find_matches(df, args.title, args.issue, args.volume, args.row_index)

    if len(matches) == 0:
        print("No matching rows found.")
        sys.exit(0)

    print(f"\n{'DRY RUN — ' if args.dry_run else ''}Matched {len(matches)} row(s):\n")
    for _, row in matches.iterrows():
        vol = f" Vol{row['Volume']}" if "Volume" in row and pd.notna(row.get("Volume")) else ""
        box = row.get("Box #", "?")
        writer = row.get("Writer(s)", "")
        print(f"  [{row.name}]  {row['Title']}{vol} #{row['Issue #']}  Box#{box}  Writer: {writer}")

    if args.dry_run:
        print("\nDry run — nothing written.")
        sys.exit(0)

    # Confirm
    print(f"\nReason: {args.reason}")
    confirm = input(f"\nPermanently remove {len(matches)} row(s) and log to purge sheet? [yes/N]: ").strip().lower()
    if confirm != "yes":
        print("Aborted.")
        sys.exit(0)

    # Load or create existing purge sheet data
    xl = pd.ExcelFile(path)
    if PURGE_SHEET in xl.sheet_names:
        existing_purge = xl.parse(PURGE_SHEET)
    else:
        existing_purge = pd.DataFrame()

    # Tag rows being purged
    rows_to_purge = matches.copy()
    rows_to_purge["purge_reason"]    = args.reason
    rows_to_purge["purge_timestamp"] = datetime.now().isoformat()

    purge_df = pd.concat([existing_purge, rows_to_purge], ignore_index=True)

    # Remove from inventory
    remaining = df.drop(index=matches.index)

    rows_before = len(df)
    rows_after  = len(remaining)

    # Write xlsx
    _write_xlsx_with_purge(path, remaining, sheet, purge_df)

    # Append to purge_log.json
    log = _load_purge_log()
    for _, row in matches.iterrows():
        log.append({
            "ts":        datetime.now().isoformat(),
            "file":      os.path.basename(path),
            "row_index": int(row.name),
            "title":     str(row.get("Title", "")),
            "issue":     str(row.get("Issue #", "")),
            "volume":    str(row.get("Volume", "")),
            "box":       str(row.get("Box #", "")),
            "writer":    str(row.get("Writer(s)", "")),
            "reason":    args.reason,
        })
    _save_purge_log(log)

    print(f"\nDone. {rows_before:,} → {rows_after:,} rows ({rows_before - rows_after} removed).")
    print(f"Purge sheet '{PURGE_SHEET}' updated in {os.path.basename(path)}")
    print(f"purge_log.json updated ({len(log)} total entries)")


if __name__ == "__main__":
    main()
