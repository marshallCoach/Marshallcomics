#!/usr/bin/env python3
"""Fill Writer(s)/Artist(s)/Cover Artist for 11 residual books, from
owner-confirmed source pages (imagecomics / DC / Marvel / fandom infoboxes).
Content-keyed (Title + Issue [+ Year/Box where needed]) so it survives the
weekly reingest's row shifts. BLANK-ONLY: never overwrites an existing credit,
so filled twin copies and the 2009 JLA / 2011 JLA are left untouched. Flattens
formulas, writes a new timestamped canonical, contamination-checks that only
the three credit columns changed.

Held back on purpose (not written):
  - Star Trek: Lore War #1 Artist — the source put the writer (C. Cantwell) in
    the artist slot; needs a real interior-artist source. Cover is written.
"""
import openpyxl, glob, os, datetime

SRC = max((f for f in glob.glob("attached_assets/comics_inventory_*.xlsx")
           if " copy" not in f and not os.path.basename(f).startswith("~$")),
          key=os.path.getmtime)

def ni(v): return str(v or "").strip().replace(".0", "")

# each: titles (exact) or contains; iss list; optional year/box; w/a/c (None=skip)
CREDITS = [
 {"titles":["Icon & Rocket: Season One"], "iss":[str(i) for i in range(1,7)],
  "w":"Reginald Hudlin", "a":"Doug Braithwaite", "c":"Taurin Clarke"},
 {"titles":["Mask","M.A.S.K."], "iss":["3"], "box":"103",
  "w":"Dan Watters", "a":"Pierluigi Casolino", "c":"Pye Parr"},
 {"titles":["Doom Patrol and Suicide Squad Special"], "iss":["1"],
  "w":"Paul Kupperberg & John Ostrander", "a":"Erik Larsen", "c":"Erik Larsen"},
 {"titles":["Justice League of America 80-Page Giant"], "iss":["1"], "year":"2011",
  "w":"Adam Glass", "a":"Dennis Calero", "c":"Stanley Lau"},
 {"contains":"Fury of Firestorm", "iss":["1"], "year":"2011",
  "w":"Ethan Van Sciver & Gail Simone", "a":"Yildiray Cinar", "c":"Ethan Van Sciver"},
 {"titles":["What If? Magik"], "iss":["1"],
  "w":"Leah Williams", "a":"Filipe Andrade", "c":"Jeff Dekal"},
 {"titles":["Stargate Universe"], "iss":["2","3","4","5"],
  "w":"Mark L. Haynes & J. C. Vaughn", "a":"Eliseu Gouveia & Greg LaRocque", "c":None},
 {"titles":["Years of Future Past"], "iss":["1","2"],
  "w":"Marguerite Bennett", "a":"Mike Norton", "c":None},
 {"titles":["Star Trek: Lore War"], "iss":["1"],
  "w":None, "a":"Davide Tinto", "c":"Davide Tinto"}, # Tinto did interior + cover
 {"titles":["Ultraman x The Avengers","Ultraman X The Avengers"], "iss":["1"],
  "w":"Kyle Higgins & Mat Groom", "a":None, "c":None},
 {"titles":["Vision and the Scarlet Witch"], "iss":["1"],
  "w":None, "a":"Lorenzo Tammetta", "c":"Russell Dauterman"},
]

srcvals = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory"))
cached = [list(r) for r in srcvals.iter_rows(values_only=True)]

wb = openpyxl.load_workbook(SRC)
ws = next(w for w in wb.worksheets if w.title.startswith("✅ Clean Inventory"))
for ri, row in enumerate(ws.iter_rows()):
    for cj, c in enumerate(row):
        if isinstance(c.value, str) and c.value.startswith("=") and ri < len(cached) and cj < len(cached[ri]):
            c.value = cached[ri][cj]

H = [c.value for c in ws[1]]; C = {h: i for i, h in enumerate(H) if h}
def col(n): return C[n] + 1
TCOL, ICOL, YCOL, BCOL = col("Title"), col("Issue #"), col("Year"), col("Box #")
WCOL, ACOL, CCOL = col("Writer(s)"), col("Artist(s)"), col("Cover Artist")

def matches(spec, t, i, y, b):
    if i not in spec["iss"]: return False
    if "year" in spec and ni(y) != spec["year"]: return False
    if "box"  in spec and str(b or "").strip() != spec["box"]: return False
    if "titles" in spec and t in spec["titles"]: return True
    if "contains" in spec and spec["contains"].lower() in t.lower(): return True
    return False

set_log, skip_log = [], []
for r in range(2, ws.max_row + 1):
    t = str(ws.cell(r, TCOL).value or "").strip()
    i = ni(ws.cell(r, ICOL).value)
    y = ws.cell(r, YCOL).value; b = ws.cell(r, BCOL).value
    for spec in CREDITS:
        if not matches(spec, t, i, y, b): continue
        for coln, key in ((WCOL,"w"), (ACOL,"a"), (CCOL,"c")):
            val = spec.get(key)
            if not val: continue
            cur = str(ws.cell(r, coln).value or "").strip()
            if cur:
                skip_log.append((r, t, i, key, cur))
            else:
                ws.cell(r, coln, val); set_log.append((r, t, i, key, val))
        break

if not set_log:
    print("Nothing to write (all target fields already filled?). No file written.")
    raise SystemExit(0)

out = f"attached_assets/comics_inventory_{datetime.datetime.now():%d%m_%H%M}.xlsx"
wb.save(out)

so = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
          if w.title.startswith("✅ Clean Inventory"))
no = next(w for w in openpyxl.load_workbook(out, read_only=True, data_only=True).worksheets
          if w.title.startswith("✅ Clean Inventory"))
allowed = {C["Writer(s)"], C["Artist(s)"], C["Cover Artist"]}
off = sum(1 for a, b in zip(so.iter_rows(values_only=True), no.iter_rows(values_only=True))
          for j in range(len(a)) if j not in allowed and (a[j] or "") != (b[j] or ""))
print(f"SOURCE: {os.path.basename(SRC)}")
print(f"OUTPUT: {os.path.basename(out)}")
print(f"fields written: {len(set_log)}   skipped (already filled): {len(skip_log)}")
for r, t, i, k, v in set_log: print(f"   set  r{r} {t[:32]:32} #{i:>2} {k}={v}")
for r, t, i, k, v in skip_log: print(f"   keep r{r} {t[:32]:32} #{i:>2} {k} (had: {v[:30]})")
print(f"contamination (non-credit cells changed): {off}")
