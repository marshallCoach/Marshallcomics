#!/usr/bin/env python3
"""
brb_gcd_year_fix.py — replace SERIES-RANGE Year values ("1997-1998") with the
actual publication year of that specific issue, from GCD's per-issue dates.

Why the inverted lookup: resolving the series first (find_series) picks the
wrong volume for long-running titles — e.g. "The Flash" 1997 resolved to a
2002-2008 series. Instead this searches EVERY GCD series whose name matches the
title, collects the publication year of every issue carrying that issue number,
and uses the declared range as a FILTER. If exactly one candidate year falls
inside the declared range, that's the answer; anything else is left alone.

Only touches rows whose Year is a YYYY-YYYY range. Single-year rows, blanks and
anything ambiguous are never modified. New-file-only, dry-run by default.

Usage:
    python3 brb_gcd_year_fix.py                 # dry-run on newest xlsx
    python3 brb_gcd_year_fix.py --apply
    python3 brb_gcd_year_fix.py <xlsx> --apply --out <file>
"""
import argparse, glob, os, re, sqlite3, datetime
from collections import defaultdict
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
DB = os.path.join(ROOT, "gcd_local.sqlite")
TOLERANCE = 1  # cover dates run ahead of release; allow +/- 1 yr at the edges

RANGE_RE = re.compile(r"\d{4}\s*[-–—]\s*\d{4}")


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def tight(t):
    return re.sub(r"[^a-z0-9]", "", re.sub(r"&", " and ", str(t or "").lower()))


def norm_issue(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def year_span(v):
    n = [int(x) for x in re.findall(r"\d{4}", str(v or "")) if 1900 < int(x) < 2100]
    return (min(n), max(n)) if n else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=None)
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    xlsx = os.path.abspath(args.xlsx) if args.xlsx else latest_xlsx()
    print(f"Source: {os.path.basename(xlsx)}   (mode: {'APPLY' if args.apply else 'DRY-RUN'})")

    conn = sqlite3.connect(DB)
    by_name = defaultdict(list)
    for sid, name in conn.execute("SELECT id, name FROM gcd_series"):
        by_name[tight(name)].append(sid)
    # issue-number -> years, per series, built lazily per title
    issue_cache = {}

    def candidate_years(title, inum):
        key = tight(title)
        if key not in issue_cache:
            idx = defaultdict(set)
            for sid in by_name.get(key, []):
                for num, kd, pd in conn.execute(
                        "SELECT number, key_date, publication_date FROM gcd_issue WHERE series_id = ?", (sid,)):
                    m = re.search(r"(\d{4})", str(kd or pd or ""))
                    if m:
                        idx[norm_issue(num)].add(int(m.group(1)))
            issue_cache[key] = idx
        return issue_cache[key].get(inum, set())

    wb = openpyxl.load_workbook(xlsx)
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("✅ Clean Inventory"))
    H = [c.value for c in next(ws.iter_rows(max_row=1))]
    ti, ii, yi = (H.index(c) for c in ("Title", "Issue #", "Year"))

    fixed = ambiguous = unresolved = 0
    samples = []
    for row in ws.iter_rows(min_row=2):
        raw = str(row[yi].value or "")
        if not RANGE_RE.search(raw):
            continue
        span = year_span(raw)
        years = candidate_years(str(row[ti].value or "").strip(), norm_issue(row[ii].value))
        inwin = {y for y in years if span and span[0] - TOLERANCE <= y <= span[1] + TOLERANCE}
        if len(inwin) == 1:
            y = inwin.pop()
            if args.apply:
                row[yi].value = y
            fixed += 1
            if len(samples) < 12:
                samples.append((str(row[ti].value), norm_issue(row[ii].value), raw, y))
        elif inwin:
            ambiguous += 1
        else:
            unresolved += 1

    print(f"\n  Year ranges resolved to a single year: {fixed}")
    print(f"  Ambiguous (several candidate years in range, left alone): {ambiguous}")
    print(f"  Unresolved (issue/date absent from GCD subset, left alone): {unresolved}")
    print("\n  samples:")
    for t, i, old, new in samples:
        print(f"    {t} #{i}: {old!r} -> {new}")

    if not args.apply:
        print("\n[DRY RUN] No file written.")
        return

    stamp = datetime.datetime.now().strftime("%d%m_%H%M")
    wb["📝 Data Integrity Log"].append([stamp, "Year ranges -> exact year (GCD)",
        f"{fixed} rows whose Year held a series RANGE replaced with that issue's actual publication "
        f"year from GCD per-issue key_date/publication_date. {ambiguous} ambiguous and {unresolved} "
        f"not-in-GCD rows left unchanged. Method: match issue number across all same-named GCD "
        f"series, filter candidates by the declared range (+/-1yr)."])
    out = args.out or os.path.join(os.path.dirname(xlsx), f"comics_inventory_{stamp}.xlsx")
    wb.save(out)
    print(f"\n  Written: {out}")


if __name__ == "__main__":
    main()
