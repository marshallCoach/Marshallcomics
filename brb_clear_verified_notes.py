#!/usr/bin/env python3
"""Clear the 'check title' notes on rows whose title is confirmed by GCD, using
the CORRECT openpyxl clearing idiom (cell.value = None; passing None to
ws.cell(r,c,None) is a getter and silently no-ops). Also clears the now-stale
Verify-Duplicate flag on the two kept twins. Non-destructive: flattens formulas,
writes a new timestamped canonical, confirms only Issue Note / Verify Duplicate
changed and the count of 'check title' notes dropped as expected."""
import openpyxl, glob, os, re, sqlite3, datetime

SRC = max((f for f in glob.glob("attached_assets/comics_inventory_*.xlsx")
           if " copy" not in f and not os.path.basename(f).startswith("~$")),
          key=os.path.getmtime)

def norm(s): return re.sub(r"[^a-z0-9]+", " ", str(s).lower()).strip()
gcd = set(norm(n) for (n,) in sqlite3.connect("gcd_local.sqlite").execute("select name from gcd_series"))

srcvals = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory"))
cached = [list(r) for r in srcvals.iter_rows(values_only=True)]

wb = openpyxl.load_workbook(SRC)
ws = next(w for w in wb.worksheets if w.title.startswith("✅ Clean Inventory"))
for ri, row in enumerate(ws.iter_rows()):
    for cj, c in enumerate(row):
        if isinstance(c.value, str) and c.value.startswith("=") and ri < len(cached) and cj < len(cached[ri]):
            c.value = cached[ri][cj]

H = [c.value for c in ws[1]]; C = {h: i for i, h in enumerate(H) if h}
TCOL, NCOL, VDCOL = C["Title"]+1, C["Issue Note"]+1, C["⚠ Verify Duplicate"]+1
def sval(r, col):
    v = ws.cell(r, col).value
    return "" if v is None else str(v).strip()

cleared = 0; kept = []
for r in range(2, ws.max_row + 1):
    note = sval(r, NCOL); nl = note.lower()
    if "check title" not in nl:
        continue
    if "cover" in nl or "dupe" in nl:            # keep cover/dupe flags
        continue
    if norm(sval(r, TCOL)) in gcd:
        ws.cell(r, NCOL).value = None            # CORRECT clear
        cleared += 1
    else:
        kept.append((r, sval(r, TCOL)))

# clear stale Verify-Duplicate on the two kept twins (dup partner removed)
vd_cleared = 0
for r in range(2, ws.max_row + 1):
    key = (sval(r, TCOL), sval(r, C["Issue #"]+1), sval(r, C["Year"]+1))
    if key in {("Aliens vs. Avengers","1","2024"), ("Fantastic Four: Empyre","0","2020")} and sval(r, VDCOL):
        ws.cell(r, VDCOL).value = None; vd_cleared += 1

out = f"attached_assets/comics_inventory_{datetime.datetime.now():%d%m_%H%M}.xlsx"
wb.save(out)

# verify
no = next(w for w in openpyxl.load_workbook(out, read_only=True, data_only=True).worksheets
          if w.title.startswith("✅ Clean Inventory"))
nr = list(no.iter_rows(values_only=True)); Ci = {h:i for i,h in enumerate(nr[0]) if h}
remain = sum(1 for r in nr[1:] if r[Ci["Issue Note"]] and "check title" in str(r[Ci["Issue Note"]]).lower())
# contamination: only Issue Note / Verify Duplicate may differ
so = list(srcvals.iter_rows(values_only=True))
allowed = {C["Issue Note"], C["⚠ Verify Duplicate"]}
off = sum(1 for a,b in zip(so, nr) for j in range(len(a)) if j not in allowed and (a[j] or "")!=(b[j] or ""))
print(f"SOURCE: {os.path.basename(SRC)}\nOUTPUT: {os.path.basename(out)}")
print(f"notes cleared (GCD-verified): {cleared}   Verify-Duplicate flags cleared: {vd_cleared}")
print(f"'check title' notes remaining: {remain}   (kept: mismatch/no-GCD + cover/dupe)")
print(f"CONTAMINATION (non Note/VDup diffs): {off}  (must be 0)")
