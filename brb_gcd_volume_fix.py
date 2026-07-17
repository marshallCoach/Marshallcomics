#!/usr/bin/env python3
"""
brb_gcd_volume_fix.py — apply GCD-derived Volume numbers to the inventory,
restricted to the two independently-verified high-confidence slices from
brb_gcd_volume_check.py:

  forward: declared Volume is 1 (the default-fill value) but GCD says the
           book belongs to a later same-name series
  reverse: declared Volume is HIGHER than GCD's derived number (the
           flat-per-title over-numbering pattern, e.g. one-shot minis
           declared Vol 5-7)

Both directions are gated on GCD_volume_count_for_title <= 3 — where GCD has
few same-name series for the title+publisher, the counting-convention gap
(GCD numbering every special/facsimile) cannot manufacture a false fix.
Rows with count > 3 are NEVER touched regardless of direction.

Safety (same pattern as brb_gcd_fill.py):
  - dry-run by default; --apply writes a NEW *_VOLUME_FIXED.xlsx, never
    the source in place
  - only rows that resolve to exactly ONE GCD series by year overlap
  - prints a per-title summary and flags Absolute Batman (three separate
    data corrections have converged on that title this month)

Usage:
    python3 brb_gcd_volume_fix.py                # dry-run on newest xlsx
    python3 brb_gcd_volume_fix.py --apply        # write *_VOLUME_FIXED.xlsx
    python3 brb_gcd_volume_fix.py <xlsx> --apply --out <file>
"""
import argparse, os, re, sqlite3, sys
from collections import defaultdict, Counter
import openpyxl

from brb_gcd_volume_check import tight, pub_match, parse_year_range, latest_xlsx, DB

MAX_SERIES_COUNT = 3  # convention-noise gate — never touch titles above this
WATCH_TITLES = {"Absolute Batman"}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=None)
    ap.add_argument("--apply", action="store_true", help="write the fixed file (default is dry-run)")
    ap.add_argument("--out", default=None)
    ap.add_argument("--approved", default=None, metavar="CSV",
                    help="CSV of Title,Decision rows (human-reviewed). ONLY titles marked 'y' "
                         "are processed, with the <=3-series gate and direction rules bypassed: "
                         "every mismatching row of an approved title moves to the GCD-derived number.")
    args = ap.parse_args()

    approved = None
    if args.approved:
        import csv as _csv
        with open(args.approved) as f:
            approved = {r["Title"] for r in _csv.DictReader(f) if r["Decision"].strip().lower() == "y"}
        print(f"Approved-titles mode: {len(approved)} titles from {os.path.basename(args.approved)}")

    xlsx = os.path.abspath(args.xlsx) if args.xlsx else latest_xlsx()
    print(f"Source: {os.path.basename(xlsx)}   (mode: {'APPLY' if args.apply else 'DRY-RUN'})")
    conn = sqlite3.connect(DB)

    # series pools merged across spelling variants + aliases (same as the check)
    gcd_by_key = defaultdict(dict)
    for sid, mt, name, yb, ye, pub in conn.execute(
        "SELECT s.id, s.matched_title, s.name, s.year_began, s.year_ended, p.name "
        "FROM gcd_series s LEFT JOIN gcd_publisher p ON p.id = s.publisher_id"):
        gcd_by_key[tight(mt)][sid] = {"name": name, "yb": yb, "ye": ye, "pub": pub}
    try:
        aliases = dict(conn.execute("SELECT alias, matched_title FROM gcd_title_alias"))
    except sqlite3.OperationalError:
        aliases = {}

    def series_for(title):
        pool = dict(gcd_by_key.get(tight(title), {}))
        if title in aliases:
            pool.update(gcd_by_key.get(tight(aliases[title]), {}))
        return list(pool.values())

    wb = openpyxl.load_workbook(xlsx)
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("✅ Clean Inventory"))
    H = [c.value for c in next(ws.iter_rows(max_row=1))]
    ti, ii, yi, vi, pi, bi = (H.index(c) for c in ("Title", "Issue #", "Year", "Volume", "Publisher", "Box #"))

    numbering_cache = {}
    fixed_fwd = fixed_rev = 0
    fixes = []
    for row in ws.iter_rows(min_row=2):
        title = str(row[ti].value or "").strip()
        if not title:
            continue
        try:
            declared_n = int(float(row[vi].value))
        except (ValueError, TypeError):
            continue

        pub = str(row[pi].value or "").strip()
        ck = (title, pub)
        if ck not in numbering_cache:
            cands = [s for s in series_for(title) if pub_match(pub, s["pub"]) and s["yb"]]
            by_year = {}
            for s in sorted(cands, key=lambda s: (s["yb"], s["name"])):
                by_year.setdefault(s["yb"], s)
            numbering_cache[ck] = sorted(by_year.values(), key=lambda s: s["yb"])
        ordered = numbering_cache[ck]
        if not ordered:
            continue
        if approved is not None:
            if title not in approved:
                continue
        elif len(ordered) > MAX_SERIES_COUNT:
            continue

        yr = parse_year_range(row[yi].value)
        if not yr:
            continue
        hits = [i for i, s in enumerate(ordered) if s["yb"] <= yr[1] and (s["ye"] or 2100) >= yr[0]]
        if len(hits) != 1:
            continue
        derived_n = hits[0] + 1
        if derived_n == declared_n:
            continue

        if declared_n == 1 and derived_n > 1:
            direction = "forward"
            fixed_fwd += 1
        elif declared_n > derived_n:
            direction = "reverse"
            fixed_rev += 1
        elif approved is not None:
            direction = "approved"  # declared < derived, non-default — allowed only on reviewed titles
            fixed_fwd += 1
        else:
            continue  # declared < derived but not the default-1 pattern — out of scope

        if args.apply:
            row[vi].value = derived_n
        s = ordered[hits[0]]
        fixes.append((title, str(row[ii].value), str(row[bi].value), declared_n, derived_n,
                      f"{s['name']} ({s['yb']}-{s['ye'] or '?'})", direction))

    total = fixed_fwd + fixed_rev
    print(f"\nVolume fixes in scope: {total}  (forward: {fixed_fwd}, reverse: {fixed_rev})")

    by_title = Counter((t, d, g) for t, _, _, d, g, _, _ in fixes)
    print(f"\n-- per-title summary (top 25) --")
    for (t, d, g), n in by_title.most_common(25):
        print(f"  {n:>3}  {t}: Vol {d} -> {g}")

    watch = [f for f in fixes if f[0] in WATCH_TITLES]
    if watch:
        print(f"\n⚠ WATCH-TITLE rows (converged corrections — double-check these by hand):")
        for t, iss, box, d, g, series, direction in watch:
            print(f"  {t} #{iss} (Box {box}): Vol {d} -> {g}  [{series}, {direction}]")

    if not args.apply:
        print("\n[DRY RUN] No file written. Re-run with --apply to write.")
        return

    out = args.out or xlsx.replace(".xlsx", "_VOLUME_FIXED.xlsx")
    wb.save(out)
    print(f"\nWritten: {out}")


if __name__ == "__main__":
    main()
