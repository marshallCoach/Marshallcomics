#!/usr/bin/env python3
"""
brb_gcd_fill.py — fill blank Writer(s)/Artist(s) from the local GCD subset.

Offline companion to the Comic Vine overnight fill (run_overnight_v2.py): no
API, no rate limit — queries gcd_local.sqlite (built by gcd_build_local_db.py
+ gcd_add_credits.py). Targets the ~637 rows Comic Vine genuinely can't fill,
and is the right source for volume-collision titles CV mis-matched (GCD's
Publisher+Year series resolution is more reliable there).

Safety, matching the hard lessons from the CV run:
  - Fills ONLY blank cells (never overwrites an existing credit — the exact
    guard whose absence let the CV run clobber correct writers).
  - Never writes the source xlsx in place — always a new *_GCD_FILLED.xlsx.
  - Only writes when GCD resolved the series by Title AND the issue exists in
    it; ambiguous/not-found rows are left blank and reported, not guessed.

Usage:
    python3 brb_gcd_fill.py                 # newest xlsx in attached_assets/, dry-run
    python3 brb_gcd_fill.py --apply         # write the *_GCD_FILLED.xlsx
    python3 brb_gcd_fill.py <xlsx> --apply --out <file>
"""
import argparse, glob, os, sys
import openpyxl
from brb_gcd_lookup import lookup

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def is_blank(v):
    return v is None or str(v).strip() in ("", "nan", "None")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=None)
    ap.add_argument("--apply", action="store_true", help="write the filled file (default is dry-run)")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    xlsx = os.path.abspath(args.xlsx) if args.xlsx else latest_xlsx()
    if not xlsx or not os.path.exists(xlsx):
        print(f"ERROR: no xlsx found in {ASSETS}"); sys.exit(1)
    if not os.path.exists(os.path.join(ROOT, "gcd_local.sqlite")):
        print("ERROR: gcd_local.sqlite not found — run gcd_build_local_db.py + gcd_add_credits.py first."); sys.exit(1)
    print(f"Source: {os.path.basename(xlsx)}   (mode: {'APPLY' if args.apply else 'DRY-RUN'})")

    wb = openpyxl.load_workbook(xlsx)
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("✅ Clean Inventory"))
    H = [c.value for c in next(ws.iter_rows(max_row=1))]
    ci = {name: H.index(name) + 1 for name in ("Title", "Issue #", "Year", "Publisher", "Writer(s)")}
    ai = H.index("Artist(s)") + 1 if "Artist(s)" in H else None
    cvi = H.index("Cover Artist") + 1 if "Cover Artist" in H else None

    filled_w = filled_a = filled_c = not_found = no_gap = 0
    fills = []
    for row in ws.iter_rows(min_row=2):
        title = str(row[ci["Title"] - 1].value or "").strip()
        if not title:
            continue
        w_blank = is_blank(row[ci["Writer(s)"] - 1].value)
        a_blank = ai is not None and is_blank(row[ai - 1].value)
        c_blank = cvi is not None and is_blank(row[cvi - 1].value)
        if not w_blank and not a_blank and not c_blank:
            no_gap += 1
            continue

        issue = row[ci["Issue #"] - 1].value
        year = row[ci["Year"] - 1].value
        pub = row[ci["Publisher"] - 1].value
        r = lookup(title, issue, year, pub)
        if not r.get("found"):
            not_found += 1
            continue

        cr = r["issue"]
        did = []
        if w_blank and cr.get("writer"):
            if args.apply:
                row[ci["Writer(s)"] - 1].value = cr["writer"]
            filled_w += 1; did.append(f"W={cr['writer']}")
        if a_blank and cr.get("artist"):
            if args.apply:
                row[ai - 1].value = cr["artist"]
            filled_a += 1; did.append(f"A={cr['artist']}")
        if c_blank and cr.get("cover_artist"):
            if args.apply:
                row[cvi - 1].value = cr["cover_artist"]
            filled_c += 1; did.append(f"C={cr['cover_artist']}")
        if did:
            fills.append((title, str(issue), r["series"]["name"], r["series"]["year_began"], " ".join(did)))

    print(f"\nRows with a Writer/Artist/Cover gap that GCD resolved:")
    print(f"  Writers filled:       {filled_w}")
    print(f"  Artists filled:       {filled_a}")
    print(f"  Cover Artists filled: {filled_c}")
    print(f"  Gap rows GCD couldn't resolve (series/issue not found): {not_found}")
    print(f"  Rows with no gap (skipped): {no_gap}")

    print(f"\n-- sample of fills (first 30) --")
    for t, i, sname, sy, did in fills[:30]:
        print(f"  {t} #{i}  [GCD: {sname} ({sy})]  {did}")
    if len(fills) > 30:
        print(f"  ... and {len(fills) - 30} more")

    if not args.apply:
        print("\n[DRY RUN] No file written. Re-run with --apply to write.")
        return

    out = args.out or xlsx.replace(".xlsx", "_GCD_FILLED.xlsx")
    wb.save(out)
    print(f"\nWritten: {out}")


if __name__ == "__main__":
    main()
