#!/usr/bin/env python3
"""Action the review flags in the promoted canonical (Issue Note column).

Non-destructive: writes a new timestamped canonical, flattens formulas from a
data_only read first (so the VF-value formula column isn't nulled), and confirms
only the intended columns changed.

What it does:
  1. Restore "Batgirl and the Birds of Prey: Rebirth" (row's title was shortened,
     colliding with the base series #1 — the Rebirth one-shot is a distinct book).
  2. Mark the two remaining duplicate collisions reviewed (⚠ Verify Duplicate)
     so validation passes; they are listed for the user to confirm keep-vs-remove.
  3. Clear the "check title" note on rows whose title EXACTLY matches a GCD series
     name (title verified) — but keep any note that also mentions an incorrect
     cover or a dupe. Title mismatches / no-GCD-match rows keep their flag.
Everything is logged to FLAGS_ACTION_LOG.md.
"""
import openpyxl, glob, os, re, sqlite3, difflib, collections, datetime

SRC = max((f for f in glob.glob("attached_assets/comics_inventory_*.xlsx")
           if " copy" not in f and not os.path.basename(f).startswith("~$")),
          key=os.path.getmtime)


def norm(s):
    return re.sub(r"[^a-z0-9]+", " ", str(s).lower()).strip()


# --- GCD authoritative series names ---
con = sqlite3.connect("gcd_local.sqlite")
gcd_names = set(norm(n) for (n,) in con.execute("select name from gcd_series"))

# --- flatten formulas via cached values ---
srcvals = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory"))
cached = [list(r) for r in srcvals.iter_rows(values_only=True)]

wb = openpyxl.load_workbook(SRC)
ws = next(w for w in wb.worksheets if w.title.startswith("✅ Clean Inventory"))
for ri, row in enumerate(ws.iter_rows()):
    for cj, cell in enumerate(row):
        if isinstance(cell.value, str) and cell.value.startswith("="):
            if ri < len(cached) and cj < len(cached[ri]):
                cell.value = cached[ri][cj]

H = [c.value for c in ws[1]]
C = {h: i for i, h in enumerate(H) if h}
TCOL = C["Title"] + 1
NCOL = C["Issue Note"] + 1
VDCOL = C["⚠ Verify Duplicate"] + 1


def cell(row, col):
    return ws.cell(row, col).value


def sval(row, col):
    v = cell(row, col)
    return "" if v is None else str(v).strip()


verified, kept_mismatch, kept_nomatch, kept_cover = [], [], [], []

# 1) Restore Batgirl Rebirth title (find the row by its current signature)
for row in range(2, ws.max_row + 1):
    if (sval(row, TCOL) == "Batgirl and the Birds of Prey"
            and sval(row, C["Issue #"] + 1) == "1"
            and sval(row, C["Year"] + 1) == "2016"
            and sval(row, NCOL).lower() == "rebirth"):
        ws.cell(row, TCOL, "Batgirl and the Birds of Prey: Rebirth")
        ws.cell(row, NCOL, None)

# 2) Mark the two remaining collisions reviewed
DUP_REVIEW = [("Aliens vs. Avengers", "1", "2024", "75"),
              ("Fantastic Four: Empyre", "0", "2020", "UNKNOWN — needs physical reassignment")]
dup_marked = []
for row in range(2, ws.max_row + 1):
    key = (sval(row, TCOL), sval(row, C["Issue #"] + 1), sval(row, C["Year"] + 1), sval(row, C["Box #"] + 1))
    if key in DUP_REVIEW and not sval(row, VDCOL):
        ws.cell(row, VDCOL, "⚠ Verify Duplicate")
        dup_marked.append((row, *key))

# 3) Clear verified "check title" notes
for row in range(2, ws.max_row + 1):
    note = sval(row, NCOL)
    nl = note.lower()
    if "check title" not in nl:
        continue
    title = sval(row, TCOL)
    if "cover" in nl or "dupe" in nl:
        kept_cover.append((row, title, note)); continue
    if norm(title) in gcd_names:
        ws.cell(row, NCOL, None)
        verified.append((row, title))
    else:
        # closeness to any GCD name
        best, br = None, 0.0
        nt = norm(title)
        for k in gcd_names:
            r = difflib.SequenceMatcher(None, nt, k).ratio()
            if r > br:
                br, best = r, k
        if best and br >= 0.86:
            kept_mismatch.append((row, title, best, round(br, 2)))
        else:
            kept_nomatch.append((row, title, sval(row, C["Year"] + 1), sval(row, C["Publisher"] + 1)))

out = f"attached_assets/comics_inventory_{datetime.datetime.now():%d%m_%H%M}.xlsx"
wb.save(out)

# --- contamination check: only Title / Issue Note / Verify Duplicate may differ ---
so = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
          if w.title.startswith("✅ Clean Inventory"))
no = next(w for w in openpyxl.load_workbook(out, read_only=True, data_only=True).worksheets
          if w.title.startswith("✅ Clean Inventory"))
allowed = {C["Title"], C["Issue Note"], C["⚠ Verify Duplicate"]}
off = 0
for a, b in zip(so.iter_rows(values_only=True), no.iter_rows(values_only=True)):
    for j in range(len(a)):
        if j in allowed:
            continue
        if (a[j] or "") != (b[j] or ""):
            off += 1

print(f"SOURCE: {os.path.basename(SRC)}")
print(f"OUTPUT: {os.path.basename(out)}")
print(f"verified notes cleared: {len(verified)}")
print(f"kept — title mismatch (needs rename decision): {len(kept_mismatch)}")
print(f"kept — no GCD match: {len(kept_nomatch)}")
print(f"kept — cover/dupe flag: {len(kept_cover)}")
print(f"dup groups marked reviewed: {len(dup_marked)}")
print(f"CONTAMINATION (non-allowed cell diffs): {off}  (must be 0)")

with open("FLAGS_ACTION_LOG.md", "w") as f:
    f.write(f"# Flag action — {datetime.datetime.now():%Y-%m-%d %H:%M}\n\n")
    f.write(f"Source `{os.path.basename(SRC)}` → `{os.path.basename(out)}`\n\n")
    f.write(f"- **{len(verified)}** 'check title' notes cleared (title confirmed against GCD series names)\n")
    f.write(f"- **{len(kept_mismatch)}** kept — GCD suggests a different title (your call)\n")
    f.write(f"- **{len(kept_nomatch)}** kept — no GCD match, could not verify offline\n")
    f.write(f"- **{len(kept_cover)}** kept — flagged for an incorrect cover / dupe\n\n")
    f.write("## Title rename proposals (GCD)\n\n| Row | Current | GCD suggests | Ratio |\n|---|---|---|---|\n")
    for row, t, b, r in sorted(kept_mismatch):
        f.write(f"| {row} | {t} | {b} | {r} |\n")
    f.write("\n## No GCD match — verify manually / via Comic Vine\n\n| Row | Title | Year | Publisher |\n|---|---|---|---|\n")
    for row, t, y, p in sorted(kept_nomatch):
        f.write(f"| {row} | {t} | {y} | {p} |\n")
    f.write("\n## Incorrect-cover / dupe flags kept\n\n| Row | Title | Note |\n|---|---|---|\n")
    for row, t, n in sorted(kept_cover):
        f.write(f"| {row} | {t} | {n} |\n")
    f.write("\n## Duplicate collisions marked reviewed — confirm keep vs remove\n\n| Row | Title | Issue | Year | Box |\n|---|---|---|---|---|\n")
    for row, t, i, y, b in dup_marked:
        f.write(f"| {row} | {t} | {i} | {y} | {b} |\n")
print("Logged: FLAGS_ACTION_LOG.md")
