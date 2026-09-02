#!/usr/bin/env python3
"""Normalize two Check-12 title splits to one spelling per series.
Only the Title column is mutated (these rows carry no 'check title' flag, so
Issue Note is left untouched). Row-keyed with an expected-current-title guard;
flattens formulas; writes a new timestamped canonical; contamination check
confirms only Title changed."""
import openpyxl, glob, os, datetime

SRC = max((f for f in glob.glob("attached_assets/comics_inventory_*.xlsx")
           if " copy" not in f and not os.path.basename(f).startswith("~$")),
          key=os.path.getmtime)

# row -> (expected current title, corrected title)
RENAMES = {
    5539: ("Justice League of America: 80-Page Giant", "Justice League of America 80-Page Giant"),
    9845: ("Ultraman x the Avengers", "Ultraman x The Avengers"),
    9846: ("Ultraman X The Avengers", "Ultraman x The Avengers"),
    9847: ("Ultraman X The Avengers", "Ultraman x The Avengers"),
    9848: ("Ultraman X The Avengers", "Ultraman x The Avengers"),
}

srcvals = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory"))
cached = [list(r) for r in srcvals.iter_rows(values_only=True)]

wb = openpyxl.load_workbook(SRC)
ws = next(w for w in wb.worksheets if w.title.startswith("✅ Clean Inventory"))
for ri, row in enumerate(ws.iter_rows()):
    for cj, c in enumerate(row):
        if isinstance(c.value, str) and c.value.startswith("=") and ri < len(cached) and cj < len(cached[ri]):
            c.value = cached[ri][cj]

H = [c.value for c in ws[1]]; C = {h: i for i, h in enumerate(H) if h}
TCOL = C["Title"] + 1

applied, skipped = [], []
for row, (expect, new) in RENAMES.items():
    cur = str(ws.cell(row, TCOL).value or "").strip()
    if cur != expect:
        skipped.append((row, cur, expect)); continue
    ws.cell(row, TCOL, new)
    applied.append((row, expect, new))

out = f"attached_assets/comics_inventory_{datetime.datetime.now():%d%m_%H%M}.xlsx"
wb.save(out)

so = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
          if w.title.startswith("✅ Clean Inventory"))
no = next(w for w in openpyxl.load_workbook(out, read_only=True, data_only=True).worksheets
          if w.title.startswith("✅ Clean Inventory"))
allowed = {C["Title"]}
off = sum(1 for a, b in zip(so.iter_rows(values_only=True), no.iter_rows(values_only=True))
          for j in range(len(a)) if j not in allowed and (a[j] or "") != (b[j] or ""))
print(f"SOURCE: {os.path.basename(SRC)}\nOUTPUT: {os.path.basename(out)}")
print(f"renames applied: {len(applied)}   skipped (title moved): {len(skipped)}")
for r, e, n in applied: print(f"   row {r}: {e!r} -> {n!r}")
for r, cur, e in skipped: print(f"   SKIP row {r}: found {cur!r}, expected {e!r}")
print(f"contamination (non-Title cells changed): {off}")
