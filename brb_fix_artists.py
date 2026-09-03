#!/usr/bin/env python3
"""Owner-confirmed artist corrections (run after / instead of brb_fill_credits.py,
idempotent either way):
  - Icon & Rocket: Season One #1-6  Artist -> "Doug Braithwaite"  (Taurin Clarke
    was the COVER artist only; interior is Braithwaite). OVERWRITES.
  - Star Trek: Lore War #1          Artist -> "Davide Tinto"      (Tinto did
    interior + cover). SETS.
Only the Artist(s) column changes (contamination check). Content-keyed."""
import openpyxl, glob, os, datetime

SRC = max((f for f in glob.glob("attached_assets/comics_inventory_*.xlsx")
           if " copy" not in f and not os.path.basename(f).startswith("~$")),
          key=os.path.getmtime)
def ni(v): return str(v or "").strip().replace(".0", "")

# (title match, issues, artist value)
FIXES = [
 ("Icon & Rocket: Season One", [str(i) for i in range(1, 7)], "Doug Braithwaite"),
 ("Star Trek: Lore War",       ["1"],                          "Davide Tinto"),
]

srcvals = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory"))
cached = [list(r) for r in srcvals.iter_rows(values_only=True)]
wb = openpyxl.load_workbook(SRC)
ws = next(w for w in wb.worksheets if w.title.startswith("✅ Clean Inventory"))
for ri, row in enumerate(ws.iter_rows()):
    for cj, c in enumerate(row):
        if isinstance(c.value, str) and c.value.startswith("=") and ri < len(cached) and cj < len(cached[ri]):
            c.value = cached[ri][cj]
H = [c.value for c in ws[1]]; C = {h: i for i, h in enumerate(H) if h}
TCOL, ICOL, ACOL = C["Title"]+1, C["Issue #"]+1, C["Artist(s)"]+1

applied = []
for r in range(2, ws.max_row + 1):
    t = str(ws.cell(r, TCOL).value or "").strip(); i = ni(ws.cell(r, ICOL).value)
    for title, issues, artist in FIXES:
        if t == title and i in issues:
            old = str(ws.cell(r, ACOL).value or "").strip()
            if old != artist:
                ws.cell(r, ACOL, artist); applied.append((r, t, i, old, artist))
            break

if not applied:
    print("No change (already correct). Nothing written."); raise SystemExit(0)

out = f"attached_assets/comics_inventory_{datetime.datetime.now():%d%m_%H%M}.xlsx"
wb.save(out)
so = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
          if w.title.startswith("✅ Clean Inventory"))
no = next(w for w in openpyxl.load_workbook(out, read_only=True, data_only=True).worksheets
          if w.title.startswith("✅ Clean Inventory"))
allowed = {C["Artist(s)"]}
off = sum(1 for a, b in zip(so.iter_rows(values_only=True), no.iter_rows(values_only=True))
          for j in range(len(a)) if j not in allowed and (a[j] or "") != (b[j] or ""))
print(f"SOURCE: {os.path.basename(SRC)}\nOUTPUT: {os.path.basename(out)}")
for r, t, i, o, n in applied: print(f"   r{r} {t[:30]:30} #{i:>2} Artist: {o or '(blank)'!r} -> {n!r}")
print(f"contamination (non-Artist cells changed): {off}")
