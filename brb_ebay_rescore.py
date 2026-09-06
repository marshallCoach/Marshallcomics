#!/usr/bin/env python3
"""
brb_ebay_rescore.py — repair eBay medians that landed in the wrong price cluster.

WHY (found 2407, via Roberto checking Fantastic Four x Gargoyles #1 against
keycollectorcomics — listed there at ~$6, our data claimed $222):

The eBay query is `"<title> #<issue> comic <year>"`, which is loose. For titles
built from common words ("Fantastic Four", "X-Men", "Avengers", "JSA") or with
odd glyphs (the multiplication sign in "Fantastic Four x Gargoyles"), eBay
returns a MIX of two different books:

    Fantastic Four x Gargoyles #1 -> [5.19, 5.98, 194.95, 249.56, 299.99, 299.99, 899.99]
                                      ^^^^^^^^^^^^  the real book    ^^^^ Fantastic Four keys

The existing outlier trim drops prices >3x the median. That works when noise is
a minority — but when the WRONG comps outnumber the right ones, the median lands
in the wrong cluster and the trim deletes the LEGITIMATE sales instead. Exactly
backwards.

METHOD — split, then choose by the NM prior:
  1. sort the comps, find the largest multiplicative gap between neighbours
  2. if that gap is >= GAP_RATIO and both sides hold >= MIN_SIDE comps, the set
     is bimodal: two different books got mixed
  3. pick the cluster whose geometric mean is closest (in log space) to the
     book's own Est. Raw Value (NM). That prior is independent of eBay, so it
     breaks the tie without circular reasoning
  4. recompute median/avg/low/high/count from the chosen cluster only

Direction is NOT always "take the cheap one":
    Fantastic Four x Gargoyles  NM $6   -> low cluster  ($5-6)    correct
    JSA #1                      NM $275 -> high cluster ($150+)   correct

Entries that are bimodal but whose NM is missing/zero cannot be adjudicated —
they are marked needs_review instead of guessed.

READ-ONLY on the source: writes a NEW json (never edits in place) plus a review
xlsx. Nothing is silently overwritten; re-run freely.

Usage:
    python3 brb_ebay_rescore.py                 # report only
    python3 brb_ebay_rescore.py --apply         # write rescored json + review sheet
"""
import argparse, json, math, os, shutil
from statistics import median as _median

ROOT = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(ROOT, "ebay_pricing_results.json")
GAP_RATIO = 4.0     # neighbour ratio that counts as a cluster break
MIN_SIDE = 2        # comps required on each side to call it bimodal
MIN_COMPS = 4


def gmean(xs):
    return math.exp(sum(math.log(x) for x in xs) / len(xs))


def split_clusters(ps):
    """Return (low, high) if bimodal, else None."""
    if len(ps) < MIN_COMPS:
        return None
    best = (0.0, -1)
    for i in range(len(ps) - 1):
        if ps[i] > 0:
            r = ps[i + 1] / ps[i]
            if r > best[0]:
                best = (r, i)
    ratio, idx = best
    if ratio < GAP_RATIO:
        return None
    lo, hi = ps[: idx + 1], ps[idx + 1:]
    if len(lo) < MIN_SIDE or len(hi) < MIN_SIDE:
        return None
    return lo, hi


def stats(ps):
    return {"median": round(_median(ps), 2), "avg": round(sum(ps) / len(ps), 2),
            "low": min(ps), "high": max(ps), "count": len(ps)}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--out", default=os.path.join(ROOT, "ebay_pricing_results_rescored.json"))
    ap.add_argument("--review", default=os.path.join(ROOT, "ebay_rescore_review.xlsx"))
    args = ap.parse_args()

    data = json.load(open(SRC))
    fixed, review, unresolved = 0, [], 0
    out = json.loads(json.dumps(data))          # deep copy

    for k, v in data.items():
        ps = sorted(p for p in (v.get("prices") or []) if p and p > 0)
        sp = split_clusters(ps)
        if not sp:
            continue
        lo, hi = sp
        nm = v.get("nm_value") or 0
        old_med = v.get("median") or v.get("avg")
        if not nm:
            unresolved += 1
            out[k]["needs_review"] = True
            out[k]["review_reason"] = "bimodal comps, no NM value to adjudicate"
            review.append((k, nm, old_med, None, lo, hi, "NO NM — cannot choose"))
            continue
        # choose the cluster nearest the NM prior in log space
        pick, other = (lo, hi) if abs(math.log(gmean(lo) / nm)) <= abs(math.log(gmean(hi) / nm)) else (hi, lo)
        st = stats(pick)
        if abs((st["median"] or 0) - (old_med or 0)) < 0.01:
            continue
        out[k].update(st)
        out[k]["prices"] = pick
        out[k]["rescored"] = True
        out[k]["rescore_note"] = (f"bimodal comps split at >={GAP_RATIO}x; kept the cluster nearest "
                                  f"NM ${nm:g}; dropped {len(other)} comps from the other book")
        fixed += 1
        review.append((k, nm, old_med, st["median"], lo, hi,
                       "kept LOW" if pick is lo else "kept HIGH"))

    print(f"Source entries          : {len(data):,}")
    print(f"  bimodal & RESCORED    : {fixed:,}")
    print(f"  bimodal, unresolvable : {unresolved:,}  (no NM prior — flagged needs_review)")

    moved = [(k, o, n) for k, _, o, n, _, _, _ in review if n]
    drops = sorted((((o - n), k, o, n) for k, o, n in moved if o and n and o > n), reverse=True)
    if drops:
        print(f"\n  Largest corrections (inflated -> real):")
        for d, k, o, n in drops[:12]:
            print(f"    {k[:44]:<45} ${o:>8.2f} -> ${n:>7.2f}")
    ups = sorted(((n - o), k, o, n) for k, o, n in moved if o and n and n > o)
    if ups:
        print(f"\n  Corrected UPWARD (median had sat in the cheap contamination):")
        for d, k, o, n in ups[:6]:
            print(f"    {k[:44]:<45} ${o:>8.2f} -> ${n:>7.2f}")

    if not args.apply:
        print("\n[REPORT ONLY] re-run with --apply to write the rescored json + review sheet")
        return

    json.dump(out, open(args.out, "w"))
    print(f"\n  Written: {os.path.basename(args.out)}")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook(); sh = wb.active; sh.title = "Rescored"
        sh.append(["Title|||Issue", "NM $", "old median", "new median", "low cluster", "high cluster", "decision"])
        for c in sh[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="C00000")
        sh.freeze_panes = "A2"
        for i, w in enumerate((44, 8, 12, 12, 34, 40, 26), 1):
            sh.column_dimensions[chr(64 + i)].width = w
        for k, nm, o, n, lo, hi, dec in review:
            sh.append([k, nm, o, n, str([round(p, 2) for p in lo]), str([round(p, 2) for p in hi]), dec])
        wb.save(args.review)
        print(f"  Review sheet: {os.path.basename(args.review)}  ({len(review)} rows)")
    except ImportError:
        print("  (openpyxl missing — skipped review sheet)")


if __name__ == "__main__":
    main()
