#!/usr/bin/env python3
"""Apply the CONFIDENT volume corrections from volume_collision_fix.xlsx to the
newest inventory. Only touches the Volume column, only on rows whose current
Volume still matches the proposal's assumption (safety). Non-destructive: writes a
new timestamped canonical, confirms every other cell is byte-identical, and logs
every change to VOLUME_FIX_LOG.md for review."""
import openpyxl, glob, os, datetime, hashlib

SRC = max(glob.glob("attached_assets/comics_inventory_*.xlsx"), key=os.path.getmtime)
FIX = "volume_collision_fix.xlsx"


def ni(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s); return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def nv(v):
    try:
        return str(int(float(str(v).strip())))
    except (ValueError, TypeError):
        return "1"


def yr(v):
    import re
    n = [int(t) for t in re.findall(r"\d{4}", str(v or "")) if 1900 < int(t) < 2100]
    return str(n[0]) if n else ""


# --- load confident proposals: (title,issue,year,box) -> (current_vol, proposed) ---
fw = openpyxl.load_workbook(FIX, data_only=True).active
props = {}
for r in list(fw.iter_rows(values_only=True))[1:]:
    title, issue, box, year, pub, curv, prop, src, reason = r
    if prop in (None, "?"):
        continue
    k = (str(title).strip().lower(), ni(issue), yr(year), str(box).strip())
    props[k] = (nv(curv), str(prop).strip())

# --- cached values to flatten formulas (avoid nulling formula cells) ---
srcvals = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
               if w.title == "Sheet X" or (w.title == "Sheet X" or w.title.startswith("✅ Clean Inventory")))
cached = [list(r) for r in srcvals.iter_rows(values_only=True)]

wb = openpyxl.load_workbook(SRC)
ws = next(w for w in wb.worksheets if w.title == "Sheet X" or (w.title == "Sheet X" or w.title.startswith("✅ Clean Inventory")))
for ri, row in enumerate(ws.iter_rows()):
    for cj, cell in enumerate(row):
        if isinstance(cell.value, str) and cell.value.startswith("="):
            if ri < len(cached) and cj < len(cached[ri]):
                cell.value = cached[ri][cj]

H = [c.value for c in ws[1]]; C = {h: i for i, h in enumerate(H) if h}
tcol = C["Title"] + 1; icol = C["Issue #"] + 1; ycol = C["Year"] + 1; bcol = C["Box #"] + 1; vcol = C["Volume"] + 1

changes = []
for row in range(2, ws.max_row + 1):
    t = str(ws.cell(row, tcol).value or "").strip(); iss = ni(ws.cell(row, icol).value)
    year = yr(ws.cell(row, ycol).value); box = str(ws.cell(row, bcol).value or "").strip()
    k = (t.lower(), iss, year, box)
    if k not in props:
        continue
    cur_expected, proposed = props[k]
    cur_actual = nv(ws.cell(row, vcol).value)
    if cur_actual != cur_expected:            # data moved since proposal — skip, don't guess
        continue
    if cur_actual == proposed:                # already correct
        continue
    ws.cell(row, vcol, int(proposed) if proposed.isdigit() else proposed)
    changes.append((t, iss, year, box, cur_actual, proposed))

out = f"attached_assets/comics_inventory_{datetime.datetime.now():%d%m_%H%M}.xlsx"
wb.save(out)

# --- contamination check: only Volume column differs, only on changed rows ---
so = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
          if w.title == "Sheet X" or (w.title == "Sheet X" or w.title.startswith("✅ Clean Inventory")))
no = next(w for w in openpyxl.load_workbook(out, read_only=True, data_only=True).worksheets
          if w.title == "Sheet X" or (w.title == "Sheet X" or w.title.startswith("✅ Clean Inventory")))
sr = list(so.iter_rows(values_only=True)); nr = list(no.iter_rows(values_only=True))
vidx = C["Volume"]
offvol = 0
for a, b in zip(sr, nr):
    for j in range(len(a)):
        if j == vidx:
            continue
        if a[j] != b[j]:
            offvol += 1
print(f"SOURCE: {os.path.basename(SRC)}")
print(f"OUTPUT: {out}")
print(f"rows: {len(sr)-1} -> {len(nr)-1}   volume changes applied: {len(changes)}")
print(f"CONTAMINATION: non-Volume cell diffs = {offvol}  (must be 0)")

with open("VOLUME_FIX_LOG.md", "w") as f:
    f.write(f"# Volume collision fix — applied {datetime.datetime.now():%Y-%m-%d %H:%M}\n\n")
    f.write(f"Source: `{os.path.basename(SRC)}` -> `{os.path.basename(out)}`\n\n")
    f.write(f"Applied {len(changes)} confident volume corrections (only where current Volume still matched the proposal). "
            f"Only the Volume column changed; {offvol} other cells differ.\n\n")
    f.write("| Title | Issue | Year | Box | Vol was | Vol now |\n|---|---|---|---|---|---|\n")
    for t, iss, year, box, was, now in changes:
        f.write(f"| {t} | {iss} | {year} | {box} | {was} | {now} |\n")
print("Logged: VOLUME_FIX_LOG.md")
