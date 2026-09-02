#!/usr/bin/env python3
"""Verify the remaining 'check title' rows (those GCD couldn't confirm) against
the live Comic Vine proxy on :5001. Exact title match -> clear the note. A
different CV title -> report as a rename proposal (never auto-applied). No CV
match -> leave the flag. Non-destructive: flattens formulas, writes a new
timestamped canonical, only the Issue Note column changes."""
import openpyxl, glob, os, re, sqlite3, difflib, datetime, json, time, urllib.parse, urllib.request

SRC = max((f for f in glob.glob("attached_assets/comics_inventory_*.xlsx")
           if " copy" not in f and not os.path.basename(f).startswith("~$")),
          key=os.path.getmtime)
PROXY = "http://localhost:5001/api/covers/search"

def norm(s): return re.sub(r"[^a-z0-9]+", " ", str(s).lower()).strip()
gcd = set(norm(n) for (n,) in sqlite3.connect("gcd_local.sqlite").execute("select name from gcd_series"))

def cv_title(title, issue):
    q = urllib.parse.urlencode({"title": title, "issue": issue or "1", "refresh": "1"})
    try:
        with urllib.request.urlopen(f"{PROXY}?{q}", timeout=40) as r:
            d = json.load(r)
        m = d.get("match")
        return (m or {}).get("volume_name")
    except Exception as e:
        return f"__ERR__{e}"

srcvals = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory"))
cached = [list(r) for r in srcvals.iter_rows(values_only=True)]
wb = openpyxl.load_workbook(SRC)
ws = next(w for w in wb.worksheets if w.title.startswith("✅ Clean Inventory"))
for ri, row in enumerate(ws.iter_rows()):
    for cj, c in enumerate(row):
        if isinstance(c.value, str) and c.value.startswith("=") and ri < len(cached) and cj < len(cached[ri]):
            c.value = cached[ri][cj]
H = [c.value for c in ws[1]]; C = {h: i for i, h in enumerate(H) if h}
TCOL, NCOL, ICOL = C["Title"]+1, C["Issue Note"]+1, C["Issue #"]+1
def sval(r, col):
    v = ws.cell(r, col).value
    return "" if v is None else str(v).strip()

targets = []
for r in range(2, ws.max_row + 1):
    nl = sval(r, NCOL).lower()
    if "check title" in nl and "cover" not in nl and "dupe" not in nl and norm(sval(r, TCOL)) not in gcd:
        targets.append(r)

seen = {}  # (title,issue) -> cv title, to avoid duplicate queries
cleared, proposals, nomatch, errors = 0, [], [], []
for r in targets:
    t, iss = sval(r, TCOL), sval(r, ICOL)
    key = (t, iss)
    if key not in seen:
        seen[key] = cv_title(t, iss); time.sleep(0.5)
    cv = seen[key]
    if isinstance(cv, str) and cv.startswith("__ERR__"):
        errors.append((r, t, cv[7:])); continue
    if not cv:
        nomatch.append((r, t, iss)); continue
    if norm(cv) == norm(t):
        ws.cell(r, NCOL).value = None; cleared += 1
    else:
        ratio = round(difflib.SequenceMatcher(None, norm(t), norm(cv)).ratio(), 2)
        proposals.append((r, t, cv, ratio))

out = f"attached_assets/comics_inventory_{datetime.datetime.now():%d%m_%H%M}.xlsx"
wb.save(out)
so = list(srcvals.iter_rows(values_only=True))
no = list(next(w for w in openpyxl.load_workbook(out, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory")).iter_rows(values_only=True))
off = sum(1 for a,b in zip(so,no) for j in range(len(a)) if j != C["Issue Note"] and (a[j] or "")!=(b[j] or ""))
print(f"SOURCE: {os.path.basename(SRC)}\nOUTPUT: {os.path.basename(out)}")
print(f"targets: {len(targets)}   CV-verified notes cleared: {cleared}")
print(f"rename proposals: {len(proposals)}   no CV match: {len(nomatch)}   errors: {len(errors)}")
print(f"CONTAMINATION (non-Note diffs): {off}  (must be 0)")

with open("CV_VERIFY_LOG.md", "w") as f:
    f.write(f"# CV title verification — {datetime.datetime.now():%Y-%m-%d %H:%M}\n\n")
    f.write(f"`{os.path.basename(SRC)}` -> `{os.path.basename(out)}` · {cleared} notes cleared (CV title matched)\n\n")
    f.write("## Rename proposals (CV authoritative title — your call)\n\n| Row | Current | Comic Vine | Ratio |\n|---|---|---|---|\n")
    for r,t,cv,ra in sorted(proposals): f.write(f"| {r} | {t} | {cv} | {ra} |\n")
    f.write("\n## No CV match — still unverified\n\n| Row | Title | Issue |\n|---|---|---|\n")
    for r,t,i in sorted(nomatch): f.write(f"| {r} | {t} | {i} |\n")
    if errors:
        f.write("\n## Query errors\n\n| Row | Title | Error |\n|---|---|---|\n")
        for r,t,e in errors: f.write(f"| {r} | {t} | {e} |\n")
print("Logged: CV_VERIFY_LOG.md")
