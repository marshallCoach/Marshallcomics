#!/usr/bin/env python3
"""Apply the user-sourced title corrections (each backed by a Fandom / eBay /
Google-Books reference) + integrate the Aliens vs. Avengers flags. Flattens
formulas, writes a new timestamped canonical, flags the cross-box duplicates the
corrections create (non-destructive), then you validate."""
import openpyxl, glob, os, datetime

SRC = max((f for f in glob.glob("attached_assets/comics_inventory_*.xlsx")
           if " copy" not in f and not os.path.basename(f).startswith("~$")),
          key=os.path.getmtime)

# (title-contains, issue or None) -> (new_title, new_volume or None, new_year or None)
RENAMES = [
    ("Doom Patrol and Suicide Squad", "1", "Doom Patrol and Suicide Squad Special", None, None),
    ("Generations: The Best", "1", "Generations: Wolverine & All-New Wolverine", None, None),
    ("Nova: Annihilation Conquest", "1", "Nova", None, None),
    ("Star Trek: La'an", "1", "Star Trek: Lore War", None, None),
    ("The Vision and Scarlet Witch", "1", "Vision and the Scarlet Witch", "3", None),
    ("What If? Magic", "1", "What If? Magik", None, None),
    ("Fantastic Four: Empyre", "1", "Empyre", None, None),
    ("Fantastic Four: Empyre", "2", "Empyre", None, None),
    ("New Warriors: Giant-Size Spectacular", "1", "New Warriors", "2", "1999"),
]
# after renames, flag every row in these (title, issue, year) groups as reviewed
DUP_GROUPS = [("Doom Patrol and Suicide Squad Special", "1", "1988"),
              ("Empyre", "1", "2020"), ("Empyre", "2", "2020")]

srcvals = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory"))
cached = [list(r) for r in srcvals.iter_rows(values_only=True)]
wb = openpyxl.load_workbook(SRC)
ws = next(w for w in wb.worksheets if w.title.startswith("✅ Clean Inventory"))
for ri, row in enumerate(ws.iter_rows()):
    for cj, c in enumerate(row):
        if isinstance(c.value, str) and c.value.startswith("=") and ri < len(cached) and cj < len(cached[ri]):
            c.value = cached[ri][cj]
H = [c.value for c in ws[1]]; C = {h: i for i, h in enumerate(H) if h}
TCOL, NCOL, ICOL, YCOL, VCOL, VDCOL = (C["Title"]+1, C["Issue Note"]+1, C["Issue #"]+1,
                                       C["Year"]+1, C["Volume"]+1, C["⚠ Verify Duplicate"]+1)
def sv(r, c):
    v = ws.cell(r, c).value
    return "" if v is None else str(v).strip()

applied = []
for r in range(2, ws.max_row + 1):
    t, iss = sv(r, TCOL), sv(r, ICOL)
    for cont, wiss, nt, nv, ny in RENAMES:
        if cont in t and (wiss is None or iss == wiss):
            ws.cell(r, TCOL, nt)
            if nv is not None: ws.cell(r, VCOL, int(nv) if nv.isdigit() else nv)
            if ny is not None: ws.cell(r, YCOL, int(ny) if ny.isdigit() else ny)
            ws.cell(r, NCOL).value = None
            applied.append((r, t, nt)); break
    # integrate Aliens vs. Avengers: clear their check notes
    if "Aliens vs. Avengers" in sv(r, TCOL) and sv(r, NCOL):
        ws.cell(r, NCOL).value = None

flagged = 0
for r in range(2, ws.max_row + 1):
    if (sv(r, TCOL), sv(r, ICOL), sv(r, YCOL)) in DUP_GROUPS and not sv(r, VDCOL):
        ws.cell(r, VDCOL, "⚠ Verify Duplicate"); flagged += 1

out = f"attached_assets/comics_inventory_{datetime.datetime.now():%d%m_%H%M}.xlsx"
wb.save(out)
so = list(srcvals.iter_rows(values_only=True))
no = list(next(w for w in openpyxl.load_workbook(out, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory")).iter_rows(values_only=True))
allowed = {C["Title"], C["Issue Note"], C["Volume"], C["Year"], C["⚠ Verify Duplicate"]}
off = sum(1 for a,b in zip(so,no) for j in range(len(a)) if j not in allowed and (a[j] or "")!=(b[j] or ""))
print(f"OUTPUT: {os.path.basename(out)}   renames: {len(applied)}   dup-rows flagged: {flagged}")
for r,a,b in applied: print(f"   row {r}: {a!r} -> {b!r}")
print(f"CONTAMINATION (unexpected col diffs): {off}  (must be 0)")
