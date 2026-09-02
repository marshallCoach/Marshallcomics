#!/usr/bin/env python3
"""Two user-confirmed title fixes: 'GODS' -> 'G.O.D.S.' (Hickman 2023) and the
'Fantastic Four: Empyre' #0 -> 'Empyre: Fantastic Four' (correct title per
Marvel Fandom). Matched by content; clears the 'check title' note. Flattens
formulas, writes a new timestamped canonical, only Title/Issue Note change."""
import openpyxl, glob, os, datetime

SRC = max((f for f in glob.glob("attached_assets/comics_inventory_*.xlsx")
           if " copy" not in f and not os.path.basename(f).startswith("~$")),
          key=os.path.getmtime)

srcvals = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory"))
cached = [list(r) for r in srcvals.iter_rows(values_only=True)]
wb = openpyxl.load_workbook(SRC)
ws = next(w for w in wb.worksheets if w.title.startswith("✅ Clean Inventory"))
for ri, row in enumerate(ws.iter_rows()):
    for cj, c in enumerate(row):
        if isinstance(c.value, str) and c.value.startswith("=") and ri < len(cached) and cj < len(cached[ri]):
            c.value = cached[ri][cj]
H = [c.value for c in ws[1]]; C = {h: i for i, h in enumerate(H) if h}
TCOL, NCOL, ICOL = C["Title"]+1, C["Issue Note"]+1, C["Issue #"]+1
def sval(r, col):
    v = ws.cell(r, col).value
    return "" if v is None else str(v).strip()

applied = []
for r in range(2, ws.max_row + 1):
    t, iss = sval(r, TCOL), sval(r, ICOL)
    if t == "GODS":
        ws.cell(r, TCOL, "G.O.D.S."); ws.cell(r, NCOL).value = None
        applied.append((r, "GODS", "G.O.D.S."))
    elif t == "Fantastic Four: Empyre" and iss == "0":
        ws.cell(r, TCOL, "Empyre: Fantastic Four"); ws.cell(r, NCOL).value = None
        applied.append((r, "Fantastic Four: Empyre #0", "Empyre: Fantastic Four #0"))

out = f"attached_assets/comics_inventory_{datetime.datetime.now():%d%m_%H%M}.xlsx"
wb.save(out)
so = list(srcvals.iter_rows(values_only=True))
no = list(next(w for w in openpyxl.load_workbook(out, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory")).iter_rows(values_only=True))
allowed = {C["Title"], C["Issue Note"]}
off = sum(1 for a,b in zip(so,no) for j in range(len(a)) if j not in allowed and (a[j] or "")!=(b[j] or ""))
print(f"OUTPUT: {os.path.basename(out)}")
for r,a,b in applied: print(f"   row {r}: {a} -> {b}")
print(f"CONTAMINATION (non Title/Note): {off}  (must be 0)")
