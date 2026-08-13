#!/usr/bin/env python3
"""
brb_volume_repair.py — repair the (title, volume) blocks whose declared years
span more years than any single volume could have run.

Found by the legacy-credit sweep: 5 blocks / 215 rows declare an impossible
span, e.g. Black Panther Vol 7 carrying rows dated 2005-2023. These blocks are
what blocks credit filling — a wrong Volume means every Fandom candidate fails
the year gate, so the row can never be resolved.

THE AMBIGUITY: for a row like "Black Panther #2, year 2005, vol 7, John Ridley"
both the year and the volume cannot be right (Ridley's BP #2 is Vol 8, 2021;
Hudlin's 2005 BP #2 is Vol 4). Guessing either way would corrupt data.

METHOD — two independent votes, from data already in the row:
  vote A: the row's declared YEAR      vs the candidate page's release year
  vote B: the row's existing WRITER    vs the candidate page's writer
Candidates are the real volumes that actually have that issue number
(volume_index.json, enumerated from Fandom allpages — no guessing which exist).

  both votes agree on ONE volume  -> CONFIDENT, auto-applied
  only one vote, or votes split   -> REVIEW, written to a sheet, NOT applied

That mirrors the project rule that an uncertain call is never applied silently.
Only the Volume cell is ever written; Year is only ever PROPOSED in the review
sheet, because which of the two is wrong is exactly what needs human eyes.

Fandom API — free, does not touch the Comic Vine quota.

Usage:
    python3 brb_volume_repair.py                 # dry-run + review sheet
    python3 brb_volume_repair.py --apply
"""
import argparse, glob, json, os, re, time, urllib.parse, urllib.request
import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
API = "https://marvel.fandom.com/api.php"
UA = "MarshallComicsInventory/1.0"
DELAY = 0.25
YEAR_TOL = 2

# the 5 impossible-span blocks (title, declared volume)
BLOCKS = [("Captain America", 9), ("Fantastic Four", 3), ("Fantastic Four", 5),
          ("Black Panther", 7), ("Captain America", 7)]


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx"))
         if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def blank(v):
    return v is None or str(v).strip() in ("", "nan", "None")


def ni(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def nv(v):
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def yr(v):
    n = [int(t) for t in re.findall(r"\d{4}", str(v or "")) if 1900 < int(t) < 2100]
    return n[0] if n else None


def surname(s):
    """Compare people by surname — 'C.F. Villa' vs 'Carlos Villa' etc."""
    s = re.sub(r"[^a-z ]", " ", str(s or "").lower())
    parts = [p for p in s.split() if len(p) > 2]
    return parts[-1] if parts else ""


_cache = {}


def fandom(page):
    if page in _cache:
        return _cache[page]
    out = None
    try:
        u = (f"{API}?action=parse&page={urllib.parse.quote(page)}"
             "&prop=wikitext&format=json&formatversion=2&redirects=1")
        with urllib.request.urlopen(urllib.request.Request(u, headers={"User-Agent": UA}), timeout=30) as r:
            d = json.load(r)
        wt = d.get("parse", {}).get("wikitext", "")
        if wt:
            def f(*names):
                for n in names:
                    m = re.search(r"\|\s*" + n + r"\s*=\s*([^\n|]+)", wt)
                    if m:
                        v = re.sub(r"\[\[|\]\]", "", m.group(1)).split("|")[-1].strip()
                        if v:
                            return v
                return None
            out = {"year": yr(f("ReleaseDate", "CoverDate", "Year")),
                   "writer": f("Writer1_1", "Writer_1", "Writer1"),
                   "artist": f("Penciler1_1", "Penciler_1", "Artist1_1")}
    except Exception:
        pass
    _cache[page] = out
    time.sleep(DELAY)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=None)
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--out", default=None)
    ap.add_argument("--review", default="volume_repair_review.xlsx")
    args = ap.parse_args()

    idx = json.load(open(os.path.join(ROOT, "volume_index.json")))
    xlsx = os.path.abspath(args.xlsx) if args.xlsx else latest_xlsx()
    print(f"Source: {os.path.basename(xlsx)}  (mode: {'APPLY' if args.apply else 'DRY-RUN'})")

    wb = openpyxl.load_workbook(xlsx)
    ws = next(wb[n] for n in wb.sheetnames if (n == "Sheet X" or n.startswith("✅ Clean Inventory")))
    H = [c.value for c in next(ws.iter_rows(max_row=1))]
    C = {n: H.index(n) for n in H if n}
    ti, ii, yi, vi, wi = C["Title"], C["Issue #"], C["Year"], C["Volume"], C["Writer(s)"]

    blocks = {(t, v) for t, v in BLOCKS}
    targets = []
    for n, row in enumerate(ws.iter_rows(min_row=2), 2):
        t = str(row[ti].value or "").strip()
        if (t, nv(row[vi].value)) in blocks:
            targets.append((n, row))
    print(f"  rows in the 5 impossible-span blocks: {len(targets)}")

    conf, review = [], []
    for n, row in targets:
        t = str(row[ti].value or "").strip()
        iss, y, dv = ni(row[ii].value), yr(row[yi].value), nv(row[vi].value)
        rw = surname(row[wi].value)
        cands = idx.get(t, {}).get(iss, [])
        scored = []
        for v in cands:
            info = fandom(f"{t} Vol {v} {iss}")
            if not info:
                continue
            ay = info["year"] and y and abs(info["year"] - y) <= YEAR_TOL
            aw = rw and surname(info["writer"]) == rw
            scored.append({"vol": v, "year": info["year"], "writer": info["writer"],
                           "year_ok": bool(ay), "writer_ok": bool(aw)})
        both = [s for s in scored if s["year_ok"] and s["writer_ok"]]
        wonly = [s for s in scored if s["writer_ok"] and not s["year_ok"]]
        yonly = [s for s in scored if s["year_ok"] and not s["writer_ok"]]
        rec = {"row": n, "title": t, "issue": iss, "year": y, "vol": dv,
               "writer": str(row[wi].value or ""), "cands": scored}
        if len(both) == 1:
            rec["proposed"], rec["why"] = both[0]["vol"], f"year AND writer agree (Fandom {both[0]['year']}, {both[0]['writer']})"
            conf.append((rec, row))
        elif len(wonly) == 1 and not both and not yonly:
            rec["proposed"] = wonly[0]["vol"]
            rec["why"] = (f"writer agrees ({wonly[0]['writer']}) but Fandom year {wonly[0]['year']} "
                          f"vs declared {y} — YEAR likely wrong")
            review.append(rec)
        elif len(yonly) == 1 and not both:
            rec["proposed"] = yonly[0]["vol"]
            rec["why"] = f"year agrees (Fandom {yonly[0]['year']}) but writer differs ({yonly[0]['writer']} vs row)"
            review.append(rec)
        else:
            rec["proposed"] = ""
            rec["why"] = f"ambiguous — {len(both)} both-match, {len(wonly)} writer-only, {len(yonly)} year-only"
            review.append(rec)

    changed = sum(1 for r, _ in conf if r["proposed"] != r["vol"])
    print(f"\n  CONFIDENT (year+writer agree)  : {len(conf)}   of which volume changes: {changed}")
    print(f"  REVIEW (ambiguous / conflicting): {len(review)}")
    print("\n  sample confident fixes:")
    for r, _ in [c for c in conf if c[0]['proposed'] != c[0]['vol']][:10]:
        print(f"    row{r['row']:<6} {r['title'][:20]:<21}#{r['issue']:<5} yr={r['year']}  "
              f"vol {r['vol']} -> {r['proposed']}   {r['why'][:52]}")

    if args.apply:
        for r, row in conf:
            row[vi].value = r["proposed"]
        out = args.out or xlsx.replace(".xlsx", "_VOLFIX.xlsx")
        wb.save(out)
        print(f"\n  Written: {out}")
    else:
        print("\n  [DRY-RUN] nothing written")

    # review sheet — every candidate shown, per the never-guess-silently rule
    rb = openpyxl.Workbook(); sh = rb.active; sh.title = "Needs review"
    hdr = ["Sheet row", "Title", "Issue", "Declared year", "Declared vol", "Row writer",
           "PROPOSED vol", "Why", "All candidates (vol: year / writer)", "Approve? (y/n)"]
    sh.append(hdr)
    for c in sh[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="C00000")
    sh.freeze_panes = "A2"
    for i, w in enumerate((10, 22, 8, 14, 13, 24, 13, 60, 70, 12), 1):
        sh.column_dimensions[chr(64 + i)].width = w
    for r in review:
        cs = "; ".join(f"v{c['vol']}: {c['year']} / {c['writer']}" for c in r["cands"])
        sh.append([r["row"], r["title"], r["issue"], r["year"], r["vol"], r["writer"],
                   r["proposed"], r["why"], cs, ""])
    rb.save(args.review)
    print(f"  Review sheet: {args.review}  ({len(review)} rows)")


if __name__ == "__main__":
    main()
