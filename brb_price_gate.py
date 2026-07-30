#!/usr/bin/env python3
"""
brb_price_gate.py — suppress eBay prices that are inflated by wrong-issue matches.

The problem (Roberto, 2707): for reused issue numbers — Annuals especially — the
eBay search matches an OLDER, valuable issue of the same number, so a modern
cheap book (e.g. the 2018 Justice League Annual #1, ~$5) shows the original
annual's price (~$150). The comps are all the wrong book, so the rescore's
bimodal split can't help — there's no cheap cluster.

Sanity gate, anchored on the book's own catalogued NM value (independent of eBay):
  For any priced book with eBay median >= $10:
    - CONFIRMED KEY  -> trust it (keys can be genuinely valuable).
    - modern (year >= 2015) NON-key with median >= 3x its NM estimate:
        * SUPPRESS  if ratio >= 5 OR comps < 3  (near-certain wrong-issue match):
          blank the eBay median/avg so the app falls back to the NM estimate,
          keep the original under _suppressed for reversibility.
        * FLAG      otherwise (ratio 3-5 with >=3 comps): keep but mark for review.

Non-destructive & reversible: rewrites ebay_pricing_results.json (the working
overlay, regenerated anyway) with a _suppressed record + a review xlsx. Re-run
after any fresh eBay pull.

Usage:
    python3 brb_price_gate.py            # report only
    python3 brb_price_gate.py --apply    # suppress + write review sheet
"""
import argparse, glob, json, os, re
import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
EBAY = os.path.join(ROOT, "ebay_pricing_results.json")


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def num(v):
    try:
        return float(str(v).replace("$", "").replace(",", "").strip())
    except (ValueError, TypeError, AttributeError):
        return None


def norm(v):
    s = str(v).strip().lstrip("#")
    try:
        return str(float(s))
    except (ValueError, TypeError):
        return s


def yes(v):
    return str(v).strip().upper() in ("YES", "Y", "TRUE", "1")


def yr(v):
    n = [int(t) for t in re.findall(r"\d{4}", str(v or "")) if 1900 < int(t) < 2100]
    return n[0] if n else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--review", default="price_gate_review.xlsx")
    args = ap.parse_args()

    d = json.load(open(EBAY))
    xlsx = latest_xlsx()
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("✅ Clean Inventory"))
    rows = list(ws.iter_rows(values_only=True))
    H = list(rows[0]); C = {n: H.index(n) for n in H if n}

    def g(r, n):
        i = C.get(n); return r[i] if i is not None else None

    suppressed, flagged = [], []
    for r in rows[1:]:
        t = str(g(r, "Title") or "").strip()
        iss = norm(g(r, "Issue #"))
        key = f"{t}|||{iss}"
        e = d.get(key)
        if not e:
            continue
        med = e.get("median") or e.get("avg")
        if not med or med < 10:
            continue
        if yes(g(r, "Key Issue?")):
            continue
        nm = num(g(r, "Est. Raw Value (NM) $")) or 0
        y = yr(g(r, "Year")) or 0
        cnt = e.get("count") or 0
        ratio = med / nm if nm else 99
        # Suppress inflated prices from year 2000 on (the wrong-issue contamination
        # era — cheap digital/New-52-era books mismatched to older valuable issues).
        # Pre-2000 vintage is left alone (its eBay value is more likely genuine).
        # Confirmed KEYS are always exempt above, so real value is preserved.
        if y < 2000 or ratio < 3:
            continue
        rec = dict(title=t, issue=g(r, "Issue #"), year=y, nm=nm, med=med, ratio=ratio, cnt=cnt, delta=med - nm)
        # Roberto (2707): go with the LOWER value for ALL borderline modern
        # non-key books (manage expectations) — suppress the inflated eBay price
        # and fall back to NM. Only surface for review the ones with a >= $25
        # delta (median - NM); smaller deltas aren't worth attention.
        suppressed.append((key, rec))

    review = [(k, r) for k, r in suppressed if r["delta"] >= 25]
    print(f"eBay-priced modern non-key books >$10 with median >= 3x NM:")
    print(f"  SUPPRESSED (all -> fall back to lower NM value): {len(suppressed)}")
    print(f"  of which worth your review (>= $25 delta)      : {len(review)}")
    print(f"\n  top suppressions (eBay median -> NM fallback):")
    for key, rec in sorted(suppressed, key=lambda x: -x[1]["med"])[:12]:
        print(f"    {rec['title'][:30]:<31} #{rec['issue']} ({rec['year']})  ${rec['med']:.0f} -> ${rec['nm']:.0f}  ({rec['cnt']}c)")

    if args.apply:
        seen = set()
        for key, rec in suppressed:
            if key in seen:
                continue
            seen.add(key)
            e = d[key]
            e["_suppressed"] = {"median": e.get("median"), "avg": e.get("avg"),
                                "reason": f"wrong-issue: median {rec['med']:.0f} vs NM {rec['nm']:.0f} ({rec['cnt']} comps)"}
            e["median"] = None
            e["avg"] = None
            e["price_gate"] = "suppressed"
        json.dump(d, open(EBAY, "w"))
        print(f"\n  Applied: blanked {len(seen)} inflated eBay medians (fall back to NM). Re-run: node gen_data.mjs")

    rb = openpyxl.Workbook(); sh = rb.active; sh.title = "Price gate review >=$25"
    sh.append(["Title", "Issue", "Year", "eBay median (suppressed)", "NM est (now shown)", "delta", "comps"])
    for c in sh[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="C00000")
    sh.freeze_panes = "A2"
    for key, rec in sorted(review, key=lambda x: -x[1]["delta"]):
        sh.append([rec["title"], rec["issue"], rec["year"], round(rec["med"], 2),
                   rec["nm"], round(rec["delta"], 2), rec["cnt"]])
    rb.save(args.review)
    print(f"  Review ({len(review)} rows, >= $25 delta): {args.review}")


if __name__ == "__main__":
    main()
