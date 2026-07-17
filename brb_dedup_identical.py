#!/usr/bin/env python3
"""
brb_dedup_identical.py — drop field-identical same-box duplicate rows.

A dup group = same (Title, Issue #, Year, Box #) with 2+ rows AND identical
collector-identity fields (Condition, Signed?, Signed By, NM $, VF $, CGC). Only
those are true double-entries; groups whose fields differ are real distinct
copies and are NEVER touched. Keeps the first occurrence, drops the rest.

Safety: dry-run by default; --apply writes a NEW *_DEDUP.xlsx, never the source
in place. --titles restricts to an explicit reviewed set (comma-separated), so a
broad sweep can't happen by accident.

Usage:
    python3 brb_dedup_identical.py --titles "Sam Wilson: Captain America"          # dry-run
    python3 brb_dedup_identical.py --titles "Sam Wilson: Captain America" --apply
    python3 brb_dedup_identical.py --all --apply          # every identical group (review first!)
"""
import argparse, glob, os, sys
from collections import defaultdict
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
IDENT_COLS = ("Condition", "Signed?", "Signed By", "Est. Raw Value (NM) $", "Est. Raw Value (VF) $", "CGC Worth It?")


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=None)
    ap.add_argument("--titles", default=None, help="comma-separated titles to restrict to")
    ap.add_argument("--all", action="store_true", help="consider every title (use with care)")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()
    if not args.titles and not args.all:
        ap.error("pass --titles \"...\" or --all")

    xlsx = os.path.abspath(args.xlsx) if args.xlsx else latest_xlsx()
    print(f"Source: {os.path.basename(xlsx)}   (mode: {'APPLY' if args.apply else 'DRY-RUN'})")
    scope = {t.strip() for t in args.titles.split(",")} if args.titles else None
    if scope:
        print(f"Scope: {len(scope)} title(s)")

    wb = openpyxl.load_workbook(xlsx)
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("✅ Clean Inventory"))
    H = [c.value for c in next(ws.iter_rows(max_row=1))]
    ti, ii, yi, bi = (H.index(c) for c in ("Title", "Issue #", "Year", "Box #"))
    idf = [H.index(c) for c in IDENT_COLS]

    groups = defaultdict(list)
    for excel_row in range(2, ws.max_row + 1):
        vals = [ws.cell(row=excel_row, column=c + 1).value for c in range(len(H))]
        title = str(vals[ti] or "").strip()
        if scope and title not in scope:
            continue
        key = (title, str(vals[ii] or "").strip(), str(vals[yi] or "").strip(), str(vals[bi] or "").strip())
        groups[key].append((excel_row, tuple(str(vals[c]) for c in idf)))

    drop_rows = []
    kept_differ = 0
    for key, items in groups.items():
        if len(items) < 2:
            continue
        if len({sig for _, sig in items}) == 1:
            # identical -> keep first, drop rest
            drop_rows.extend(r for r, _ in items[1:])
        else:
            kept_differ += 1

    print(f"\n  Identical dup groups found: {sum(1 for k,v in groups.items() if len(v)>1 and len({s for _,s in v})==1)}")
    print(f"  Rows to drop (one+ per identical group): {len(drop_rows)}")
    print(f"  Differing dup groups left untouched: {kept_differ}")

    if not args.apply:
        print("\n[DRY RUN] No file written. Re-run with --apply.")
        return

    for r in sorted(drop_rows, reverse=True):  # delete bottom-up to preserve indices
        ws.delete_rows(r, 1)
    out = args.out or xlsx.replace(".xlsx", "_DEDUP.xlsx")
    wb.save(out)
    print(f"\n  Dropped {len(drop_rows)} rows. Written: {out}")


if __name__ == "__main__":
    main()
