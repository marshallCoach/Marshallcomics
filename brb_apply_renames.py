#!/usr/bin/env python3
"""Apply the confident GCD title corrections (official published titles) to the
flagged rows, and clear their 'check title' note. Safety: each edit is keyed to
the expected current title — if a row's title has changed, it's skipped, not
guessed. Non-destructive: flattens formulas, writes a new timestamped canonical,
and confirms only the Title / Issue Note columns changed."""
import openpyxl, glob, os, datetime

SRC = max((f for f in glob.glob("attached_assets/comics_inventory_*.xlsx")
           if " copy" not in f and not os.path.basename(f).startswith("~$")),
          key=os.path.getmtime)

# row -> (expected current title, corrected official title)
RENAMES = {
    819:  ("Avengers and Moon Girl",        "Avengers & Moon Girl"),
    3209: ("Dracula: Bloodhunt",            "Dracula: Blood Hunt"),
    3210: ("Dracula: Bloodhunt",            "Dracula: Blood Hunt"),
    3553: ("Falcon and Winter Soldier",     "Falcon & Winter Soldier"),
    6483: ("Nextwave: Agents of Hate",      "Nextwave: Agents of H.A.T.E."),
    8043: ("Strange Academy: Bloodhunt",    "Strange Academy: Blood Hunt"),
    8044: ("Strange Academy: Bloodhunt",    "Strange Academy: Blood Hunt"),
    10073:("Union Jack the Ripper: Bloodhunt", "Union Jack the Ripper: Blood Hunt"),
}

srcvals = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory"))
cached = [list(r) for r in srcvals.iter_rows(values_only=True)]

wb = openpyxl.load_workbook(SRC)
ws = next(w for w in wb.worksheets if w.title.startswith("✅ Clean Inventory"))
for ri, row in enumerate(ws.iter_rows()):
    for cj, c in enumerate(row):
        if isinstance(c.value, str) and c.value.startswith("=") and ri < len(cached) and cj < len(cached[ri]):
            c.value = cached[ri][cj]

H = [c.value for c in ws[1]]; C = {h: i for i, h in enumerate(H) if h}
TCOL = C["Title"] + 1; NCOL = C["Issue Note"] + 1

applied, skipped = [], []
for row, (expect, new) in RENAMES.items():
    cur = str(ws.cell(row, TCOL).value or "").strip()
    if cur != expect:
        skipped.append((row, cur, expect)); continue
    ws.cell(row, TCOL, new)
    ws.cell(row, NCOL, None)          # flag resolved
    applied.append((row, expect, new))

out = f"attached_assets/comics_inventory_{datetime.datetime.now():%d%m_%H%M}.xlsx"
wb.save(out)

so = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
          if w.title.startswith("✅ Clean Inventory"))
no = next(w for w in openpyxl.load_workbook(out, read_only=True, data_only=True).worksheets
          if w.title.startswith("✅ Clean Inventory"))
allowed = {C["Title"], C["Issue Note"]}
off = sum(1 for a, b in zip(so.iter_rows(values_only=True), no.iter_rows(values_only=True))
          for j in range(len(a)) if j not in allowed and (a[j] or "") != (b[j] or ""))
print(f"SOURCE: {os.path.basename(SRC)}\nOUTPUT: {os.path.basename(out)}")
print(f"renames applied: {len(applied)}   skipped (title moved): {len(skipped)}")
for r, e, n in applied: print(f"   row {r}: {e!r} -> {n!r}")
for r, c, e in skipped: print(f"   SKIP row {r}: found {c!r}, expected {e!r}")
print(f"CONTAMINATION (non Title/Note diffs): {off}  (must be 0)")
