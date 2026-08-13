#!/usr/bin/env python3
"""
brb_gcd_enrich.py — two GCD-sourced cleanups in one pass, new-file-only:

  --verify-tags : replace 'Verify' / '(verify)' placeholders in Writer(s) /
                  Artist(s) with GCD's authoritative credit, and DROP the tag.
                  Cells GCD can't resolve keep their tag (still need a human).
  --fill-volume : fill BLANK Volume cells with a GCD-derived volume number
                  (publisher-filtered same-name series ordered by year; only
                  when the row resolves to exactly one series and the title has
                  <=3 GCD series — same safety gate as brb_gcd_volume_fix).

Dry-run by default; --apply writes *_ENRICHED.xlsx. Never edits source in place.
"""
import argparse, glob, os, re, sqlite3
from collections import defaultdict
import openpyxl
from brb_gcd_lookup import lookup
from brb_gcd_volume_check import tight, pub_match, parse_year_range, DB

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
MAX_SERIES = 3


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def norm_issue(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s); return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError): return s


def is_verify(v):
    return v is not None and "verif" in str(v).lower()


def is_blank(v):
    return v is None or str(v).strip() in ("", "nan", "None")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=None)
    ap.add_argument("--verify-tags", action="store_true")
    ap.add_argument("--fill-volume", action="store_true")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()
    if not args.verify_tags and not args.fill_volume:
        ap.error("pass --verify-tags and/or --fill-volume")

    xlsx = os.path.abspath(args.xlsx) if args.xlsx else latest_xlsx()
    print(f"Source: {os.path.basename(xlsx)}  (mode: {'APPLY' if args.apply else 'DRY-RUN'})")

    wb = openpyxl.load_workbook(xlsx)
    ws = next(wb[n] for n in wb.sheetnames if (n == "Sheet X" or n.startswith("✅ Clean Inventory")))
    H = [c.value for c in next(ws.iter_rows(max_row=1))]
    C = {n: H.index(n) for n in ("Title", "Issue #", "Year", "Publisher", "Writer(s)", "Artist(s)", "Volume")}

    # --- volume numbering per (title, publisher), from GCD ---
    conn = sqlite3.connect(DB)
    gcd_by_key = defaultdict(dict)
    for sid, mt, name, yb, ye, pub in conn.execute(
        "SELECT s.id, s.matched_title, s.name, s.year_began, s.year_ended, p.name "
        "FROM gcd_series s LEFT JOIN gcd_publisher p ON p.id = s.publisher_id"):
        gcd_by_key[tight(mt)][sid] = {"yb": yb, "ye": ye, "pub": pub}
    try:
        aliases = dict(conn.execute("SELECT alias, matched_title FROM gcd_title_alias"))
    except sqlite3.OperationalError:
        aliases = {}

    def derived_volume(title, pub, year):
        pool = dict(gcd_by_key.get(tight(title), {}))
        if title in aliases:
            pool.update(gcd_by_key.get(tight(aliases[title]), {}))
        cands = [s for s in pool.values() if pub_match(pub, s["pub"]) and s["yb"]]
        by_year = {}
        for s in sorted(cands, key=lambda s: s["yb"]):
            by_year.setdefault(s["yb"], s)
        ordered = sorted(by_year.values(), key=lambda s: s["yb"])
        if not ordered or len(ordered) > MAX_SERIES:
            return None
        yr = parse_year_range(year)
        if not yr:
            return None
        hits = [i for i, s in enumerate(ordered) if s["yb"] <= yr[1] and (s["ye"] or 2100) >= yr[0]]
        return hits[0] + 1 if len(hits) == 1 else None

    vt_w = vt_a = vt_unresolved = vol_filled = 0
    for row in ws.iter_rows(min_row=2):
        title = str(row[C["Title"]].value or "").strip()
        if not title:
            continue
        issue = row[C["Issue #"]].value
        year = row[C["Year"]].value
        pub = row[C["Publisher"]].value

        if args.verify_tags and (is_verify(row[C["Writer(s)"]].value) or is_verify(row[C["Artist(s)"]].value)):
            r = lookup(title, issue, year, pub)
            cr = r.get("issue") if r.get("found") else None
            did = False
            if is_verify(row[C["Writer(s)"]].value) and cr and cr.get("writer"):
                if args.apply: row[C["Writer(s)"]].value = cr["writer"]
                vt_w += 1; did = True
            if is_verify(row[C["Artist(s)"]].value) and cr and cr.get("artist"):
                if args.apply: row[C["Artist(s)"]].value = cr["artist"]
                vt_a += 1; did = True
            if not did:
                vt_unresolved += 1

        if args.fill_volume and is_blank(row[C["Volume"]].value):
            dv = derived_volume(title, pub, year)
            if dv:
                if args.apply: row[C["Volume"]].value = dv
                vol_filled += 1

    if args.verify_tags:
        print(f"\nVERIFY-TAGS: replaced {vt_w} Writer + {vt_a} Artist cells from GCD; {vt_unresolved} rows GCD couldn't verify (tag kept)")
    if args.fill_volume:
        print(f"FILL-VOLUME: filled {vol_filled} blank Volume cells from GCD-derived numbering")

    if not args.apply:
        print("\n[DRY RUN] No file written.")
        return
    out = args.out or xlsx.replace(".xlsx", "_ENRICHED.xlsx")
    wb.save(out)
    print(f"\nWritten: {out}")


if __name__ == "__main__":
    main()
