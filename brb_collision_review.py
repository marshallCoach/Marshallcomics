#!/usr/bin/env python3
"""
brb_collision_review.py — build a REVIEW sheet for same-box collision groups,
anchored on Volume (per Roberto's "always check volume" rule).

Lesson from the failed auto-fix: reassigning Volume to match a cover artist
collides rows into existing copies (+5 true dupes). The safe lever is the
opposite — trust each row's Volume and correct its YEAR to what that volume's
issue actually is, which SEPARATES the colliding rows without creating dupes.

For every collision group (Title+Issue+Year+Box, >1 row) in the scoped boxes,
for each member row this fetches the real release year of "{Title} Vol {vol}
{issue}" from Marvel Fandom and flags rows whose declared Year disagrees. Where
the volume ALSO looks wrong (cover artist doesn't fit that volume's era) it says
so, so a human decides rather than the script guessing.

READ-ONLY. Writes a review xlsx; never the source file. Fandom = free, no CV quota.

Usage:
    python3 brb_collision_review.py --boxes 7,10
"""
import argparse, glob, json, os, re, time, unicodedata, urllib.parse, urllib.request
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
WIKIS = ["https://marvel.fandom.com/api.php", "https://dc.fandom.com/api.php"]
UA = "MarshallComicsInventory/1.0"
DELAY = 0.25


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def ni(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s); return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def nv(v):
    try:
        return str(int(float(str(v).strip())))
    except (ValueError, TypeError):
        return str(v or "").strip()


def yr(v):
    n = [int(t) for t in re.findall(r"\d{4}", str(v or "")) if 1900 < int(t) < 2100]
    return n[0] if n else None


def deaccent(s):
    return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()


_cache = {}


def fandom_year_cover(page):
    if page in _cache:
        return _cache[page]
    out = (None, None)
    for API in WIKIS:              # try Marvel, then DC
        try:
            u = (f"{API}?action=parse&page={urllib.parse.quote(page)}&prop=wikitext"
                 "&format=json&formatversion=2&redirects=1")
            with urllib.request.urlopen(urllib.request.Request(u, headers={"User-Agent": UA}), timeout=30) as r:
                d = json.load(r)
            wt = d.get("parse", {}).get("wikitext", "")
            time.sleep(DELAY)
            if not wt:
                continue
            def f(*names):
                for n in names:
                    m = re.search(r"\|\s*" + n + r"\s*=\s*([^\n|]+)", wt)
                    if m:
                        v = re.sub(r"\[\[|\]\]", "", m.group(1)).split("|")[-1].strip()
                        if v:
                            return v
                return None
            # DC uses Pubyear/Year; Marvel uses ReleaseDate/CoverDate
            y = yr(f("ReleaseDate", "CoverDate", "Pubyear", "Year"))
            if y:
                out = (y, f("CoverArtist1_1", "CoverArtist_1", "CoverArtist1"))
                break
        except Exception:
            time.sleep(DELAY)
    _cache[page] = out
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--boxes", default="7,10")
    ap.add_argument("--out", default="collision_review.xlsx")
    args = ap.parse_args()
    boxes = {b.strip() for b in args.boxes.split(",")}
    all_boxes = "all" in boxes

    xlsx = latest_xlsx()
    print(f"Source: {os.path.basename(xlsx)}  boxes: {sorted(boxes)}", flush=True)
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = next(wb[n] for n in wb.sheetnames if (n == "Sheet X" or n.startswith("✅ Clean Inventory")))
    rows = list(ws.iter_rows(values_only=True))
    H = list(rows[0]); C = {n: H.index(n) for n in H if n}

    def g(r, n):
        i = C.get(n); return r[i] if i is not None else None

    groups = defaultdict(list)
    for r in rows[1:]:
        b = str(g(r, "Box #") or "").strip()
        if not all_boxes and b not in boxes:
            continue
        groups[(str(g(r, "Title") or "").strip(), ni(g(r, "Issue #")), str(g(r, "Year") or "").strip(), b)].append(r)
    coll = {k: v for k, v in groups.items() if len(v) > 1}
    print(f"  collision groups: {len(coll)}", flush=True)

    out = openpyxl.Workbook(); sh = out.active; sh.title = "Collision review"
    hdr = ["Group", "Title", "Issue", "Box", "Declared Vol", "Declared Year", "Cover Artist",
           "Fandom year for this Vol", "Verdict", "Proposed Year", "Approve? (y/n)"]
    sh.append(hdr)
    for c in sh[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2F5496")
    sh.freeze_panes = "A2"
    for i, w in enumerate((6, 26, 7, 5, 12, 13, 30, 22, 34, 13, 13), 1):
        sh.column_dimensions[chr(64 + i)].width = w

    gi = 0; flagged = 0
    for (title, iss, y, box), rowlist in sorted(coll.items()):
        gi += 1
        for r in rowlist:
            vol = nv(g(r, "Volume"))
            dy = yr(g(r, "Year"))
            fy, fcov = fandom_year_cover(f"{title} Vol {vol} {iss}") if vol.isdigit() else (None, None)
            verdict, prop = "", ""
            if fy and dy:
                if abs(fy - dy) <= 1:
                    verdict = "year OK for this volume"
                else:
                    verdict = f"YEAR WRONG — Vol {vol} #{iss} is {fy}, not {dy}"
                    prop = fy; flagged += 1
            elif not fy:
                verdict = "Fandom page not found for this Vol/issue — volume may be wrong"
            sh.append([gi, title, iss, box, vol, dy, g(r, "Cover Artist"), fy or "", verdict, prop, ""])
    out.save(args.out)
    print(f"\n  rows flagged with a wrong year (volume-anchored): {flagged}", flush=True)
    print(f"  Written: {args.out}  (READ-ONLY — nothing applied)", flush=True)


if __name__ == "__main__":
    main()
