#!/usr/bin/env python3
"""Apply the 4 confirmed held renames + remove the 2 confirmed true-duplicate
rows. Flattens formulas, writes a new timestamped canonical, verifies the row
count dropped by exactly 2 and the dup groups collapsed to one row each."""
import openpyxl, glob, os, datetime

SRC = max((f for f in glob.glob("attached_assets/comics_inventory_*.xlsx")
           if " copy" not in f and not os.path.basename(f).startswith("~$")),
          key=os.path.getmtime)

# expected current title -> corrected (GCD authoritative). Only flagged rows.
RENAMES = {
    "SGU: Stargate Universe":     "Stargate Universe",
    "Superman and Robin Special": "Superman & Robin Special",
    "The Iron Age: Alpha":        "Iron Age: Alpha",
    "Ultraman x Avengers":        "Ultraman x the Avengers",
}
# true-duplicate rows to remove (keep the lower-numbered twin)
REMOVE_ROWS = [889, 3779]
KEEP_ROWS   = [246, 3778]   # clear their now-stale Verify-Duplicate flag

srcvals = next(w for w in openpyxl.load_workbook(SRC, read_only=True, data_only=True).worksheets
               if w.title.startswith("✅ Clean Inventory"))
cached = [list(r) for r in srcvals.iter_rows(values_only=True)]
n_before = len(cached) - 1

wb = openpyxl.load_workbook(SRC)
ws = next(w for w in wb.worksheets if w.title.startswith("✅ Clean Inventory"))
for ri, row in enumerate(ws.iter_rows()):
    for cj, c in enumerate(row):
        if isinstance(c.value, str) and c.value.startswith("=") and ri < len(cached) and cj < len(cached[ri]):
            c.value = cached[ri][cj]

H = [c.value for c in ws[1]]; C = {h: i for i, h in enumerate(H) if h}
TCOL = C["Title"] + 1; NCOL = C["Issue Note"] + 1; VDCOL = C["⚠ Verify Duplicate"] + 1

def sval(r, col):
    v = ws.cell(r, col).value
    return "" if v is None else str(v).strip()

# 1) renames (only on rows flagged 'check title', matched by expected title)
renamed = []
for r in range(2, ws.max_row + 1):
    t = sval(r, TCOL)
    if t in RENAMES and "check title" in sval(r, NCOL).lower():
        ws.cell(r, TCOL, RENAMES[t]); ws.cell(r, NCOL, None)
        renamed.append((r, t, RENAMES[t]))

# 2) clear stale Verify-Duplicate on kept twins; verified title note on 246
for r in KEEP_ROWS:
    ws.cell(r, VDCOL, None)
ws.cell(246, NCOL, None)   # "Aliens vs. Avengers" #1 (2024) is a real Marvel title

# 3) remove the true-duplicate rows (highest index first so lower stays valid)
for r in sorted(REMOVE_ROWS, reverse=True):
    ws.delete_rows(r, 1)

out = f"attached_assets/comics_inventory_{datetime.datetime.now():%d%m_%H%M}.xlsx"
wb.save(out)

# verify
no = next(w for w in openpyxl.load_workbook(out, read_only=True, data_only=True).worksheets
          if w.title.startswith("✅ Clean Inventory"))
nr = list(no.iter_rows(values_only=True)); n_after = len(nr) - 1
Ci = {h: i for i, h in enumerate(nr[0]) if h}
def cnt(title, iss, yr):
    return sum(1 for r in nr[1:]
               if str(r[Ci["Title"]] or "").strip() == title
               and str(r[Ci["Issue #"]] or "").strip() == iss
               and str(r[Ci["Year"]] or "").strip() == yr)
print(f"SOURCE: {os.path.basename(SRC)}\nOUTPUT: {os.path.basename(out)}")
print(f"rows: {n_before} -> {n_after}  (expected -2)")
print("renames applied:")
for r, a, b in renamed: print(f"   row {r}: {a!r} -> {b!r}")
print(f"Aliens vs. Avengers #1 (2024) count: {cnt('Aliens vs. Avengers','1','2024')}  (expected 1)")
print(f"Fantastic Four: Empyre #0 (2020) count: {cnt('Fantastic Four: Empyre','0','2020')}  (expected 1)")
