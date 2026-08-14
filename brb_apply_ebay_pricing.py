#!/usr/bin/env python3
"""
brb_apply_ebay_pricing.py — merge eBay Tier-2 results into Est. Raw Value (NM) $.

Only touches rows from the specific eBay run being applied (isolated by
fetched_at date, not the whole historical ebay_pricing_results.json cache),
matched by the JSON's own stored Box field — not a Title+Issue guess — so a
same-title-different-printing collision (like Doctor Strange Annual #1,
Box 36 vs Box 31) resolves unambiguously from the data the pricing run
actually used.

Never writes the source xlsx in place (CLAUDE.md write protocol) — always
produces a new output file for review before promotion.

Usage:
    python3 brb_apply_ebay_pricing.py <xlsx> --since 2026-07-14 --out <new.xlsx>
    python3 brb_apply_ebay_pricing.py <xlsx> --since 2026-07-14 --dry-run
"""
import argparse, json, os, sys
import openpyxl


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--results", default="ebay_pricing_results.json")
    ap.add_argument("--since", required=True, help="Only apply entries fetched on/after this date (YYYY-MM-DD)")
    ap.add_argument("--out", default=None)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    with open(args.results) as f:
        results = json.load(f)

    # Isolate this run's entries: fetched in the window, and successfully priced
    # (skip the "no results" nulls — nothing to apply there).
    applicable = {
        k: v for k, v in results.items()
        if v.get("fetched_at", "")[:10] >= args.since and v.get("median") is not None
    }
    print(f"Total cache entries: {len(results)}")
    print(f"Entries fetched since {args.since} with a real price: {len(applicable)}")

    wb = openpyxl.load_workbook(args.xlsx)
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("✅ Clean Inventory"))
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    ti  = header.index("Title") + 1
    ii  = header.index("Issue #") + 1
    bi  = header.index("Box #") + 1
    nmi = header.index("Est. Raw Value (NM) $") + 1

    def norm_issue(v):
        s = str(v).strip()
        try:
            f = float(s)
            return str(int(f)) if f == int(f) else s
        except ValueError:
            return s

    applied, not_found, ambiguous = [], [], []
    for key, v in applicable.items():
        title, issue, box = v["title"], norm_issue(v["issue"]), str(v["box"])
        median = v["median"]

        matches = []
        for row in ws.iter_rows(min_row=2):
            if str(row[ti - 1].value).strip() != title:
                continue
            if norm_issue(row[ii - 1].value) != issue:
                continue
            if str(row[bi - 1].value).strip() != box:
                continue
            matches.append(row)

        if not matches:
            not_found.append((title, issue, box))
            continue
        if len(matches) > 1:
            ambiguous.append((title, issue, box, len(matches)))
            continue

        row = matches[0]
        old_val = row[nmi - 1].value
        if str(old_val) == str(median):
            continue  # already matches, no-op
        applied.append((title, issue, box, old_val, median))
        if not args.dry_run:
            row[nmi - 1].value = median

    print(f"\nApplied: {len(applied)}")
    print(f"Not found in xlsx (title/issue/box no longer matches): {len(not_found)}")
    print(f"Ambiguous (multiple rows share Title+Issue+Box — skipped, needs manual review): {len(ambiguous)}")

    if ambiguous:
        print("\n-- ambiguous, skipped --")
        for t, i, b, n in ambiguous:
            print(f"  {t} #{i} (Box {b}): {n} matching rows")

    print("\n-- sample of applied changes (first 25) --")
    for t, i, b, old, new in applied[:25]:
        print(f"  {t} #{i} (Box {b}): {old!r} -> {new}")
    if len(applied) > 25:
        print(f"  ... and {len(applied) - 25} more")

    if args.dry_run:
        print("\n[DRY RUN] No file written.")
        return

    out = args.out or args.xlsx.replace(".xlsx", "_EBAY_APPLIED.xlsx")
    wb.save(out)
    print(f"\nWritten: {out}")


if __name__ == "__main__":
    main()
