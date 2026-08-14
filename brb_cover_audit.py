#!/usr/bin/env python3
"""brb_cover_audit.py — find covers whose stored date doesn't match the row's
year (wrong-volume/era match, e.g. a 2018 cover on a 2015 book). READ-ONLY."""
import glob,os,json,re,openpyxl
from openpyxl.styles import Font,PatternFill
covers=json.load(open("covers.json"))
x=max(glob.glob("attached_assets/comics_inventory_*.xlsx"),key=os.path.getmtime)
wb=openpyxl.load_workbook(x,read_only=True,data_only=True)
ws=next(wb[n] for n in wb.sheetnames if n.startswith("✅ Clean Inventory"))
rows=list(ws.iter_rows(values_only=True));H=list(rows[0]);C={n:H.index(n) for n in H if n}
def g(r,n):
    i=C.get(n);return r[i] if i is not None else None
def yr(v):
    n=[int(t) for t in re.findall(r"\d{4}",str(v or "")) if 1900<int(t)<2100];return n[0] if n else None
def ni(v):
    s=str(v).strip().lstrip("#")
    try:
        f=float(s);return str(int(f)) if f==int(f) else s
    except:return s
def cover(t,iss,vol):
    for k in (f"{t}|||{iss}|||{vol}",f"{t}|||{iss}",f"{t}|||#{iss}"):
        e=covers.get(k)
        if e and e.get("url"): return e,k
    return None,None
sus=[];seen=set()
for r in rows[1:]:
    t=str(g(r,'Title') or '').strip();iss=ni(g(r,'Issue #'));vol=str(g(r,'Volume') or '1').strip()
    ry=yr(g(r,'Year'))
    if not t or not ry: continue
    e,k=cover(t,iss,vol)
    if not e: continue
    cy=yr(e.get('date'))
    if not cy: continue          # no date to compare
    if abs(cy-ry)>2:
        key=(t,iss,vol)
        if key in seen: continue
        seen.add(key)
        sus.append((abs(cy-ry),t,g(r,'Issue #'),ry,cy,str(g(r,'Box #')),k,e.get('source','')))
sus.sort(reverse=True)
print(f"COVERS whose stored date is >2yr off the row's year (likely wrong): {len(sus)}")
print(f"  {'Δyr':>4} {'row yr':>6} {'cover yr':>8}  Title #issue (box)")
for d,t,i,ry,cy,box,k,src in sus[:30]:
    print(f"  {d:>4} {ry:>6} {cy:>8}  {t[:30]} #{i} (Box {box})")
out=openpyxl.Workbook();sh=out.active;sh.title="Wrong-era covers"
sh.append(["Title","Issue","Box","Row year","Cover date year","Δ years","cover source","key"])
for c in sh[1]: c.font=Font(bold=True,color="FFFFFF");c.fill=PatternFill("solid",fgColor="C00000")
for d,t,i,ry,cy,box,k,src in sus:
    sh.append([t,i,box,ry,cy,d,src,k])
out.save("cover_audit.xlsx")
print(f"\n  Written: cover_audit.xlsx ({len(sus)} rows)")
