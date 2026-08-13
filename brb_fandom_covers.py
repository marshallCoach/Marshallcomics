#!/usr/bin/env python3
"""
brb_fandom_covers.py — fetch covers straight from Marvel/DC Fandom for rows CV
cannot resolve (modern reused-title books: Titans 2024, Flash 2023, etc.).

Why this works NOW: the cover key is Title|||Issue|||Volume, and after the
volume audit the volumes are correct, so "{Title} Vol {vol} {issue}" is a real
Fandom page whose infobox Image is the cover. CV's search buries these under
reprints/foreign editions; Fandom addresses them directly by volume.

Method per missing-cover row:
  1. page = "{Title} Vol {volume} {issue}", try Marvel then DC Fandom
  2. read the infobox Image (Image1 / Image) filename from wikitext
  3. resolve File:<name> to its static URL via imageinfo
  4. write covers.json[Title|||Issue|||Volume] = {url, large, date, source}

Only fills rows whose cover is currently missing (null/absent under every key).
Never overwrites an existing URL. Writes covers.json + copies to public/.
Fandom = free (no Comic Vine / eBay quota).

Usage:
    python3 brb_fandom_covers.py            # fill all missing
    python3 brb_fandom_covers.py --limit 20 # small test
"""
import argparse, glob, json, os, re, time, urllib.parse, urllib.request, shutil
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
COVERS = os.path.join(ROOT, "covers.json")
PUBLIC = os.path.join(ROOT, "artifacts/comics-inventory/public/covers.json")
WIKIS = ["https://marvel.fandom.com/api.php", "https://dc.fandom.com/api.php"]
UA = "MarshallComicsInventory/1.0"
DELAY = 0.6


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def ni(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s); return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def nv(v):
    try:
        return str(int(float(str(v).strip())))
    except (ValueError, TypeError):
        return (str(v or "").strip() or "1")


def api(base, params):
    u = base + "?" + urllib.parse.urlencode(params)
    with urllib.request.urlopen(urllib.request.Request(u, headers={"User-Agent": UA}), timeout=30) as r:
        return json.load(r)


def fandom_cover(title, vol, issue):
    """Return (url, date, wiki) for {title} Vol {vol} {issue}, or None."""
    page = f"{title} Vol {vol} {issue}"
    for base in WIKIS:
        try:
            d = api(base, {"action": "parse", "page": page, "prop": "wikitext",
                           "format": "json", "formatversion": 2, "redirects": 1})
            time.sleep(DELAY)
            wt = d.get("parse", {}).get("wikitext", "")
            if not wt:
                continue
            m = re.search(r"\|\s*Image1?\s*=\s*([^\n|]+\.(?:jpg|png|jpeg))", wt, re.I)
            if not m:
                continue
            fn = m.group(1).strip()
            dm = re.search(r"\|\s*(?:ReleaseDate|CoverDate|Pubyear|Year)\s*=\s*([^\n|]+)", wt)
            date = dm.group(1).strip() if dm else ""
            ii = api(base, {"action": "query", "titles": "File:" + fn, "prop": "imageinfo",
                            "iiprop": "url", "format": "json", "formatversion": 2})
            time.sleep(DELAY)
            for p in ii.get("query", {}).get("pages", []):
                if "imageinfo" in p:
                    url = p["imageinfo"][0]["url"].split("/revision/")[0]
                    return (url, date, base.split("//")[1].split(".")[0])
        except Exception:
            time.sleep(DELAY)
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0)
    args = ap.parse_args()

    covers = json.load(open(COVERS))
    xlsx = latest_xlsx()
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = next(wb[n] for n in wb.sheetnames if (n == "Sheet X" or n.startswith("✅ Clean Inventory")))
    rows = list(ws.iter_rows(values_only=True))
    H = list(rows[0]); C = {n: H.index(n) for n in H if n}

    def g(r, n):
        i = C.get(n); return r[i] if i is not None else None

    def has_cover(t, iss, vol):
        for k in (f"{t}|||{iss}|||{vol}", f"{t}|||{iss}", f"{t}|||#{iss}"):
            e = covers.get(k)
            if e and e.get("url"):
                return True
        return False

    todo = []
    seen = set()
    for r in rows[1:]:
        t = str(g(r, "Title") or "").strip()
        iss = ni(g(r, "Issue #"))
        vol = nv(g(r, "Volume"))
        if not t or not iss:
            continue
        if has_cover(t, iss, vol):
            continue
        key = (t, iss, vol)
        if key in seen:
            continue
        seen.add(key)
        todo.append((t, iss, vol))
    if args.limit:
        todo = todo[: args.limit]
    print(f"Source: {os.path.basename(xlsx)}   missing-cover rows to try: {len(todo)}", flush=True)

    filled = 0
    for i, (t, iss, vol) in enumerate(todo, 1):
        res = fandom_cover(t, vol, iss)   # (title, vol, issue) — order matters!
        if res:
            url, date, wiki = res
            covers[f"{t}|||{iss}|||{vol}"] = {"url": url, "large": url, "date": date, "source": f"fandom-{wiki}"}
            filled += 1
            print(f"  [{i}/{len(todo)}] ✓ {t} Vol {vol} #{iss}  ({wiki})", flush=True)
        if i % 20 == 0:
            json.dump(covers, open(COVERS, "w"))
            print(f"  ...flushed at {i}, filled {filled}", flush=True)

    json.dump(covers, open(COVERS, "w"))
    shutil.copy(COVERS, PUBLIC)
    print(f"\n  Filled {filled} / {len(todo)} covers from Fandom.", flush=True)
    print(f"  Wrote {COVERS} (+ copied to public/). Run: node gen_data.mjs", flush=True)


if __name__ == "__main__":
    main()
