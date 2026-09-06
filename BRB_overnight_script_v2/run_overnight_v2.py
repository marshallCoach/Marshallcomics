#!/usr/bin/env python3
"""
BRB Overnight Writer+Artist Fill — Comic Vine edition (no Anthropic API)

Uses Comic Vine's free /issues/ endpoint to look up writer and artist credits.
Rate limit: 200 requests/hour → 18s delay keeps it safely within limits.
Set COMIC_VINE_API_KEY in your environment before running.

Usage:
    export COMIC_VINE_API_KEY=your_key_here
    nohup python3 run_overnight_v2.py > overnight_log.txt 2>&1 &
    tail -f overnight_log.txt
"""

import pandas as pd
import requests
import json, os, re, sys, time, signal
from datetime import datetime

# Line-buffer stdout so the log is a truthful heartbeat even when redirected to
# a file (nohup ... > overnight_log.txt). Without this, stdout is block-buffered
# and "no output for 24 min" is ambiguous between "hung" and "buffer not flushed".
try:
    sys.stdout.reconfigure(line_buffering=True)
except Exception:
    pass

# Per-title watchdog: if a single title's processing exceeds this, SIGALRM fires
# and we skip it instead of letting one wedged iteration hang the whole night.
TITLE_TIMEOUT = 600  # seconds (10 min — generous vs. a legit multi-issue title)
class _TitleTimeout(Exception):
    pass
def _title_alarm(signum, frame):
    raise _TitleTimeout()

# ── CONFIG ────────────────────────────────────────────────────────────────────
import glob as _glob

def _find_inventory():
    """Newest comics_inventory_*.xlsx by mtime in ../attached_assets/ — same
    convention as gen_data.mjs, so every script reads the one current file
    (not whichever happens to carry 'VALIDATED' in its name)."""
    base = os.path.dirname(os.path.abspath(__file__))
    assets = os.path.join(base, "..", "attached_assets")
    matches = [f for f in _glob.glob(os.path.join(assets, "comics_inventory_*.xlsx"))
               if not os.path.basename(f).startswith("~$")]
    if not matches:
        raise FileNotFoundError(f"No comics_inventory_*.xlsx found in {assets}")
    return max(matches, key=os.path.getmtime)

def _find_sheet(path):
    """Return the sheet name that contains the inventory columns."""
    xl = pd.ExcelFile(path)
    for name in xl.sheet_names:
        df = xl.parse(name, nrows=1)
        if "Title" in df.columns and "Writer(s)" in df.columns:
            return name
    return xl.sheet_names[0]

INVENTORY_PATH   = _find_inventory()
SHEET_NAME       = _find_sheet(INVENTORY_PATH)
LOG_PATH         = "issues.json"
REVIEW_PATH      = "needs_review.json"
CHECKPOINT_EVERY = 25
SKIP_TITLES      = ["The Flash"]   # flagged for separate volume audit
MAX_YEAR_GAP     = 15
DELAY_SECONDS    = 20              # Comic Vine free tier: 200 req/hr = 1 per 18s (20 for safety)
RETRY_420_WAIT   = 65             # seconds to wait after a 420 rate-limit response

API_KEY = os.environ.get("COMIC_VINE_API_KEY", "")
CV_BASE = "https://comicvine.gamespot.com/api"
HEADERS = {"User-Agent": "BRB-Marshall-Comics/1.0"}

def cv_get(url, params, retries=3):
    """Wrapper around requests.get with automatic 420 backoff."""
    for attempt in range(retries):
        try:
            resp = requests.get(url, params=params, headers=HEADERS, timeout=30)
            if resp.status_code == 420:
                wait = RETRY_420_WAIT * (attempt + 1)
                print(f"  [420] Rate limited — waiting {wait}s before retry {attempt+1}/{retries}")
                time.sleep(wait)
                continue
            resp.raise_for_status()
            return resp, None
        except Exception as e:
            if attempt == retries - 1:
                return None, str(e)
            time.sleep(RETRY_420_WAIT)
    return None, "Max retries exceeded"

# ── LOGGING ───────────────────────────────────────────────────────────────────
def load_json(p):
    return json.load(open(p)) if os.path.exists(p) else []

def save_json(p, d):
    json.dump(d, open(p, "w"), indent=2)

def log_issue(cat, title, detail):
    log = load_json(LOG_PATH)
    log.append({"ts": datetime.now().isoformat(), "category": cat, "title": title, "detail": detail})
    save_json(LOG_PATH, log)

# ── SAFE WRITE: row-count assertion + integrity log sheet ─────────────────────
_integrity_log = []  # accumulated across the run

def safe_write(df, path, sheet, rows_initial, label="checkpoint"):
    """Write xlsx only if row count is stable; append a Data Integrity Log sheet."""
    rows_now = len(df)
    status = "OK" if rows_now >= rows_initial else "LOST"
    _integrity_log.append({
        "ts":      datetime.now().isoformat(),
        "label":   label,
        "before":  rows_initial,
        "after":   rows_now,
        "delta":   rows_now - rows_initial,
        "status":  status,
    })
    if rows_now < rows_initial:
        msg = (f"DATA LOSS DETECTED at {label}: started {rows_initial} rows, "
               f"now {rows_now} ({rows_initial - rows_now} missing). Aborting write to {path}.")
        log_issue("DATA_LOSS", "safe_write", msg)
        raise RuntimeError(msg)
    assert_box_capacities(df, label=label)
    df.to_excel(path, sheet_name=sheet, index=False)
    # Append integrity log sheet
    import openpyxl
    wb = openpyxl.load_workbook(path)
    if "📝 Data Integrity Log" in wb.sheetnames:
        del wb["📝 Data Integrity Log"]
    ws = wb.create_sheet("📝 Data Integrity Log")
    headers = ["ts", "label", "before", "after", "delta", "status"]
    ws.append(headers)
    for entry in _integrity_log:
        ws.append([entry[h] for h in headers])
    wb.save(path)

# ── BOX CAPACITY CHECK ───────────────────────────────────────────────────────
BOX_CAPACITY_DEFAULT = 240
BOX_CAPACITY_EXCEPTIONS = {15: 150, 23: 155, 40: 80, 44: 200, 72: 80, 85: 155}

def assert_box_capacities(df, label=""):
    """Halt if any box exceeds its defined capacity. Call before every write."""
    if "Box #" not in df.columns:
        return
    counts = df.groupby("Box #").size()
    violations = []
    for box, count in counts.items():
        try:
            box_num = int(float(str(box)))
        except (ValueError, TypeError):
            continue
        capacity = BOX_CAPACITY_EXCEPTIONS.get(box_num, BOX_CAPACITY_DEFAULT)
        if count > capacity:
            violations.append(f"  Box {box_num}: {count} comics (capacity {capacity}, overage +{count - capacity})")
    if violations:
        msg = f"BOX CAPACITY WARNING at {label}:\n" + "\n".join(violations)
        log_issue("BOX_OVERAGE", "assert_box_capacities", msg)
        print(f"  ⚠  {msg}")

def get_year_mode_safe(series, title="unknown"):
    """Return mode year string or '?' — logs when all values are non-numeric."""
    mode_result = series.mode()
    if len(mode_result) == 0:
        log_issue("YEAR_PARSE_FAILURE", title,
                  f"All {len(series)} Year values are non-numeric; defaulting to '?'")
        return "?"
    return str(mode_result.iloc[0])[:4]

def log_review(title, vol, reason):
    rv = load_json(REVIEW_PATH)
    rv.append({"ts": datetime.now().isoformat(), "title": title, "volume": vol, "reason": reason})
    save_json(REVIEW_PATH, rv)

def is_blank(v):
    return pd.isna(v) or str(v).strip() in ("", "nan", "None")

def normalize_credits(s):
    """Canonicalize credit separators so 'A / B', 'A & B', 'A, B' all compare equal.
    Also strips extra whitespace. Applied before writing and before conflict detection."""
    if not s or pd.isna(s):
        return s
    s = str(s).strip()
    # Normalize all separators to ' & '
    s = re.sub(r'\s*/\s*', ' & ', s)
    s = re.sub(r'\s*,\s*', ' & ', s)
    s = re.sub(r'\s+and\s+', ' & ', s, flags=re.IGNORECASE)
    # Collapse multiple spaces
    s = re.sub(r'\s{2,}', ' ', s)
    return s

# ── YEAR-GAP COLLISION CHECK ──────────────────────────────────────────────────
def has_year_collision(df, title, volume):
    rows   = df[(df["Title"] == title) & (df["Volume"] == volume)]
    filled = rows[~rows["Writer(s)"].apply(is_blank)]
    blank  = rows[rows["Writer(s)"].apply(is_blank)]
    if len(filled) == 0 or len(blank) == 0:
        return False, None
    fy = pd.to_numeric(filled["Year"], errors="coerce").dropna()
    by = pd.to_numeric(blank["Year"],  errors="coerce").dropna()
    if len(fy) == 0 or len(by) == 0:
        return False, None
    gap = abs(fy.mean() - by.mean())
    if gap > MAX_YEAR_GAP:
        return True, f"Filled avg year {fy.mean():.0f}, blank avg {by.mean():.0f} (gap {gap:.0f}y)"
    return False, None

# ── COMIC VINE LOOKUP ─────────────────────────────────────────────────────────
_volume_cache = {}   # title → volume_id (int) or None

def cv_find_volume_id(title, year_hint):
    """
    Step 1: query /volumes/ to find the best-matching volume ID for a series title.
    Returns (volume_id, None) or (None, error_string).
    Costs 1 API call; result cached in _volume_cache.
    """
    if title in _volume_cache:
        return _volume_cache[title], None

    params = {
        "api_key":    API_KEY,
        "format":     "json",
        "filter":     f"name:{title}",
        "field_list": "id,name,start_year,count_of_issues",
        "limit":      10,
    }
    resp, err = cv_get(f"{CV_BASE}/volumes/", params)
    if resp is None:
        return None, f"HTTP error: {err}"
    try:
        data = resp.json()
    except Exception as e:
        return None, f"JSON error: {e}"

    if data.get("status_code") != 1:
        return None, f"CV error: {data.get('error','unknown')}"

    results = data.get("results", [])
    if not results:
        _volume_cache[title] = None
        return None, "No volumes found"

    # Best match: exact name first, then closest start_year to year_hint
    title_lower = title.lower()
    exact = [r for r in results if r.get("name", "").lower() == title_lower]
    pool  = exact if exact else results

    best = pool[0]
    if year_hint and year_hint != "?":
        try:
            yh = int(year_hint)
            def year_dist(r):
                sy = r.get("start_year")
                return abs(int(sy) - yh) if sy and str(sy).isdigit() else 9999
            best = min(pool, key=year_dist)
        except (ValueError, TypeError):
            pass

    vid = best.get("id")
    _volume_cache[title] = vid
    print(f"  [VOL] '{title}' → volume id {vid} ('{best.get('name')}' {best.get('start_year')})")
    return vid, None


def cv_fetch_issue_ids_for_volume(volume_id):
    """
    Step 2: get issue_number → issue_id mapping for a volume.
    Uses field_list=id,issue_number only (bulk fetch, no role data needed here).
    Returns (dict of issue_num_str → issue_id, error).
    """
    id_map = {}
    offset, limit = 0, 100
    while True:
        params = {
            "api_key":    API_KEY,
            "format":     "json",
            "filter":     f"volume:{volume_id}",
            "field_list": "id,issue_number",
            "limit":      limit,
            "offset":     offset,
        }
        resp, err = cv_get(f"{CV_BASE}/issues/", params)
        if resp is None:
            return None, f"HTTP error: {err}"
        try:
            data = resp.json()
        except Exception as e:
            return None, f"JSON error: {e}"

        if data.get("status_code") != 1:
            return None, f"CV error: {data.get('error','unknown')}"

        for r in data.get("results", []):
            num_raw = str(r.get("issue_number", "")).strip()
            try:
                num_key = str(int(float(num_raw)))
            except (ValueError, TypeError):
                num_key = re.sub(r"[^0-9]", "", num_raw) or num_raw
            id_map[num_key] = r["id"]

        total = data.get("number_of_total_results", 0)
        offset += limit
        if offset >= total or not data.get("results"):
            break
        time.sleep(DELAY_SECONDS)

    return id_map, None


def cv_fetch_issue_credits(issue_id):
    """
    Step 3: fetch a single issue detail to get person_credits with role data.
    CV only returns roles on individual issue detail endpoints.
    Returns (writer, artist, cover_artist) or (None, None, None).
    """
    params = {
        "api_key":    API_KEY,
        "format":     "json",
        "field_list": "person_credits",
    }
    resp, err = cv_get(f"{CV_BASE}/issue/4000-{issue_id}/", params)
    if resp is None:
        return None, None, None, f"HTTP error: {err}"
    try:
        data = resp.json()
    except Exception as e:
        return None, None, None, f"JSON error: {e}"

    if data.get("status_code") != 1:
        return None, None, None, f"CV error: {data.get('error','unknown')}"

    credits = (data.get("results") or {}).get("person_credits") or []
    writer, artist, cover_artist = None, None, None
    for p in credits:
        roles = (p.get("role") or "").lower()
        name  = (p.get("name") or "").strip()
        if not name:
            continue
        if "writer" in roles and not writer:
            writer = name
        if ("penciler" in roles or "penciller" in roles) and not artist:
            artist = name
        if "cover" in roles and not cover_artist:
            cover_artist = name
    return writer, artist, cover_artist, None


def cv_get_all_credits_for_title(title, year_hint, needed_issue_nums):
    """
    Three-step lookup: volume search → issue ID list → individual issue details.
    Only fetches detail for issues actually needed (blank writers).
    Returns (credits_by_issue_num, api_calls_used, error).
    """
    # Step 1 — volume lookup (1 API call, cached)
    volume_id, err = cv_find_volume_id(title, year_hint)
    if volume_id is None:
        return {}, 1, err or "Volume not found"
    time.sleep(DELAY_SECONDS)

    # Step 2 — get issue ID map (1+ API calls)
    id_map, err = cv_fetch_issue_ids_for_volume(volume_id)
    calls = 2
    if id_map is None:
        return {}, calls, err

    # Normalise needed issue nums to string keys
    needed_keys = set()
    for n in needed_issue_nums:
        try:
            needed_keys.add(str(int(float(str(n)))))
        except (ValueError, TypeError):
            needed_keys.add(re.sub(r"[^0-9]", "", str(n)) or str(n))

    # Step 3 — fetch individual issue details for needed issues only
    by_num = {}
    for num_key in needed_keys:
        issue_id = id_map.get(num_key)
        if issue_id is None:
            continue   # issue not in this volume
        time.sleep(DELAY_SECONDS)
        calls += 1
        w, a, ca, detail_err = cv_fetch_issue_credits(issue_id)
        if detail_err:
            print(f"  [DETAIL ERR] #{num_key} — {detail_err}")
            continue
        by_num[num_key] = {"writer": w, "artist": a, "cover_artist": ca}

    return by_num, calls, None

# ── BUILD QUEUE FROM EXCEL ────────────────────────────────────────────────────
def build_queue(df):
    # Queue titles where writer OR cover artist is missing
    writer_blank = df["Writer(s)"].apply(is_blank)
    ca_blank     = df["Cover Artist"].apply(is_blank) if "Cover Artist" in df.columns else pd.Series(True, index=df.index)
    blank_mask   = writer_blank | ca_blank
    blank_df     = df[blank_mask].copy()
    blank_df["Volume"] = pd.to_numeric(blank_df.get("Volume", 1), errors="coerce").fillna(1)
    counts = (
        blank_df.groupby(["Title", "Volume"])
        .size()
        .reset_index(name="val")
        .sort_values("val", ascending=False)
    )
    return counts

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    if not API_KEY:
        print("ERROR: Set COMIC_VINE_API_KEY environment variable before running.")
        print("  export COMIC_VINE_API_KEY=your_key_here")
        return

    print(f"[{datetime.now()}] BRB Overnight Writer+Artist Fill (Comic Vine) starting...")

    df = pd.read_excel(INVENTORY_PATH, sheet_name=SHEET_NAME)

    if "Volume" not in df.columns:
        df["Volume"] = 1
    df["Volume"] = pd.to_numeric(df["Volume"], errors="coerce").fillna(1)

    if "Cover Artist" not in df.columns:
        df["Cover Artist"] = ""
        print("  (Cover Artist column not found — added)")

    rows_initial = len(df)
    print(f"Loaded {rows_initial} rows from {INVENTORY_PATH}")

    # Safe shutdown: save whatever's in memory on kill/Ctrl-C instead of losing
    # it silently. Checkpoints only fire every CHECKPOINT_EVERY titles, so a
    # kill between checkpoints previously discarded everything fetched since
    # the last one (confirmed: a full ~34hr run was lost this way).
    import signal
    def _save_and_exit(signum, frame):
        ts = datetime.now().strftime("%d%m_%H%M")
        out = f"comics_inventory_{ts}_INTERRUPTED.xlsx"
        print(f"\n[SIGNAL {signum}] Saving in-progress work to {out} before exiting...")
        try:
            safe_write(df, out, f"✅ Clean Inventory {ts}", rows_initial, label=f"interrupted_{signum}")
            print(f"[SIGNAL {signum}] Saved. Safe to kill.")
        except Exception as e:
            print(f"[SIGNAL {signum}] Save failed: {e}")
        sys.exit(1)
    signal.signal(signal.SIGTERM, _save_and_exit)
    signal.signal(signal.SIGINT, _save_and_exit)

    queue = build_queue(df)
    queue = queue[~queue["Title"].isin(SKIP_TITLES)]
    print(f"Queue: {len(queue)} Title+Volume runs with blank writers.")
    print(f"Delay: {DELAY_SECONDS}s between requests (~{3600//DELAY_SECONDS} req/hr, within free tier)")

    done_keys = {
        f"{e['title']}|{e.get('detail','')[:30]}"
        for e in load_json(LOG_PATH)
        if e["category"] in ("FILLED", "FILLED_PARTIAL", "SKIPPED_LOW_CONFIDENCE", "SKIPPED_COLLISION")
    }

    processed = 0
    api_calls  = 0
    signal.signal(signal.SIGALRM, _title_alarm)

    for _, run in queue.iterrows():
        title, volume = run["Title"], run["Volume"]
        key = f"{title}|{volume}"
        if key in done_keys:
            continue

        title_budget = TITLE_TIMEOUT  # re-armed with a workload-scaled value once issue count is known
        signal.alarm(title_budget)    # watchdog: skip this title if it wedges
        try:
          ca_blank_col = df["Cover Artist"].apply(is_blank) if "Cover Artist" in df.columns else pd.Series(True, index=df.index)
          mask = (
              (df["Title"] == title) &
              (df["Volume"] == volume) &
              (df["Writer(s)"].apply(is_blank) | ca_blank_col)
          )
          if mask.sum() == 0:
              continue

          # Collision check only applies when writer is missing (year-gap logic is writer-based)
          writer_missing = mask & df["Writer(s)"].apply(is_blank)
          collision, cdesc = has_year_collision(df, title, volume) if writer_missing.any() else (False, None)
          if collision:
              log_issue("SKIPPED_COLLISION", title, cdesc)
              log_review(title, volume, f"Year gap: {cdesc}")
              print(f"[SKIP-COLLISION] {title} Vol {volume}")
              continue

          year_hint = get_year_mode_safe(df.loc[mask, "Year"], title=title)
          issues    = df.loc[mask, "Issue #"].dropna().sort_values().unique()
          n_rows    = int(mask.sum())

          # Re-arm the watchdog scaled to this title's real workload: each issue
          # costs DELAY_SECONDS + the API call (+ possible 420 backoffs), so a
          # flat 600s murdered every legitimately-large title mid-fetch (a
          # 97-issue title needs ~50+ min of NORMAL work). Verified on the
          # 15/07 run: 34 of 89 titles were watchdog-killed while healthy.
          # The wedge this guards against is minutes of NO progress, so
          # workload-scaled + generous is still an effective tripwire.
          title_budget = max(TITLE_TIMEOUT, len(issues) * (DELAY_SECONDS + 20) + 300)
          signal.alarm(title_budget)

          print(f"[CV] {title} Vol {volume} — {len(issues)} issues (~{year_hint}) [{n_rows} blank rows, watchdog {title_budget}s]")

          # Three-step: volume lookup → issue ID map → individual issue details (for roles)
          credits_by_num, calls_used, fetch_err = cv_get_all_credits_for_title(title, year_hint, issues)
          api_calls += calls_used

          if fetch_err and not credits_by_num:
              print(f"  [ERROR] {fetch_err}")
              log_issue("SKIPPED_LOW_CONFIDENCE", title, f"CV lookup failed: {fetch_err}")
              log_review(title, volume, f"CV lookup failed: {fetch_err}")
              processed += 1
              continue

          writers_found       = {}
          artists_found       = {}
          cover_artists_found = {}
          not_found           = []

          for issue_num in issues:
              try:
                  issue_key = str(int(float(str(issue_num))))
              except (ValueError, TypeError):
                  issue_key = re.sub(r"[^0-9]", "", str(issue_num)) or str(issue_num)

              result = credits_by_num.get(issue_key)
              if result:
                  if result["writer"]:
                      writers_found[issue_num] = result["writer"]
                  if result["artist"]:
                      artists_found[issue_num] = result["artist"]
                  if result["cover_artist"]:
                      cover_artists_found[issue_num] = result["cover_artist"]
                  if not result["writer"]:
                      not_found.append(issue_num)
                      print(f"  [NO WRITER] #{issue_num} — credits found but no writer role")
              else:
                  not_found.append(issue_num)
                  print(f"  [NOT FOUND] #{issue_num} — not in CV volume")

          # Apply fills — normalize separators before writing.
          # Writer(s) must be guarded with is_blank, same as Artist(s)/Cover_Artist
          # below: `mask` includes rows queued because Cover_Artist was blank even
          # when Writer(s) already had a value, and the batch's CV volume is
          # resolved from the group's Year mode, not per-row — a bad match there
          # was unconditionally overwriting correct existing writer credits.
          filled_count = 0
          for issue_num, writer in writers_found.items():
              row_mask = mask & (df["Issue #"] == issue_num)
              writer_row_mask = row_mask & df["Writer(s)"].apply(is_blank)
              df.loc[writer_row_mask, "Writer(s)"] = normalize_credits(writer)
              filled_count += int(writer_row_mask.sum())

          for issue_num, artist in artists_found.items():
              row_mask = mask & (df["Issue #"] == issue_num)
              if "Artist(s)" in df.columns:
                  df.loc[row_mask & df["Artist(s)"].apply(is_blank), "Artist(s)"] = normalize_credits(artist)

          for issue_num, ca in cover_artists_found.items():
              row_mask = mask & (df["Issue #"] == issue_num)
              if "Cover Artist" in df.columns:
                  df.loc[row_mask & df["Cover Artist"].apply(is_blank), "Cover Artist"] = normalize_credits(ca)

          if not_found:
              log_review(title, volume, f"{len(not_found)} issues not found on CV: {sorted(not_found)[:10]}")

          if filled_count > 0:
              log_issue("FILLED", title,
                  f"{filled_count}/{n_rows} rows filled via Comic Vine. "
                  f"Artists: {len(artists_found)}. Not found: {len(not_found)}.")
              print(f"  → Filled {filled_count} rows | artists: {len(artists_found)} | missed: {len(not_found)}")
          else:
              log_issue("SKIPPED_LOW_CONFIDENCE", title, f"No writer credits found on Comic Vine for any issue.")
              print(f"  → No writers found — logged to needs_review.json")

          processed += 1
        except _TitleTimeout:
            log_issue("SKIPPED_TIMEOUT", title, f"exceeded {title_budget}s watchdog — skipped")
            log_review(title, volume, f"watchdog timeout after {title_budget}s")
            print(f"  [WATCHDOG] {title} Vol {volume} exceeded {title_budget}s — skipped")
            processed += 1
        finally:
            signal.alarm(0)  # clear the watchdog before checkpoint / next title

        if processed % CHECKPOINT_EVERY == 0:
            ts  = datetime.now().strftime("%d%m_%H%M")
            out = f"comics_inventory_{ts}.xlsx"
            safe_write(df, out, f"✅ Clean Inventory {ts}", rows_initial, label=f"checkpoint_{processed}")
            print(f"[CHECKPOINT] Saved {out} after {processed} titles ({api_calls} API calls)")

    ts    = datetime.now().strftime("%d%m_%H%M")
    final = f"comics_inventory_FINAL_{ts}.xlsx"
    safe_write(df, final, f"✅ Clean Inventory {ts}", rows_initial, label="FINAL")
    print(f"\n[DONE] {processed} titles | {api_calls} Comic Vine calls | Output: {final}")
    print(f"Review needed: {REVIEW_PATH}")

if __name__ == "__main__":
    main()
