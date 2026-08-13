#!/usr/bin/env python3
"""
brb_reconcile_volumes.py — confirm the Volume column against Comic Vine's own
volume identity, deterministically, with ZERO extra API calls and ZERO AI.

How it works (the "piggyback" approach):
  1. Every cover fetch already captured Comic Vine's volume_id + cover_date into
     covers.json (see covers.ts / fetchCovers.mjs). No new pass is needed.
  2. For each Title, cluster its inventory rows by the CV volume_id, order the
     clusters by earliest cover_date, and auto-number them Vol 1, 2, 3, …
     (publishers number chronologically, so this matches official numbering the
     large majority of the time).
  3. Compare that derived number to the inventory's Volume column. Agreements are
     confirmed and need nothing. DISAGREEMENTS are the worklist — usually small,
     and the only tier worth a human/AI/GCD lookup.

Requires covers.json entries to carry volume_id — i.e. run a cover fetch with the
updated code first (pnpm exec tsx fetchCovers.mjs --retry-nulls, or a batch). Rows
without a volume_id yet are reported as "no CV volume data" (not as mismatches).

Usage:
    python3 brb_reconcile_volumes.py [xlsx]           # report
    python3 brb_reconcile_volumes.py --csv out.csv    # also write the worklist
"""
import sys, os, json, glob, re, argparse
from collections import defaultdict

ROOT = os.path.dirname(os.path.abspath(__file__))
COVERS = os.path.join(ROOT, "covers.json")
ASSETS = os.path.join(ROOT, "attached_assets")


def latest_xlsx():
    # Newest by mtime, same convention as gen_data.mjs (the reingest source) so
    # this reconciles against the exact file the app data reflects.
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx"))
         if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def norm_issue(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else str(f)
    except ValueError:
        return s


def cover_lookup(covers, title, issue, volume):
    """Find the covers.json entry for a row, trying the same key forms the app
    uses (3-part with the row's Volume, then 2-part with/without '#')."""
    vol = str(volume or "1").strip()
    cands = [
        f"{title}|||{issue}|||{vol}",
        f"{title}|||{issue}",
        f"{title}|||#{issue}",
        f"{title}|||{norm_issue(issue)}",
    ]
    for k in cands:
        if k in covers and covers[k]:
            return covers[k]
    return None


def derive_volume_numbers(items):
    """Given a title's rows (each a dict with 'vid' and 'cdate'), return a map
    {vid -> derived_volume_number}: distinct CV volumes ordered by earliest
    cover_date get 1, 2, 3, … Pure/deterministic — the testable core."""
    first_date = {}
    for it in items:
        if it.get("vid") is None:
            continue
        d = it.get("cdate") or "9999"
        if it["vid"] not in first_date or d < first_date[it["vid"]]:
            first_date[it["vid"]] = d
    ordered = sorted(first_date, key=lambda v: (first_date[v], str(v)))
    return {vid: i + 1 for i, vid in enumerate(ordered)}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=None)
    ap.add_argument("--csv", default=None, help="write the mismatch worklist to this CSV")
    args = ap.parse_args()

    xlsx = os.path.abspath(args.xlsx) if args.xlsx else latest_xlsx()
    if not xlsx or not os.path.exists(xlsx):
        print(f"ERROR: no xlsx found (looked in {ASSETS})"); sys.exit(1)
    if not os.path.exists(COVERS):
        print(f"ERROR: covers.json not found at {COVERS}"); sys.exit(1)

    import openpyxl
    covers = json.load(open(COVERS))
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = next((wb[n] for n in wb.sheetnames if (n == "Sheet X" or n.startswith("✅ Clean Inventory"))), None)
    if ws is None:
        print("ERROR: inventory sheet not found"); sys.exit(1)
    rows = list(ws.iter_rows(values_only=True))
    H = list(rows[0])
    ti, ii = H.index("Title"), H.index("Issue #")
    vi = H.index("Volume") if "Volume" in H else None
    bi = H.index("Box #") if "Box #" in H else None

    # Gather per-title rows with their CV volume identity + cover_date.
    by_title = defaultdict(list)
    have_cv = 0
    for r in rows[1:]:
        title = str(r[ti] or "").strip()
        if not title:
            continue
        issue = norm_issue(r[ii])
        declared = str(r[vi] or "").strip() if vi is not None else ""
        entry = cover_lookup(covers, title, issue, declared)
        vid = entry.get("volume_id") if entry else None
        cdate = entry.get("date") if entry else None
        if vid is not None:
            have_cv += 1
        by_title[title].append({
            "issue": issue, "declared": declared, "vid": vid, "cdate": cdate or "",
            "box": str(r[bi] or "").strip() if bi is not None else "",
        })

    total_rows = sum(len(v) for v in by_title.values())
    confirmed = mismatches = no_cv = single_vol = 0
    worklist = []

    for title, items in by_title.items():
        vids = {it["vid"] for it in items if it["vid"] is not None}
        if not vids:
            no_cv += len(items)
            continue
        if len(vids) == 1:
            # One CV volume for the whole title — every row should read the same
            # volume. Derived number is 1 (or the single declared value if set).
            single_vol += 1
        # Order distinct CV volumes chronologically by earliest cover_date.
        derived_num = derive_volume_numbers(items)

        for it in items:
            if it["vid"] is None:
                no_cv += 1
                continue
            dnum = derived_num[it["vid"]]
            decl = it["declared"]
            # Compare as integers where possible.
            try:
                agree = int(float(decl)) == dnum
            except (ValueError, TypeError):
                agree = (decl == str(dnum))
            if agree:
                confirmed += 1
            else:
                mismatches += 1
                worklist.append({
                    "Title": title, "Issue": it["issue"], "Box": it["box"],
                    "Declared_Volume": decl or "(blank)", "CV_Derived_Volume": dnum,
                    "CV_volume_id": it["vid"], "cover_date": it["cdate"],
                })

    print(f"\n{'='*64}")
    print(f"  VOLUME RECONCILIATION — {os.path.basename(xlsx)}")
    print(f"{'='*64}")
    print(f"  Inventory rows                 : {total_rows:,}")
    print(f"  Rows with CV volume_id captured: {have_cv:,}  ({100*have_cv/total_rows:.1f}%)")
    print(f"  {'-'*60}")
    print(f"  ✓ Confirmed (declared == CV-derived) : {confirmed:,}")
    print(f"  ✗ Mismatches (the worklist)          : {mismatches:,}")
    print(f"  · No CV volume data yet (re-fetch)    : {no_cv:,}")
    if have_cv == 0:
        print(f"\n  ⚠ No rows have volume_id yet. Run a cover fetch with the updated")
        print(f"    code first:  pnpm exec tsx fetchCovers.mjs --retry-nulls")
    if worklist:
        print(f"\n  ── WORKLIST (first 25) ──────────────────────────────────────")
        for w in sorted(worklist, key=lambda x: x["Title"])[:25]:
            print(f"    {w['Title']} #{w['Issue']} (Box {w['Box']}): "
                  f"declared Vol {w['Declared_Volume']} → CV-derived Vol {w['CV_Derived_Volume']}")
        if len(worklist) > 25:
            print(f"    ... and {len(worklist)-25} more")
    print(f"{'='*64}\n")

    if args.csv and worklist:
        import csv
        with open(args.csv, "w", newline="") as f:
            wr = csv.DictWriter(f, fieldnames=list(worklist[0].keys()))
            wr.writeheader(); wr.writerows(worklist)
        print(f"Worklist written to {args.csv} ({len(worklist)} rows)")


if __name__ == "__main__":
    main()
