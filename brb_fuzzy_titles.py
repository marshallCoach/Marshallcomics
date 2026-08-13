#!/usr/bin/env python3
# ⚠ RESULT: THIS APPROACH DOES NOT WORK for the Marshall inventory. Tested 2107.
# Loose scoring -> 201 "high-confidence" matches that are mostly garbage (foreign-
# language GCD duplicates, shared-common-word neighbours like every "Free Comic
# Book Day: X" -> "Doctor Who: Free Comic Book Day"). Strict string-similarity ->
# 8 matches, most of which are genuinely DIFFERENT books (Ultimate Avengers 2/3,
# ST:TNG-Hive). Root cause: the unmatched titles are mostly NEW books absent from
# GCD, so there is no correct answer to match against, and the real misspellings
# are homophone/semantic (Muse->Mutant, Theory->Canary) which string distance
# cannot reverse. Use Comic Vine search as the oracle instead. Kept for the record.
"""
brb_fuzzy_titles.py — propose likely-correct titles for inventory titles that
GCD can't match, by blended fuzzy matching against the GCD catalog.

Reproduces the kinds of fixes Roberto made by hand (compound-split, homophone,
and->or, initialism spacing, wrong subtitle) FOR BOOKS GCD ACTUALLY HAS. New
2024-2025 books aren't in the local dump, so no local method can guess them —
those are flagged 'no-corpus' for a Comic Vine pass instead.

Score per candidate = max of:
  - difflib ratio on the normalized full title
  - token Jaccard (shared words / union) — survives word-order + subtitle swaps
  - anchor bonus: a shared distinctive multi-word run (e.g. "best of the best")
Only distinct-name GCD series are candidates. Output is a review xlsx with a
confidence tier; nothing is applied.

Usage:
    python3 brb_fuzzy_titles.py                 # newest xlsx, all unmatched titles
    python3 brb_fuzzy_titles.py --out guesses.xlsx
"""
import argparse, glob, os, re, sqlite3, difflib
from collections import defaultdict
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
DB = os.path.join(ROOT, "gcd_local.sqlite")
STOP = {"the", "a", "an", "of", "and", "vs", "no"}


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def tight(t):
    return re.sub(r"[^a-z0-9]", "", re.sub(r"&", " and ", str(t or "").lower()))


def norm(t):
    return re.sub(r"[^a-z0-9]+", " ", str(t or "").lower()).strip()


def toks(t):
    return {w for w in norm(t).split() if w not in STOP and len(w) > 1}


def anchor_bonus(a, b):
    # longest shared word-run of length >=2 gives a strong signal (subtitles)
    wa, wb = norm(a).split(), norm(b).split()
    best = 0
    for i in range(len(wa)):
        for j in range(len(wb)):
            k = 0
            while i + k < len(wa) and j + k < len(wb) and wa[i + k] == wb[j + k]:
                k += 1
            best = max(best, k)
    return 0.15 * best if best >= 2 else 0.0


def score(inv, cand):
    d = difflib.SequenceMatcher(None, norm(inv), norm(cand)).ratio()
    ta, tb = toks(inv), toks(cand)
    j = len(ta & tb) / len(ta | tb) if (ta | tb) else 0
    return min(1.0, max(d, j) + anchor_bonus(inv, cand))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=None)
    ap.add_argument("--out", default="fuzzy_title_guesses.xlsx")
    args = ap.parse_args()

    xlsx = os.path.abspath(args.xlsx) if args.xlsx else latest_xlsx()
    print(f"Source: {os.path.basename(xlsx)}")

    conn = sqlite3.connect(DB)
    have = {r[0] for r in conn.execute("SELECT DISTINCT matched_title FROM gcd_series")}
    have_tight = {tight(x) for x in have}
    try:
        alias_t = {tight(a) for a, in conn.execute("SELECT alias FROM gcd_title_alias")}
    except sqlite3.OperationalError:
        alias_t = set()
    cand_names = sorted({r[0] for r in conn.execute("SELECT DISTINCT name FROM gcd_series") if r[0]})

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = next(wb[n] for n in wb.sheetnames if (n == "Sheet X" or n.startswith("✅ Clean Inventory")))
    rows = list(ws.iter_rows(values_only=True))
    H = list(rows[0])
    ti, yi = H.index("Title"), H.index("Year")
    counts, years = defaultdict(int), {}
    for r in rows[1:]:
        t = str(r[ti] or "").strip()
        if t:
            counts[t] += 1
            years.setdefault(t, str(r[yi]))

    unmatched = [t for t in counts if tight(t) not in have_tight and tight(t) not in alias_t]
    print(f"GCD-unmatched inventory titles: {len(unmatched)}")

    out = openpyxl.Workbook()
    sh = out.active
    sh.title = "Fuzzy guesses"
    sh.append(["Current Title", "Rows", "Year", "Best GCD Guess", "Score", "Tier", "2nd Guess", "Score2"])
    from openpyxl.styles import Font, PatternFill
    for c in sh[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2F5496")
    sh.freeze_panes = "A2"

    n_hi = n_md = n_no = 0
    for t in sorted(unmatched, key=lambda x: -counts[x]):
        scored = sorted(((score(t, c), c) for c in cand_names), reverse=True)[:2]
        (s1, c1), (s2, c2) = (scored + [(0, ""), (0, "")])[:2]
        yr = years.get(t, "")
        # a newish year with a weak match = almost certainly not-in-corpus
        recent = bool(re.search(r"202[3-6]", yr))
        if s1 >= 0.80 and tight(c1) != tight(t):
            tier = "HIGH"; n_hi += 1
        elif s1 >= 0.62 and not recent:
            tier = "REVIEW"; n_md += 1
        else:
            tier = "no-corpus (try Comic Vine)"; n_no += 1
            c1 = c2 = ""; s1 = s2 = 0
        sh.append([t, counts[t], yr, c1, round(s1, 2) or "", tier, c2 if tier != "no-corpus (try Comic Vine)" else "", round(s2, 2) or ""])

    out.save(args.out)
    print(f"  HIGH-confidence guesses:   {n_hi}")
    print(f"  REVIEW (0.62-0.80):        {n_md}")
    print(f"  no-corpus (need CV):       {n_no}")
    print(f"Written: {args.out}")


if __name__ == "__main__":
    main()
