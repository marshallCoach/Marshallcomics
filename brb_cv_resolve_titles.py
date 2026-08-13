#!/usr/bin/env python3
"""
brb_cv_resolve_titles.py — propose correct titles for GCD-unmatched inventory
titles using Comic Vine's own volume identity as the oracle.

Two phases:
  Phase 1 (default, LOCAL, no API): mine volume_name already captured in
    covers.json from prior cover fetches — Comic Vine's authoritative title for
    that book, free. A sanity gate (shared significant tokens + difflib ratio)
    rejects CV mis-resolutions of short/ambiguous titles ("Die" -> "Pokémon Die
    Ersten Abenteuer").
  Phase 2 (--fetch, needs the api-server proxy on :5001 + COMIC_VINE_API_KEY):
    for titles with no cached volume_name, call /covers/search live and read
    match.volume_name. Roberto runs this — Code can't make credentialed calls.

Output: cv_title_guesses.xlsx — Current Title, CV proposed title, similarity,
source (cache/fetched), rows. Nothing is applied; confirm then batch-rename via
brb_merge_titles.py.

Usage:
    python3 brb_cv_resolve_titles.py            # phase 1, cache-mine only
    python3 brb_cv_resolve_titles.py --fetch    # + live CV for uncached (proxy up)
"""
import argparse, glob, json, os, re, difflib, time, urllib.request, urllib.parse
from collections import defaultdict
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
COVERS = os.path.join(ROOT, "covers.json")
PROXY = "http://localhost:5001/api/covers/search"
STOP = {"the", "a", "an", "of", "and", "or", "vs", "no", "1", "2", "3"}


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def tight(t):
    return re.sub(r"[^a-z0-9]", "", re.sub(r"&", " and ", str(t or "").lower()))


def norm(t):
    return re.sub(r"[^a-z0-9]+", " ", str(t or "").lower()).strip()


def toks(t):
    return {w for w in norm(t).split() if w not in STOP and len(w) > 1}


def sane(inv, cv):
    """Reject CV mis-resolutions: require shared significant tokens OR high ratio."""
    if not cv:
        return False, 0.0
    ta, tb = toks(inv), toks(cv)
    shared = ta & tb
    ratio = difflib.SequenceMatcher(None, norm(inv), norm(cv)).ratio()
    ok = (len(shared) >= 1 and (len(shared) / max(1, len(ta))) >= 0.5) or ratio >= 0.6
    return ok, round(ratio, 2)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=None)
    ap.add_argument("--fetch", action="store_true", help="live CV call for uncached titles (needs proxy + key)")
    ap.add_argument("--out", default="cv_title_guesses.xlsx")
    args = ap.parse_args()

    xlsx = os.path.abspath(args.xlsx) if args.xlsx else latest_xlsx()
    print(f"Source: {os.path.basename(xlsx)}")
    covers = json.load(open(COVERS))

    # cache: tight(title) -> best CV volume_name seen for it
    cv_by_title = {}
    for k, v in covers.items():
        if not isinstance(v, dict) or not v.get("volume_name"):
            continue
        t = k.split("|||")[0].strip()
        cv_by_title.setdefault(tight(t), []).append(v["volume_name"])

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = next(wb[n] for n in wb.sheetnames if (n == "Sheet X" or n.startswith("✅ Clean Inventory")))
    rows = list(ws.iter_rows(values_only=True))
    H = list(rows[0])
    ti, ii, yi, pi = (H.index(c) for c in ("Title", "Issue #", "Year", "Publisher"))
    # GCD-unmatched detection reuses the sqlite matched_title set
    import sqlite3
    conn = sqlite3.connect(os.path.join(ROOT, "gcd_local.sqlite"))
    have = {tight(r[0]) for r in conn.execute("SELECT DISTINCT matched_title FROM gcd_series")}
    try:
        have |= {tight(a) for a, in conn.execute("SELECT alias FROM gcd_title_alias")}
    except sqlite3.OperationalError:
        pass

    meta = {}
    counts = defaultdict(int)
    for r in rows[1:]:
        t = str(r[ti] or "").strip()
        if not t:
            continue
        counts[t] += 1
        meta.setdefault(t, (r[ii], r[yi], r[pi]))
    targets = [t for t in counts if tight(t) not in have]
    print(f"GCD-unmatched titles: {len(targets)}")

    out = openpyxl.Workbook(); sh = out.active; sh.title = "CV title guesses"
    sh.append(["Current Title", "Rows", "CV proposed title", "Similarity", "Source", "Accept? (y/n)"])
    from openpyxl.styles import Font, PatternFill
    for c in sh[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2F5496")
    sh.freeze_panes = "A2"
    for i, w in enumerate((38, 6, 38, 10, 10, 12), 1):
        sh.column_dimensions[chr(64 + i)].width = w

    from_cache = fetched = rejected = uncached = 0
    for t in sorted(targets, key=lambda x: -counts[x]):
        cvname = ""; src = ""
        cached = cv_by_title.get(tight(t))
        if cached:
            # pick the cached name most similar to the inventory title
            cvname = max(cached, key=lambda n: difflib.SequenceMatcher(None, norm(t), norm(n)).ratio())
            src = "cache"
        elif args.fetch:
            iss, yr, pub = meta[t]
            q = urllib.parse.urlencode({"title": t, "issue": str(iss or "1"),
                                        "year": str(yr or ""), "publisher": str(pub or "")})
            try:
                with urllib.request.urlopen(f"{PROXY}?{q}", timeout=30) as r:
                    d = json.load(r)
                cvname = ((d.get("match") or {}).get("volume_name")) or ""
                src = "fetched"; fetched += 1
                time.sleep(19)  # respect CV 200/hr
            except Exception as e:
                src = f"fetch-err"; cvname = ""
        else:
            uncached += 1
            continue

        ok, ratio = sane(t, cvname)
        if not ok or tight(cvname) == tight(t):
            rejected += 1
            continue
        if src == "cache":
            from_cache += 1
        sh.append([t, counts[t], cvname, ratio, src, ""])

    out.save(args.out)
    print(f"  proposals from cache:  {from_cache}")
    if args.fetch:
        print(f"  proposals from live CV: {fetched}")
    else:
        print(f"  titles with no cached CV name (need --fetch): {uncached}")
    print(f"  CV names rejected by sanity gate: {rejected}")
    print(f"Written: {args.out}")


if __name__ == "__main__":
    main()
