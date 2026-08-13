#!/usr/bin/env python3
"""
brb_merge_titles.py — collapse title spelling-variants into a canonical form.

Driven by a CSV map (variant,canonical) so the same tool serves CHECK 12's
punctuation splits, the near-identical pairs, and the GCD-unmatched misspelling
worklist. Touches ONLY the Title column, exact string match.

Safety (same discipline as the other fills):
  - dry-run by default; --apply writes a NEW *_TITLES_MERGED.xlsx, never the
    source in place.
  - Reports, per mapping, how many rows change.
  - WARNS when a merge makes two rows collide on Title+Issue+Year+Box (a
    same-box duplicate that was hidden by the spelling split) — these are real
    pre-existing dupes surfaced, not created; they're left for the validator.

Usage:
    python3 brb_merge_titles.py --map title_merge_map.csv            # dry-run
    python3 brb_merge_titles.py --map title_merge_map.csv --apply
    python3 brb_merge_titles.py <xlsx> --map title_merge_map.csv --apply --out <file>
"""
import argparse, csv, glob, os, sys
from collections import Counter, defaultdict
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def norm_issue(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=None)
    ap.add_argument("--map", required=True, help="CSV with columns: variant,canonical")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    xlsx = os.path.abspath(args.xlsx) if args.xlsx else latest_xlsx()
    print(f"Source: {os.path.basename(xlsx)}   (mode: {'APPLY' if args.apply else 'DRY-RUN'})")

    with open(args.map) as f:
        rmap = {r["variant"]: r["canonical"] for r in csv.DictReader(f) if r.get("variant") and r.get("canonical")}
    print(f"Mappings loaded: {len(rmap)}")

    wb = openpyxl.load_workbook(xlsx)
    ws = next(wb[n] for n in wb.sheetnames if (n == "Sheet X" or n.startswith("✅ Clean Inventory")))
    H = [c.value for c in next(ws.iter_rows(max_row=1))]
    ti, ii, yi, bi = (H.index(c) for c in ("Title", "Issue #", "Year", "Box #"))

    changed = Counter()
    # build post-merge key census to detect newly-colliding rows
    key_after = defaultdict(int)
    for row in ws.iter_rows(min_row=2):
        t = str(row[ti].value or "").strip()
        canon = rmap.get(t, t)
        key_after[(canon, norm_issue(row[ii].value), str(row[yi].value or "").strip(), str(row[bi].value or "").strip())] += 1

    for row in ws.iter_rows(min_row=2):
        t = str(row[ti].value or "").strip()
        if t in rmap and rmap[t] != t:
            changed[(t, rmap[t])] += 1
            if args.apply:
                row[ti].value = rmap[t]

    print(f"\n-- rows changed per mapping --")
    for (v, c), n in sorted(changed.items(), key=lambda x: -x[1]):
        print(f"  {n:>3}  '{v}'  ->  '{c}'")
    total = sum(changed.values())
    print(f"  TOTAL rows retitled: {total}")

    # collisions the merge creates/reveals (same Title+Issue+Year+Box, count>1,
    # where a canonical target is involved)
    canon_targets = set(rmap.values())
    collisions = {k: n for k, n in key_after.items() if n > 1 and k[0] in canon_targets}
    if collisions:
        print(f"\n⚠ {len(collisions)} same-box Title+Issue+Year+Box group(s) become duplicates after merge")
        print(f"  (real dupes previously hidden by the spelling split — validator will list them):")
        for (t, i, y, b), n in list(collisions.items())[:15]:
            print(f"    '{t}' #{i} {y} Box#{b} — {n}x")

    if not args.apply:
        print("\n[DRY RUN] No file written. Re-run with --apply.")
        return
    out = args.out or xlsx.replace(".xlsx", "_TITLES_MERGED.xlsx")
    wb.save(out)
    print(f"\nWritten: {out}")


if __name__ == "__main__":
    main()
