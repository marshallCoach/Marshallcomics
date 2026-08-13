#!/usr/bin/env python3
"""
brb_volume_audit.py — find rows whose VOLUME is wrong, using the year as anchor.

Roberto's insight (2707): an incorrect/blank cover is usually a VOLUME signal —
the row is labelled Vol 1 (or some wrong volume) when it's really a modern
volume, so the cover keys off the wrong run. Reused titles (Flash, Titans,
Trinity, Transformers, G.I. Joe, Captain America) are riddled with this: a 2024
book stamped "Vol 1" when Vol 1 ended decades ago.

Method (year-anchored, per "always check volume"):
  for each row of a target title, try "{Title} Vol {v} {issue}" for v in 1..MAXV
  on Marvel AND DC Fandom; read each page's release year; pick the volume whose
  year matches the row's declared year (+/-1). If that volume differs from the
  row's Volume, propose the correction. Only apply when the corrected
  (title,issue,box,vol,year) does not already exist (never create a dupe).

Fandom = free (no Comic Vine / eBay quota). READ-ONLY unless --apply; --apply
writes a NEW xlsx, never the source.

Usage:
    python3 brb_volume_audit.py --titles "The Flash,Titans,Trinity"
    python3 brb_volume_audit.py --titles "..." --apply --out new.xlsx
"""
import argparse, glob, json, os, re, time, urllib.parse, urllib.request, shutil
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
WIKIS = ["https://marvel.fandom.com/api.php", "https://dc.fandom.com/api.php"]
UA = "MarshallComicsInventory/1.0"
DELAY = 0.2
MAXV = 12


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def ni(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s); return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def nv(v):
    try:
        return str(int(float(str(v).strip())))
    except (ValueError, TypeError):
        return str(v or "").strip()


def yr(v):
    n = [int(t) for t in re.findall(r"\d{4}", str(v or "")) if 1900 < int(t) < 2100]
    return n[0] if n else None


_cache = {}


def page_year(page):
    if page in _cache:
        return _cache[page]
    out = None
    for API in WIKIS:
        try:
            u = (f"{API}?action=parse&page={urllib.parse.quote(page)}&prop=wikitext"
                 "&format=json&formatversion=2&redirects=1")
            with urllib.request.urlopen(urllib.request.Request(u, headers={"User-Agent": UA}), timeout=30) as r:
                d = json.load(r)
            wt = d.get("parse", {}).get("wikitext", "")
            time.sleep(DELAY)
            if not wt:
                continue
            m = re.search(r"\|\s*(?:ReleaseDate|CoverDate|Pubyear|Year)\s*=\s*([^\n|]+)", wt)
            y = yr(m.group(1)) if m else None
            if y:
                out = y; break
        except Exception:
            time.sleep(DELAY)
    _cache[page] = out
    return out


def correct_volume(title, issue, year):
    """Return the volume number whose Fandom page-year matches `year` (+/-1)."""
    if not year:
        return None
    best = None
    for v in range(1, MAXV + 1):
        py = page_year(f"{title} Vol {v} {issue}")
        if py is not None and abs(py - year) <= 1:
            best = v
            break
    return best


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--titles", required=True)
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--out", default=None)
    ap.add_argument("--review", default="volume_audit_review.xlsx")
    args = ap.parse_args()
    titles = {t.strip() for t in args.titles.split(",")}

    src = latest_xlsx()
    print(f"Source: {os.path.basename(src)}  titles: {sorted(titles)}", flush=True)
    if args.apply:
        out = args.out or src.replace(".xlsx", "_VOLAUDIT.xlsx")
        shutil.copy(src, out); wb = openpyxl.load_workbook(out)
    else:
        wb = openpyxl.load_workbook(src, data_only=True)
    ws = next(wb[n] for n in wb.sheetnames if (n == "Sheet X" or n.startswith("✅ Clean Inventory")))
    H = [c.value for c in next(ws.iter_rows(max_row=1))]
    C = {n: H.index(n) for n in H if n}
    ti, ii, yi, vi, bi = C["Title"], C["Issue #"], C["Year"], C["Volume"], C["Box #"]

    # existing keys to guard against creating dupes
    exist = set()
    for row in ws.iter_rows(min_row=2):
        exist.add((str(row[ti].value or "").strip(), ni(row[ii].value), str(row[bi].value or "").strip(),
                   nv(row[vi].value), yr(row[yi].value)))

    props, applied, skipped = [], 0, 0
    for row in ws.iter_rows(min_row=2):
        t = str(row[ti].value or "").strip()
        if t not in titles:
            continue
        dv, y, iss, box = nv(row[vi].value), yr(row[yi].value), ni(row[ii].value), str(row[bi].value or "").strip()
        cv = correct_volume(t, iss, y)
        if not cv or str(cv) == dv:
            continue
        target = (t, iss, box, str(cv), y)
        collide = target in exist
        props.append(dict(title=t, issue=iss, box=box, year=y, old=dv, new=cv, collide=collide))
        if args.apply and not collide:
            row[vi].value = cv; applied += 1
            exist.discard((t, iss, box, dv, y)); exist.add(target)
        elif collide:
            skipped += 1

    print(f"\n  volume corrections proposed: {len(props)}   (applied: {applied}, skipped-collide: {skipped})", flush=True)
    for p in sorted(props, key=lambda x: (x["title"], x["issue"]))[:40]:
        print(f"    {p['title'][:20]:<21}#{p['issue']:<5} yr={p['year']}  vol {p['old']} -> {p['new']}"
              f"{'  [SKIP collide]' if p['collide'] else ''}", flush=True)

    rb = openpyxl.Workbook(); sh = rb.active; sh.title = "Volume audit"
    sh.append(["Title", "Issue", "Box", "Year", "Vol now", "Vol correct (Fandom)", "Collides?", "Approve? (y/n)"])
    for c in sh[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2F5496")
    sh.freeze_panes = "A2"
    for p in sorted(props, key=lambda x: (x["title"], x["issue"])):
        sh.append([p["title"], p["issue"], p["box"], p["year"], p["old"], p["new"], "YES" if p["collide"] else "", ""])
    rb.save(args.review)
    print(f"  Review: {args.review}", flush=True)
    if args.apply:
        wb.save(out); print(f"  Written: {out}", flush=True)


if __name__ == "__main__":
    main()
