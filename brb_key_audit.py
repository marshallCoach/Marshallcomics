#!/usr/bin/env python3
"""
brb_key_audit.py — find keys HIDING in the data: rows that meet a mechanically
verifiable Tier 1 criterion from KEY_CLASSIFICATION.md but are flagged NO.

Why this matters beyond tidiness: an unflagged key is invisible to
brb_ebay_pricing.py's --key-tier queue, so it never gets a real market price.
That is the "Absolute Batman problem" — genuine value sitting unpriced because a
flag was missing.

Only DETERMINISTIC criteria are auto-detected (facts already in the sheet, not
judgment):
  Tier 1 #4  Signed? = YES            -> ALWAYS a key (spec rule 4, no exceptions)
  Tier 1 #5  CGC-graded / CGC-bound   -> key
  Tier 1 #1  '1st Appearances' filled -> key (the issue IS a 1st appearance)
  Tier 1 #6  celebrity/actor signature noted in Signed By

Judgment tiers (2 & 3 — acclaimed runs, anniversaries, spec value) are surfaced
as CANDIDATES ONLY, never auto-applied.

READ-ONLY. Per spec rule 2 key flags are never batch-changed: this writes a
review xlsx with row, current value, proposed value and the exact reason, for
a human to approve.

Usage:
    python3 brb_key_audit.py
    python3 brb_key_audit.py --out key_audit.xlsx
"""
import argparse, glob, os, re
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
ANNIV = {"100", "200", "300", "400", "500", "600", "700", "750", "800", "900", "1000"}


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def is_yes(v):
    return str(v).strip().upper() in ("YES", "Y", "TRUE", "1")


def filled(v):
    return v is not None and str(v).strip() not in ("", "nan", "None", "-", "N/A", "NO", "No")


def norm_issue(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=None)
    ap.add_argument("--out", default="key_audit.xlsx")
    args = ap.parse_args()

    xlsx = os.path.abspath(args.xlsx) if args.xlsx else latest_xlsx()
    print(f"Source: {os.path.basename(xlsx)}")

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("✅ Clean Inventory"))
    rows = list(ws.iter_rows(values_only=True))
    H = list(rows[0])
    C = {n: H.index(n) for n in H if n}

    def g(r, name):
        i = C.get(name)
        return r[i] if i is not None else None

    total = len(rows) - 1
    before = sum(1 for r in rows[1:] if is_yes(g(r, "Key Issue?")))

    hidden, candidates = [], []
    for idx, r in enumerate(rows[1:], 2):
        if is_yes(g(r, "Key Issue?")):
            continue
        title, issue = str(g(r, "Title") or "").strip(), norm_issue(g(r, "Issue #"))
        box = str(g(r, "Box #") or "")
        reasons = []
        if is_yes(g(r, "Signed?")):
            by = str(g(r, "Signed By") or "").strip()
            reasons.append(f"Tier 1 #4 — SIGNED{(' by ' + by) if by else ''}; spec rule 4: signed is always a key")
        if "CGC" in box.upper():
            reasons.append(f"Tier 1 #5 — CGC-bound (Box '{box}')")
        elif is_yes(g(r, "CGC Worth It?")):
            reasons.append("Tier 1 #5 — flagged CGC Worth It")
        # NOTE: '1st Appearances' is really a free-text notes column — it holds
        # "InHyuk Lee cover", "David Aja cover" etc. Variant covers are an EXPLICIT
        # disqualifier in the spec, so the column alone proves nothing. Only mine it
        # for two things it genuinely evidences: an unrecorded signature (Tier 1 #4 /
        # #6) or an explicit first appearance.
        note = str(g(r, "1st Appearances") or "").strip()
        nl = note.lower()
        if note:
            if re.search(r"\bsign(ed|ature)", nl):
                reasons.append(f"Tier 1 #4/#6 — signature noted in text but Signed? is not YES: {note[:60]}")
            elif re.search(r"\b1st\b|\bfirst appearance\b", nl):
                reasons.append(f"Tier 1 #1 — first appearance stated: {note[:60]}")
        rec = dict(row=idx, title=title, issue=issue, year=str(g(r, "Year") or ""),
                   box=box, signed="YES" if is_yes(g(r, "Signed?")) else "",
                   nm=g(r, "Est. Raw Value (NM) $"), ebay=g(r, "eBay Median Sold $"))
        if reasons:
            rec["why"] = " | ".join(reasons)
            hidden.append(rec)
        elif issue in ANNIV:
            rec["why"] = f"Tier 2 #7 CANDIDATE — landmark issue number #{issue} (needs human check: meaningful content?)"
            candidates.append(rec)

    print(f"\n  Rows: {total:,}")
    print(f"  Keys currently flagged YES: {before:,}")
    print(f"\n  HIDING KEYS (meet a deterministic Tier 1 test but flagged NO): {len(hidden)}")
    from collections import Counter
    kinds = Counter()
    for h in hidden:
        for part in h["why"].split(" | "):
            kinds[part.split(" — ")[0]] += 1
    for k, v in kinds.most_common():
        print(f"     {k}: {v}")
    print(f"  Tier 2/3 CANDIDATES (judgment needed, NOT auto-applied): {len(candidates)}")
    print(f"\n  Keys if all hiding ones were confirmed: {before + len(hidden):,}  (+{len(hidden)})")

    print("\n  sample hiding keys:")
    for h in hidden[:12]:
        print(f"     row{h['row']:<6} {h['title'][:32]:<33} #{h['issue']:<5} {h['why'][:64]}")

    out = openpyxl.Workbook()
    hdr = ["Sheet row", "Title", "Issue", "Year", "Box", "Signed", "NM $", "eBay median",
           "Key Issue? NOW", "PROPOSED", "Reason (tier + fact)", "Approve? (y/n)"]
    from openpyxl.styles import Font, PatternFill
    def build(sh, data, name, colour):
        sh.title = name
        sh.append(hdr)
        for c in sh[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=colour)
        sh.freeze_panes = "A2"
        for i, w in enumerate((10, 34, 8, 12, 10, 8, 9, 12, 14, 11, 66, 13), 1):
            sh.column_dimensions[chr(64 + i)].width = w
        for d in data:
            sh.append([d["row"], d["title"], d["issue"], d["year"], d["box"], d["signed"],
                       d["nm"], d["ebay"], "NO", "YES", d["why"], ""])
    build(out.active, hidden, "Hiding keys (Tier 1)", "C00000")
    build(out.create_sheet(), candidates, "Tier 2-3 candidates", "BF8F00")
    out.save(args.out)
    print(f"\n  Written: {args.out}  (nothing changed — approve in the sheet, per spec rule 2)")


if __name__ == "__main__":
    main()
