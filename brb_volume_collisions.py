#!/usr/bin/env python3
"""brb_volume_collisions.py — the wrong-era covers are a VOLUME data problem: two
different-era issues share one Title|||Issue|||Volume key (both default to Vol 1),
so they collide on a single cover and one always shows the wrong era. This finds
every such collision and proposes the correct Volume for each row by asking the
Fandom year-gate which volume's release year matches that row's year. READ-ONLY:
writes volume_collision_fix.xlsx for review — nothing is applied to the xlsx.

Apply flow: fill the Volume column in the xlsx from the Proposed Vol column
(confident rows), save a new timestamped canonical, then re-run
orchestrate_covers.sh so each era resolves its own correct-era cover.
"""
import glob, os, re, collections
import openpyxl
from openpyxl.styles import Font, PatternFill
import brb_cover_yeargate as R

ASSETS = "attached_assets"


def yr(v):
    n = [int(t) for t in re.findall(r"\d{4}", str(v or "")) if 1900 < int(t) < 2100]
    return n[0] if n else None


def nv(v):
    try:
        return str(int(float(str(v).strip())))
    except (ValueError, TypeError):
        return "1"


def ni(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s); return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


x = max(glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")), key=os.path.getmtime)
ws = next(w for w in openpyxl.load_workbook(x, read_only=True, data_only=True).worksheets if w.title.startswith("✅ Clean Inventory"))
rows = list(ws.iter_rows(values_only=True)); H = list(rows[0]); C = {n: H.index(n) for n in H if n}


def g(r, n):
    i = C.get(n); return r[i] if i is not None else None


# group rows by the cover key they collapse to
grp = collections.defaultdict(list)
for r in rows[1:]:
    t = str(g(r, "Title") or "").strip(); iss = ni(g(r, "Issue #")); vol = nv(g(r, "Volume")); y = yr(g(r, "Year"))
    if t and iss and y:
        grp[(t, iss, vol)].append((y, str(g(r, "Volume") or ""), str(g(r, "Box #") or ""), str(g(r, "Publisher") or "")))

# collisions = one key, multiple eras (year spread > 2)
coll = {k: v for k, v in grp.items() if len({round(y) for y, *_ in v}) > 1 and (max(y for y, *_ in v) - min(y for y, *_ in v)) > 2}
print(f"collision groups: {len(coll)} — resolving correct volume per era via Fandom year-gate", flush=True)

out = openpyxl.Workbook(); sh = out.active; sh.title = "Volume collisions"
sh.append(["Title", "Issue", "Box", "Year", "Publisher", "Current Vol", "Proposed Vol", "Source", "Reason"])
for c in sh[1]:
    c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="C00000")
sh.freeze_panes = "A2"
for i, w in enumerate((28, 7, 8, 6, 14, 11, 12, 16, 40), 1):
    sh.column_dimensions[chr(64 + i)].width = w

confident = manual = 0
for (t, iss, curvol), members in sorted(coll.items()):
    for (y, rawvol, box, pub) in members:
        res = R.resolve(t, iss, y, curvol)   # (url, py, wiki, found_vol) or None
        if res:
            _, py, wiki, fv = res
            proposed = fv; src = f"fandom-{wiki} {py}"
            reason = f"era {y} matches {t} Vol {fv} ({py}); shares key with other eras {sorted({yy for yy,*_ in members})}"
            if str(fv) != nv(rawvol):
                confident += 1
                sh.append([t, iss, box, y, pub, rawvol or "(blank)", proposed, src, reason])
                cell = sh.cell(row=sh.max_row, column=7); cell.font = Font(bold=True, color="006100")
        else:
            manual += 1
            sh.append([t, iss, box, y, pub, rawvol or "(blank)", "?", "not on Marvel/DC wiki", f"needs manual volume (eras {sorted({yy for yy,*_ in members})})"])
            sh.cell(row=sh.max_row, column=7).fill = PatternFill("solid", fgColor="FFF2CC")
    if (confident + manual) % 40 == 0:
        print(f"  processed ~{confident+manual} rows...", flush=True)

out.save("volume_collision_fix.xlsx")
print(f"\nProposed volume changes (confident): {confident}", flush=True)
print(f"Rows needing manual volume (indie/off-wiki): {manual}", flush=True)
print("Written: volume_collision_fix.xlsx", flush=True)
