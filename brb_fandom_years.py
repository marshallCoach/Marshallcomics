#!/usr/bin/env python3
"""
brb_fandom_years.py — resolve per-issue publication years from the Marvel and DC
Fandom wikis, for the rows GCD could not settle.

Why this exists: brb_gcd_year_fix.py resolves a Year RANGE to an exact year when
GCD has the issue. ~299 rows remain — 202 where the issue number appears in two
volumes inside the declared range (ambiguous), and ~97 GCD simply lacks. Fandom
covers both, and its MediaWiki API is public (no key) so it does NOT consume the
Comic Vine quota — safe to run alongside a cover fetch.

Method: build the wiki page title from the row's own fields —
    "{Title} Vol {Volume} {Issue}"     (both wikis use this convention)
then read the year out of the infobox. The two wikis use different fields:
    Marvel -> ReleaseDate / CoverDate
    DC     -> Pubyear / Year
Redirects are followed (Marvel files Uncanny X-Men #141 under "X-Men Vol 1 141").
A proposal is only kept when the year is inside the declared range (+/-1yr), so a
wrong page can't silently rewrite a good row.

Read-only: writes a review xlsx. Nothing is applied.

⚠ CALIBRATION (Roberto, 2307): only trust rows flagged "in-range". The volume
sweep below WILL find a same-numbered issue in some other era when the right
volume has no such issue — e.g. X-Men #1-4 matched 1963, Green Lantern matched
1941, JLA #100-120 matched 1972. Those books are all MODERN; the sweep simply
landed on the wrong volume. Out-of-range hits are therefore BAD MATCHES, not
data findings, and must not be applied. The in-range gate is what makes this
safe: the wiki year must corroborate the row's own declared range before it is
used, so a wrong-era page can never pass.

Usage:
    python3 brb_fandom_years.py              # all unresolved rows
    python3 brb_fandom_years.py --limit 25   # sample first
"""
import argparse, glob, json, os, re, time, urllib.parse, urllib.request
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(ROOT, "attached_assets")
UA = "MarshallComicsInventory/1.0 (personal collection catalogue)"
DELAY = 1.0          # be polite to Fandom
TOLERANCE = 1
RANGE_RE = re.compile(r"\d{4}\s*[-–—]\s*\d{4}")
FIELDS = ["ReleaseDate", "CoverDate", "Pubyear", "Year", "Released", "Cover Date"]


def latest_xlsx():
    m = [f for f in glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")) if not os.path.basename(f).startswith("~$")]
    return max(m, key=os.path.getmtime) if m else None


def norm_issue(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def year_span(v):
    n = [int(x) for x in re.findall(r"\d{4}", str(v or "")) if 1900 < int(x) < 2100]
    return (min(n), max(n)) if n else None


def wiki_host(publisher):
    p = str(publisher or "").lower()
    if "marvel" in p:
        return "marvel"
    if p.startswith("dc") or p == "dc":
        return "dc"
    return None


def fetch_page(host, title):
    url = (f"https://{host}.fandom.com/api.php?action=query&prop=revisions&rvprop=content"
           f"&rvslots=main&titles={urllib.parse.quote(title)}&format=json&formatversion=2&redirects=1")
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=25) as r:
        d = json.load(r)
    p = (d.get("query", {}).get("pages") or [{}])[0]
    if p.get("missing") or "revisions" not in p:
        return None
    return p["revisions"][0]["slots"]["main"]["content"]


def year_from(content):
    for k in FIELDS:
        m = re.search(rf"\|\s*{re.escape(k)}\s*=\s*([^\n|]+)", content, re.I)
        if m:
            y = re.search(r"(19|20)\d{2}", m.group(1))
            if y:
                return int(y.group(0)), k
    return None, None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=None)
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--out", default="fandom_year_proposals.xlsx")
    args = ap.parse_args()

    xlsx = os.path.abspath(args.xlsx) if args.xlsx else latest_xlsx()
    print(f"Source: {os.path.basename(xlsx)}")

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = next(wb[n] for n in wb.sheetnames if n.startswith("✅ Clean Inventory"))
    rows = list(ws.iter_rows(values_only=True))
    H = list(rows[0])
    ti, ii, yi, vi, pi, bi = (H.index(c) for c in ("Title", "Issue #", "Year", "Volume", "Publisher", "Box #"))

    # target: rows whose Year is still a range (GCD couldn't settle them)
    targets = []
    seen = set()
    for r in rows[1:]:
        if not RANGE_RE.search(str(r[yi] or "")):
            continue
        host = wiki_host(r[pi])
        if not host:
            continue
        key = (str(r[ti]).strip(), norm_issue(r[ii]), str(r[vi]))
        if key in seen:
            continue
        seen.add(key)
        targets.append((str(r[ti]).strip(), norm_issue(r[ii]), str(r[yi]), r[vi], host, str(r[bi])))
    if args.limit:
        targets = targets[: args.limit]
    print(f"Rows to query on Fandom: {len(targets)}  (~{len(targets) * DELAY / 60:.0f} min)")

    out = openpyxl.Workbook(); sh = out.active; sh.title = "Fandom year proposals"
    sh.append(["Title", "Issue", "Declared Year", "Volume", "Wiki", "Wiki page", "Proposed Year", "Field", "In range?", "Box"])
    from openpyxl.styles import Font, PatternFill
    for c in sh[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2F5496")
    sh.freeze_panes = "A2"
    for i, w in enumerate((34, 7, 14, 8, 8, 34, 13, 12, 10, 8), 1):
        sh.column_dimensions[chr(64 + i)].width = w

    ok = miss = out_of_range = 0
    for n, (title, issue, declared, vol, host, box) in enumerate(targets, 1):
        try:
            v = int(float(vol)) if vol not in (None, "") else 1
        except (ValueError, TypeError):
            v = 1
        # Sweep volumes: the declared one first, then 1..6. The inventory's Volume
        # is itself sometimes wrong (JLA #100 is filed here as Vol 2/2010 but is
        # really Vol 1/1972), so trusting it alone loses the answer entirely.
        y = fld = None
        page = ""
        for cand in [v] + [x for x in range(1, 7) if x != v]:
            trial = f"{title} Vol {cand} {issue}"
            try:
                content = fetch_page(host, trial)
            except Exception:
                content = None
            if content:
                yy, ff = year_from(content)
                if yy:
                    y, fld, page, v = yy, ff, trial, cand
                    break
            time.sleep(DELAY)
        if not page:
            page = f"{title} Vol {v} {issue}"
        span = year_span(declared)
        inrange = bool(y and span and span[0] - TOLERANCE <= y <= span[1] + TOLERANCE)
        if y and inrange:
            ok += 1
        elif y:
            out_of_range += 1
        else:
            miss += 1
        sh.append([title, issue, declared, v, host, page, y or "", fld or "",
                   "in-range" if inrange else ("YEAR+VOL LIKELY WRONG" if y else "no page"), box])
        if n % 25 == 0:
            print(f"  [{n}/{len(targets)}] resolved={ok} out-of-range={out_of_range} missing={miss}")
        time.sleep(DELAY)

    out.save(args.out)
    print(f"\n  resolved & in declared range: {ok}")
    print(f"  found but OUT of range (suspect page/volume): {out_of_range}")
    print(f"  no page / no date: {miss}")
    print(f"Written: {args.out}")


if __name__ == "__main__":
    main()
