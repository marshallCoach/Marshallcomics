#!/usr/bin/env python3
"""brb_vision_enrich.py — add Publisher + Cover Artist (from the inventory) to the
vision character output. Run AFTER brb_vision_characters.py finishes.

Joins the vision output to the ✅ Clean Inventory sheet on Title + Issue. Where a
title+issue maps to more than one inventory row (different volumes/eras), the most
common non-blank value is used and true conflicts are joined with ' | ' so nothing
is silently dropped. Writes vision_characters_enriched.xlsx with columns ordered:
Title, Issue, Publisher, Cover Artist, Visually Verified Characters, Confidence, URL.

Usage: python3 brb_vision_enrich.py [vision_output.xlsx]
"""
import sys, glob, os, collections
import openpyxl

ASSETS = "attached_assets"


def ni(v):
    s = str(v).strip().lstrip("#")
    try:
        f = float(s); return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def pick(counter):
    """Most common non-blank value; join genuine multiples with ' | '."""
    vals = [v for v, _ in counter.most_common() if v and str(v).strip()]
    if not vals:
        return ""
    if len(vals) == 1:
        return vals[0]
    top = counter.most_common(1)[0][1]
    lead = [v for v in vals if counter[v] == top]
    return lead[0] if len(lead) == 1 else " | ".join(dict.fromkeys(vals))


def main():
    vpath = sys.argv[1] if len(sys.argv) > 1 else "vision_workqueue_visual.xlsx"
    if not os.path.exists(vpath):
        print(f"no vision output at {vpath} — run brb_vision_characters.py first"); sys.exit(2)

    inv = max(glob.glob(os.path.join(ASSETS, "comics_inventory_*.xlsx")), key=os.path.getmtime)
    ws = next(w for w in openpyxl.load_workbook(inv, read_only=True, data_only=True).worksheets if (w.title == "Sheet X" or w.title.startswith("✅ Clean Inventory")))
    rows = list(ws.iter_rows(values_only=True)); H = list(rows[0]); C = {n: H.index(n) for n in H if n}

    def g(r, n):
        i = C.get(n); return r[i] if i is not None else None

    pubmap = collections.defaultdict(collections.Counter)
    artmap = collections.defaultdict(collections.Counter)
    for r in rows[1:]:
        t = str(g(r, "Title") or "").strip().lower(); iss = ni(g(r, "Issue #"))
        if not t:
            continue
        k = (t, iss)
        pubmap[k][str(g(r, "Publisher") or "").strip()] += 1
        artmap[k][str(g(r, "Cover Artist") or "").strip()] += 1

    wb = openpyxl.load_workbook(vpath, read_only=True, data_only=True)
    vs = wb.active
    vrows = list(vs.iter_rows(values_only=True)); VH = list(vrows[0]); VC = {str(n): i for i, n in enumerate(VH) if n}
    ti = VC.get("title", 0); ii = VC.get("issue", 1)
    ui = VC.get("url"); ci = VC.get("Visually Verified Characters"); fi = VC.get("Confidence")

    out = openpyxl.Workbook(); sh = out.active; sh.title = "Characters + creators"
    from openpyxl.styles import Font, PatternFill
    sh.append(["Title", "Issue", "Publisher", "Cover Artist", "Visually Verified Characters", "Confidence", "Cover URL"])
    for c in sh[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2F5496")
    sh.freeze_panes = "A2"
    for i, w in enumerate((30, 7, 16, 26, 46, 11, 60), 1):
        sh.column_dimensions[chr(64 + i)].width = w

    matched = 0
    for r in vrows[1:]:
        t = str(r[ti]); iss = ni(r[ii]); k = (t.strip().lower(), iss)
        pub = pick(pubmap[k]) if k in pubmap else ""
        art = pick(artmap[k]) if k in artmap else ""
        if pub or art:
            matched += 1
        sh.append([t, r[ii],
                   pub, art,
                   r[ci] if ci is not None else "",
                   r[fi] if fi is not None else "",
                   r[ui] if ui is not None else ""])

    out.save("vision_characters_enriched.xlsx")
    print(f"rows: {len(vrows)-1}  matched a Publisher/Cover Artist: {matched}")
    print("Written: vision_characters_enriched.xlsx")


if __name__ == "__main__":
    main()
